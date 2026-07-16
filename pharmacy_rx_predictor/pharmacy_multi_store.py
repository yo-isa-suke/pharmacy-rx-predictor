# -*- coding: utf-8 -*-
"""
処方箋予測 — 複数店舗（A/B/C…）比較ツール
====================================================
とあるスーパーの複数の出店候補地（A点・B点・C点…）をまとめて分析し、
「医療機関ベース（ハフ按分）」と「集客ベース（来店客数）」の2トラックで比較する。

- 既存の「260702_Prescription Analysis_v2.py」のモデル/スクレイパーをそのまま再利用
  （既存ファイルは一切変更しない。UIを起動せずに関数だけ読み込む）。
- ブラウザ上で、候補地ごとに①ハフの取り分内訳・②集客の内訳・面/門前一覧（目視修正可）・
  医療機関/薬局リストを表示。
- 「数式入りExcel(.xlsx)」で書き出せる。ユニーク客数・競合・係数を編集すると集客の予測が、
  原資・重みを編集すると医療機関(ハフ)の獲得が、Excel上で自動で再計算される。
"""
import io
import json
import math
import os
import re
from dataclasses import replace

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="複数店舗比較 — 処方箋予測", page_icon="🏪", layout="wide")

CORE_FILE = os.path.join(os.path.dirname(__file__), "260702_Prescription Analysis_v2.py")


# ── 既存ファイルからモデル/スクレイパーだけを読み込む（UIは起動しない） ──────────
@st.cache_resource(show_spinner="モデルを読み込み中…")
def load_core():
    src = open(CORE_FILE, encoding="utf-8").read()
    src = re.sub(r"st\.set_page_config\(.*?\)", "", src, count=1, flags=re.DOTALL)
    for marker in ('\nst.markdown("""\n<style>', "\nst.title("):
        idx = src.find(marker)
        if idx != -1:
            src = src[:idx]
            break
    ns = {"__name__": "rx_core"}
    exec(compile(src, CORE_FILE, "exec"), ns)
    return ns


M = load_core()
MHLWScraper = M["MHLWScraper"]
run_analysis = M["run_analysis"]
PredictionAssumptions = M["PredictionAssumptions"]
HuffParams = M["HuffParams"]
FootfallParams = M["FootfallParams"]
compute_footfall_prediction = M["compute_footfall_prediction"]
classify_menkata = M["classify_menkata"]
footfall_competitor_power = M["footfall_competitor_power"]
haversine = M["haversine"]
_clinic_annual_rx_pool = M["_clinic_annual_rx_pool"]
_pharmacy_attractiveness = M["_pharmacy_attractiveness"]
facility_key = M["facility_key"]
pharmacy_key = M["pharmacy_key"]
MedFacility = M["MedFacility"]
PharmacyFacility = M["PharmacyFacility"]


def _num(v):
    """pandas由来のNaN/None/空 を None に、数値は float にする。"""
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return None if f != f else f  # NaN 除外


def _reposition(clat, clon, flat, flon, target_dist):
    """候補地→現在地の向きを保ったまま、距離だけ target_dist(m) に補正した座標を返す。"""
    cur = haversine(clat, clon, flat, flon)
    if cur <= 0 or target_dist <= 0:
        return flat, flon
    ratio = target_dist / cur
    return clat + (flat - clat) * ratio, clon + (flon - clon) * ratio


def _recs_sig(recs, numfield):
    """編集レコードの正規化シグネチャ（float揺れによる無限再実行を防ぐ比較用）。"""
    out = []
    for r in recs:
        la, lo, nv = _num(r.get("lat")), _num(r.get("lon")), _num(r.get(numfield))
        out.append(((r.get("name") or "").strip(),
                    round(la, 6) if la is not None else None,
                    round(lo, 6) if lo is not None else None,
                    round(nv, 1) if nv is not None else None))
    return out


_EXT_CATS = ["院外のみ", "院内外どちらも", "院内のみ", "不明"]


def rx_category(fac):
    """ナビィの院内/院外処方フィールドから、院外のみ/院内外どちらも/院内のみ/不明を判定。"""
    ih = (getattr(fac, "inhouse_rx", "") or "")
    op = (getattr(fac, "outpatient_rx", "") or "")
    has_in, has_out = ("有" in ih), ("有" in op)
    if has_out and has_in:
        return "院内外どちらも"
    if has_out and not has_in:
        return "院外のみ"
    if has_in and not has_out:
        return "院内のみ"
    s = getattr(fac, "rx_summary", "") or ""
    if s.startswith("院外処方あり"):
        return "院内外どちらも"
    if s == "院内処方のみ":
        return "院内のみ"
    return "不明"


def ext_coef(cat, ext_rate):
    """院外区分→院外係数。院外のみ=1.0／院内外どちらも・不明=院外率／院内のみ=0。"""
    if cat == "院外のみ":
        return 1.0
    if cat == "院内のみ":
        return 0.0
    return ext_rate


def clinic_pool(fac, cat, a, issue_rate, ext_rate, unknown_op):
    """クリニックの年間院外処方（原資）＝外来×診療日数×発行率×院外係数（美容/歯科は減係数）。"""
    op = fac.daily_outpatients
    if not op:
        op = unknown_op  # 外来不明→既定値（下目）
    days = (fac.weekly_op_days * 52.0
            if (a.annual_days_mode == "weekly" and fac.weekly_op_days) else float(a.fixed_annual_days))
    coef = ext_coef(cat, ext_rate)
    if getattr(fac, "is_cosmetic", False):
        coef *= a.cosmetic_factor
    elif getattr(fac, "facility_category", "") == "歯科診療所":
        coef *= a.dental_factor
    return op * days * issue_rate * coef


def clinic_flag(fac, high_thr):
    """外れ値・不明のアラート文字列（空=正常）。"""
    flags = []
    op = fac.daily_outpatients
    if not op:
        flags.append("外来不明→既定値使用")
    elif getattr(fac, "facility_category", "") != "病院" and op >= high_thr:
        flags.append(f"要確認：{int(op)}人/日は多め")
    cf = getattr(fac, "op_flag", "") or ""
    if cf:
        flags.append(cf)
    return " / ".join(flags)


def resolve_edit(ed, name_col, num_disp, num_store, stored, clat, clon, cat_col=None):
    """
    編集後の data_editor 内容を、保存用レコードに解決する。行ごとに：
      ・緯度/経度を編集した → その座標を採用
      ・距離(m)だけ編集した → 元座標の向きを保って距離を補正
      ・新規行 → 緯度経度があれば採用、無く距離だけなら候補地の真北に仮置き
    """
    smap = {s.get("_key"): s for s in stored if isinstance(s.get("_key"), str) and s.get("_key")}
    out = []
    for _, row in ed.iterrows():
        name = (str(row.get(name_col)) if row.get(name_col) is not None else "").strip()
        lat, lon = _num(row.get("緯度")), _num(row.get("経度"))
        dist, nv = _num(row.get("距離(m)")), _num(row.get(num_disp))
        key = row.get("_key")
        key = key if (isinstance(key, str) and key) else None
        s = smap.get(key)
        if s is not None:
            slat, slon = _num(s.get("lat")), _num(s.get("lon"))
            sdist = (haversine(clat, clon, slat, slon)
                     if (slat is not None and slon is not None) else None)
            coords_edited = (slat is not None and lat is not None and lon is not None
                             and (abs(lat - slat) > 1e-9 or abs(lon - slon) > 1e-9))
            if coords_edited:
                pass  # 編集した緯度経度を採用
            elif dist is not None and sdist is not None and abs(dist - sdist) > 1:
                lat, lon = _reposition(clat, clon, slat, slon, dist)  # 距離だけ補正
            else:
                lat, lon = slat, slon  # 変更なし
        else:
            if (lat is None or lon is None) and dist is not None and dist > 0:
                lat, lon = clat + dist / 111000.0, clon  # 新規・距離のみ→真北に仮置き
        if name or (lat is not None and lon is not None):
            rec = {"name": name, "lat": lat, "lon": lon, num_store: nv, "_key": key}
            if cat_col:
                cv = row.get(cat_col)
                rec["cat"] = cv if cv in _EXT_CATS else "不明"
            out.append(rec)
    return out


def effective_facilities(raw, clat, clon, label):
    """
    生データ＋候補地ごとの編集（座標補正・削除・手動追加）を反映した実効の医療機関/薬局リストを返す。
    編集内容は session_state['med_edit'][label] / ['ph_edit'][label] に行レコードで保持。
    """
    med_edit = st.session_state.setdefault("med_edit", {})
    ph_edit = st.session_state.setdefault("ph_edit", {})
    if label not in med_edit:
        med_edit[label] = [{"name": f.name, "lat": f.lat, "lon": f.lon,
                            "op": f.daily_outpatients, "cat": rx_category(f),
                            "_key": facility_key(f)} for f in raw["med"]]
    if label not in ph_edit:
        ph_edit[label] = [{"name": p.name, "lat": p.lat, "lon": p.lon,
                           "rx": p.annual_rx_count, "_key": pharmacy_key(p)} for p in raw["ph"]]
    raw_med_map = {facility_key(f): f for f in raw["med"]}
    raw_ph_map = {pharmacy_key(p): p for p in raw["ph"]}

    med_eff = []
    for r in med_edit[label]:
        name = (r.get("name") or "").strip()
        lat, lon = _num(r.get("lat")), _num(r.get("lon"))
        if not name or lat is None or lon is None:
            continue
        op = _num(r.get("op"))
        base = raw_med_map.get(r.get("_key"))
        if base is not None:
            f = replace(base, lat=lat, lon=lon,
                        daily_outpatients=(int(op) if op else base.daily_outpatients))
        else:
            f = MedFacility(name=name, lat=lat, lon=lon,
                            daily_outpatients=(int(op) if op else None),
                            rx_summary="院外処方あり", facility_category="診療所",
                            source="手動追加")
        cat = r.get("cat")
        f.rx_cat = cat if cat in _EXT_CATS else rx_category(f)
        f.distance_m = haversine(clat, clon, lat, lon)
        med_eff.append(f)

    ph_eff = []
    for r in ph_edit[label]:
        name = (r.get("name") or "").strip()
        lat, lon = _num(r.get("lat")), _num(r.get("lon"))
        if not name or lat is None or lon is None:
            continue
        rx = _num(r.get("rx"))
        base = raw_ph_map.get(r.get("_key"))
        if base is not None:
            p = replace(base, lat=lat, lon=lon,
                        annual_rx_count=(int(rx) if rx else base.annual_rx_count))
        else:
            p = PharmacyFacility(name=name, address="", lat=lat, lon=lon,
                                 annual_rx_count=(int(rx) if rx else None), source="手動追加")
        p.distance_m = haversine(clat, clon, lat, lon)
        ph_eff.append(p)
    return med_eff, ph_eff


@st.cache_resource
def get_scraper():
    return MHLWScraper()


# ════════════════════════════════ サイドバー ════════════════════════════════
with st.sidebar:
    st.header("共通設定")
    radius_m = st.slider("商圏半径 (m)", 500, 5000, 3000, 100,
                         help="スーパー商圏に準拠。全候補地に共通で使います。")
    fetch_all_ph = st.checkbox(
        "商圏内の薬局を全件取得（重み付けを最も正確に・ただし遅い）", value=True,
        help="ON：商圏内の全薬局の詳細（正確な座標・実績）を取得し、門前判定・実績を最も正確にします。"
             "OFF：近い順に指定件数だけ詳細取得（遠い競合も座標は取得済みで按分には入ります）。",
    )
    if fetch_all_ph:
        max_detail = 9999
        st.caption("→ 商圏内の薬局を全件取得します（速度優先にしたい場合はチェックを外すと件数指定が出ます）。")
    else:
        max_detail = int(st.slider("詳細取得件数（薬局）", 5, 120, 30, 5,
                                   help="ナビィから実績・座標を取る薬局の上限（多いほど正確・遅い）。"))
    gate_m = 50

    st.divider()
    st.subheader("🛒 集客ベースの前提（全候補地に共通）")
    ff_r65 = st.number_input("65歳以上の比率（商圏の高齢化率）", 0.0, 1.0, 0.30, 0.01,
                             help="来店客の年齢構成。会員の年齢データがあればその比率、なければ商圏の高齢化率。")
    c1, c2 = st.columns(2)
    ff_v65 = c1.number_input("65+ の月受診回数", 0.0, 6.0, 3.0, 0.1)
    ff_vu65 = c2.number_input("65- の月受診回数", 0.0, 6.0, 1.3, 0.1)
    c3, c4, c5 = st.columns(3)
    ff_issue = c3.number_input("発行率", 0.0, 1.0, 0.8054, 0.0001, format="%.4f")
    ff_ext = c4.number_input("院外率", 0.0, 1.0, 0.8313, 0.0001, format="%.4f")
    ff_use = c5.number_input("利用率", 0.0, 1.0, 0.137, 0.001, format="%.3f")
    c6, c7 = st.columns(2)
    ff_monzen = c6.number_input("門前しきい値(m)", 0, 300, 50, 10,
                                help="最寄りクリニックがこの距離以内の薬局は門前として自動判定→面競合から除外。")
    ff_decay = c7.number_input("面競合の距離減衰λ(m)", 0, 3000, 1000, 100,
                               help="遠い面競合を弱く数える。小さいほど自店シェア↑。")
    ff_main = st.number_input(
        "メイン薬局しきい値(枚/年・0=無効/既定)", 0, 100000, 0, 1000,
        help="既定0＝無効。実績が大きい面薬局は“面の強豪”なので、除外せずパワー加重で強い競合として"
             "数えます（門前かどうかは距離＝門前しきい値で判定）。門前で大量の店だけ外したい場合のみ値を入れます。",
    )

    with st.expander("⚙️ 詳細設定（ハフ按分・通常は変更不要）", expanded=False):
        huff_lambda = st.slider(
            "距離減衰 λ (m)", 150, 1200, 300, 50,
            help="発行率×院外係数を入れた新原資に合わせ、地方69シード・面型106店で再較正した値"
                 "（予測/実績 中央値0.99）。",
        )
        huff_boost = st.slider("門前ブースト", 1.0, 15.0, 6.0, 0.5)
        huff_monzen_r = st.slider(
            "門前ブースト半径 (m)", 30, 150, 50, 10,
            help="①医療機関ベースで、クリニックからこの距離以内の薬局に門前ブーストを掛けます。"
                 "実質“門前”が80〜100mにある場合はここを広げてください（既定50mは275店検証時の値のため、"
                 "広げた場合は再較正が望ましい）。",
        )
        huff_candA = st.number_input("候補店の引力（大型店は上げる）", 0.2, 10.0, 1.0, 0.1)
        st.caption("── 医療機関ベースの原資（外来→処方箋の換算）──")
        med_unknown_op = st.number_input(
            "外来不明クリニックの既定外来数(人/日)", 0, 500, 30, 5,
            help="ナビィで外来患者数が取得できないクリニックに入れる値（未入力なので下目に設定）。",
        )
        med_high_thr = st.number_input(
            "外れ値アラートしきい値（診療所・人/日）", 50, 2000, 200, 10,
            help="診療所でこの人数以上なら『要確認』を表示（月間値・誤登録の疑い）。",
        )
        st.caption(f"※ 原資 = 外来×診療日数×発行率{float(ff_issue):.4f}×院外係数"
                   "（院外のみ1.0／院内外どちらも=院外率/院内のみ0）。発行率・院外率は集客ベースと同値。")

    st.caption("※ サイドバーや面/門前を変えると、再検索なしで比較表・Excelが即更新されます。")


# ── 保存／再読込（検索結果＋手修正をローカルに保存し、再起動しても復元） ────────────
SAVE_FILE = os.path.join(os.path.dirname(__file__), "rx_session.json")
_FAC_FIELDS = ["name", "address", "lat", "lon", "daily_outpatients", "weekly_op_days",
               "rx_summary", "inhouse_rx", "outpatient_rx", "is_cosmetic",
               "facility_category", "kikan_cd", "op_flag"]
_PH_FIELDS = ["name", "address", "lat", "lon", "annual_rx_count", "kikan_cd"]


def _mk_med(d):
    return MedFacility(
        name=d.get("name") or "", address=d.get("address") or "",
        lat=d.get("lat"), lon=d.get("lon"),
        daily_outpatients=d.get("daily_outpatients"), weekly_op_days=d.get("weekly_op_days"),
        rx_summary=d.get("rx_summary") or "不明", inhouse_rx=d.get("inhouse_rx") or "—",
        outpatient_rx=d.get("outpatient_rx") or "—", is_cosmetic=bool(d.get("is_cosmetic")),
        facility_category=d.get("facility_category") or "診療所",
        kikan_cd=d.get("kikan_cd") or "", op_flag=d.get("op_flag") or "")


def _mk_ph(d):
    return PharmacyFacility(
        name=d.get("name") or "", address=d.get("address") or "",
        lat=d.get("lat"), lon=d.get("lon"),
        annual_rx_count=d.get("annual_rx_count"), kikan_cd=d.get("kikan_cd") or "")


def _build_state():
    raws = st.session_state.get("multi_raw", [])
    return {
        "candidates": [{
            "label": r["label"], "name": r["name"], "addr": r["addr"],
            "uni": r["uni"], "exposure": r.get("exposure", 1.0),
            "clat": r["clat"], "clon": r["clon"],
            "med": [{k: getattr(f, k, None) for k in _FAC_FIELDS} for f in r["med"]],
            "ph": [{k: getattr(p, k, None) for k in _PH_FIELDS} for p in r["ph"]],
        } for r in raws],
        "med_edit": st.session_state.get("med_edit", {}),
        "ph_edit": st.session_state.get("ph_edit", {}),
        "mk_multi": st.session_state.get("mk_multi", {}),
        "exp_multi": st.session_state.get("exp_multi", {}),
    }


def _apply_state(state):
    raws = []
    for c in state.get("candidates", []):
        raws.append({"label": c["label"], "name": c["name"], "addr": c["addr"],
                     "uni": c["uni"], "exposure": c.get("exposure", 1.0),
                     "clat": c["clat"], "clon": c["clon"],
                     "med": [_mk_med(d) for d in c["med"]],
                     "ph": [_mk_ph(d) for d in c["ph"]]})
    st.session_state["multi_raw"] = raws
    for k in ("med_edit", "ph_edit", "mk_multi", "exp_multi"):
        st.session_state[k] = state.get(k, {})


def state_bytes():
    return json.dumps(_build_state(), ensure_ascii=False).encode("utf-8")


def save_session():
    with open(SAVE_FILE, "w", encoding="utf-8") as fh:
        json.dump(_build_state(), fh, ensure_ascii=False)


def load_session():
    with open(SAVE_FILE, encoding="utf-8") as fh:
        _apply_state(json.load(fh))


def make_fp(uni):
    return FootfallParams(
        enabled=(uni > 0),
        unique_customers_monthly=float(uni), ratio_65plus=float(ff_r65),
        visits_month_65plus=float(ff_v65), visits_month_under65=float(ff_vu65),
        issue_rate=float(ff_issue), external_rate=float(ff_ext), use_rate=float(ff_use),
        menkata_monzen_dist=float(ff_monzen), menkata_main_rx=float(ff_main),
        competitor_decay_m=float(ff_decay),
    )


def make_hp():
    return HuffParams(lambda_m=float(huff_lambda), monzen_boost=float(huff_boost),
                      candidate_attractiveness=float(huff_candA), monzen_radius=float(huff_monzen_r))


# ── ハフの取り分内訳（クリニック1行ずつ・自店の重み/競合の重み合計を明示） ─────────
def huff_breakdown(med, ph, clat, clon, hp, a, issue_rate, ext_rate, unknown_op):
    comps = []
    for p in ph:
        if p.lat is None or p.lon is None:
            continue
        ak = _pharmacy_attractiveness(p, hp.national_avg_rx) if hp.weight_by_power else 1.0
        comps.append((p.lat, p.lon, ak))

    def bw(d, aa):
        v = math.exp(-d / hp.lambda_m)
        if d <= hp.monzen_radius:
            v *= hp.monzen_boost
        return aa * v

    rows = []
    for f in med:
        if f.lat is None or f.lon is None or not getattr(f, "in_area", True):
            continue
        d_self = haversine(clat, clon, f.lat, f.lon)
        if d_self > hp.reach_m:
            continue
        cat = getattr(f, "rx_cat", None) or rx_category(f)
        pool = clinic_pool(f, cat, a, issue_rate, ext_rate, unknown_op)
        if pool <= 0:
            continue
        self_w = bw(d_self, hp.candidate_attractiveness)
        den = self_w
        for (plat, plon, ak) in comps:
            dk = haversine(plat, plon, f.lat, f.lon)
            if dk <= hp.reach_m:
                den += bw(dk, ak)
        share = self_w / den if den > 0 else 0.0
        rows.append({"clinic": f.name, "dist": d_self, "pool": pool,
                     "self_w": self_w, "comp_w": den - self_w,
                     "share": share, "captured": pool * share})
    rows.sort(key=lambda r: r["captured"], reverse=True)
    return rows


# ── 生データ＋現在の設定/手修正から、両トラックを算出（再検索なしで即再計算） ──────
def compute_candidate(raw):
    clat, clon, uni = raw["clat"], raw["clon"], raw["uni"]
    label = raw["label"]
    a = PredictionAssumptions()
    hp = make_hp()
    # 座標補正・削除・手動追加を反映した実効の医療機関/薬局リスト（①②の両方に効く）
    med, ph = effective_facilities(raw, clat, clon, label)
    # 周知率（接触率）：館の来店客数のうち、その薬局に接触・到達する割合。手修正が優先。
    exposure = float(st.session_state.get("exp_multi", {}).get(label, raw.get("exposure", 1.0)))
    eff_uni = uni * exposure
    fp = make_fp(eff_uni)
    hb = huff_breakdown(med, ph, clat, clon, hp, a,
                        float(ff_issue), float(ff_ext), float(med_unknown_op))
    med_total = sum(r["captured"] for r in hb)
    classified = classify_menkata(ph, med, clat, clon,
                                  monzen_dist=fp.menkata_monzen_dist,
                                  main_rx_threshold=fp.menkata_main_rx, reach_m=hp.reach_m)
    override = st.session_state.get("mk_multi", {}).get(label, {})
    cpow, cn, cexc = footfall_competitor_power(classified, override, fp.competitor_decay_m,
                                               hp.national_avg_rx)
    foot = compute_footfall_prediction(fp, cpow)
    return {
        "label": label, "name": raw["name"], "addr": raw["addr"],
        "uni": uni, "exposure": exposure, "eff_uni": eff_uni,
        "med": med, "ph": ph, "clat": clat, "clon": clon, "hp": hp, "fp": fp,
        "huff_rows": hb, "med_total": med_total, "classified": classified, "override": override,
        "comp_power": cpow, "comp_n": cn, "comp_excluded": cexc,
        "foot_total": (foot["total"] if foot else None), "foot": foot,
    }


# ════════════════════════════════ 数式入りExcel ════════════════════════════════
_HDR = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="0F766E")
_HDR_FILL2 = PatternFill("solid", fgColor="B45309")
_INP_FILL = PatternFill("solid", fgColor="FFF7E0")   # 編集できる入力＝薄い黄色
_CALC_FILL = PatternFill("solid", fgColor="EEF2F1")  # 自動計算＝薄いグレー
_BOLD = Font(bold=True)


def _sheet_name(label):
    return re.sub(r"[\\/*?:\[\]]", "_", str(label))[:20]


def _build_footfall_sheet(wb, r):
    fp = r["fp"]
    ws = wb.create_sheet(f"集客_{_sheet_name(r['label'])}")
    ws["A1"] = f"② 集客ベース（来店客数）  {r['label']}  {r['name']}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "【黄色のセルは編集できます。編集すると下の「獲得」が自動で再計算されます】"
    ws["A2"].font = Font(italic=True, size=9, color="B45309")
    # 入力（B3..B12）※B5=有効客数は自動計算
    ws.cell(row=3, column=1, value="館の来店客数（月間）")
    ws.cell(row=3, column=2, value=r["uni"]).fill = _INP_FILL
    ws.cell(row=4, column=1, value="周知率（接触率）")
    ws.cell(row=4, column=2, value=r["exposure"]).fill = _INP_FILL
    ws.cell(row=5, column=1, value="有効客数 ＝ 来店客数 × 周知率")
    ec = ws.cell(row=5, column=2, value="=B3*B4")
    ec.fill = _CALC_FILL
    ec.font = _BOLD
    inputs = [
        ("65歳以上の比率", fp.ratio_65plus), ("65+ 月受診回数", fp.visits_month_65plus),
        ("65- 月受診回数", fp.visits_month_under65), ("処方箋発行率", fp.issue_rate),
        ("院外処方率", fp.external_rate), ("当該薬局利用率", fp.use_rate),
        ("面競合の距離減衰λ(m)", fp.competitor_decay_m),
    ]  # B6..B12
    for k, (lab, val) in enumerate(inputs):
        rr = 6 + k
        ws.cell(row=rr, column=1, value=lab)
        ws.cell(row=rr, column=2, value=val).fill = _INP_FILL
    computed = [
        ("年間受診延べ(回)", "=(B5*B6*B7+B5*(1-B6)*B8)*12"),
        ("院外処方プール(枚)", "=B14*B9*B10"),
        ("面競合の実効パワー", "=SUM(E22:E500)"),
        ("シェア", "=B11/(1+B16)"),
        ("獲得（年間・枚）", "=B15*B17"),
        ("獲得（月間・枚）", "=B18/12"),
    ]  # B14..B19
    for k, (lab, f) in enumerate(computed):
        rr = 14 + k
        ws.cell(row=rr, column=1, value=lab)
        c = ws.cell(row=rr, column=2, value=f)
        c.fill = _CALC_FILL
        if rr in (18, 19):
            c.font = _BOLD
    for j, htxt in enumerate(["競合薬局名", "候補地から(m)", "実績(枚)", "面=1/門前=0", "重み(自動)"], start=1):
        c = ws.cell(row=21, column=j, value=htxt)
        c.font = _HDR
        c.fill = _HDR_FILL
    for k, cl in enumerate(r["classified"]):
        rr = 22 + k
        eff = r["override"].get(cl["key"], cl["auto_menkata"])
        ws.cell(row=rr, column=1, value=cl["name"])
        ws.cell(row=rr, column=2, value=round(cl["d_cand"])).fill = _INP_FILL
        ws.cell(row=rr, column=3, value=int(cl["rx"]) if cl["rx"] else 0).fill = _INP_FILL
        ws.cell(row=rr, column=4, value=(1 if eff else 0)).fill = _INP_FILL
        ws.cell(row=rr, column=5,
                value=f"=IF(D{rr}=1,IF(C{rr}>0,C{rr}/12000,1)*EXP(-B{rr}/$B$12),0)").fill = _CALC_FILL
    for col, w in zip("ABCDE", [30, 14, 12, 14, 14]):
        ws.column_dimensions[col].width = w


def _build_medical_sheet(wb, r):
    hp = r["hp"]
    ws = wb.create_sheet(f"医療機関_{_sheet_name(r['label'])}")
    ws["A1"] = f"① 医療機関ベース（ハフ競合按分）  {r['label']}  {r['name']}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "【黄色=編集可。取り分率・獲得は自動再計算。『競合の重み合計』はアプリ計算値（λ変更は自店側のみ反映）】"
    ws["A2"].font = Font(italic=True, size=9, color="B45309")
    for k, (lab, val) in enumerate([("距離減衰λ(m)", hp.lambda_m), ("門前ブースト", hp.monzen_boost),
                                    ("候補店の引力", hp.candidate_attractiveness)]):
        ws.cell(row=4 + k, column=1, value=lab)
        ws.cell(row=4 + k, column=2, value=val)
    heads = ["クリニック名", "距離(m)", "年間院外処方(原資)", "自店の重み", "競合の重み合計", "取り分率", "獲得(枚/年)"]
    for j, htxt in enumerate(heads, start=1):
        c = ws.cell(row=9, column=j, value=htxt)
        c.font = _HDR
        c.fill = _HDR_FILL2
    rr = 10
    for row in r["huff_rows"]:
        ws.cell(row=rr, column=1, value=row["clinic"])
        ws.cell(row=rr, column=2, value=round(row["dist"]))
        ws.cell(row=rr, column=3, value=round(row["pool"])).fill = _INP_FILL
        ws.cell(row=rr, column=4, value=round(row["self_w"], 4)).fill = _INP_FILL
        ws.cell(row=rr, column=5, value=round(row["comp_w"], 4)).fill = _INP_FILL
        ws.cell(row=rr, column=6, value=f"=IF((D{rr}+E{rr})>0,D{rr}/(D{rr}+E{rr}),0)").fill = _CALC_FILL
        ws.cell(row=rr, column=7, value=f"=C{rr}*F{rr}").fill = _CALC_FILL
        rr += 1
    ws.cell(row=rr, column=6, value="合計（＝①予測）").font = _BOLD
    tot = ws.cell(row=rr, column=7, value=f"=SUM(G10:G{rr-1})")
    tot.font = _BOLD
    ws["A7"] = "ハフ按分による予測（年間・枚）"
    ws["B7"] = f"=G{rr}"
    ws["B7"].font = _BOLD
    ws["C7"] = "（月間）"
    ws["D7"] = f"=B7/12"
    for col, w in zip("ABCDEFG", [28, 10, 18, 12, 14, 10, 14]):
        ws.column_dimensions[col].width = w


def _build_summary_sheet(ws, results):
    """お客様提示用のサマリー（比較表）。"""
    ink = "16211E"
    teal = "0F766E"
    thin = Side(style="thin", color="D7DEDB")
    med_thin = Side(style="thin", color="B9C4C0")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ncols = 8
    last_col = get_column_letter(ncols)
    # タイトル帯
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = "処方箋獲得予測  ─  出店候補地の比較"
    t.font = Font(bold=True, size=18, color=teal)
    t.alignment = left
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A2:{last_col}2")
    s = ws["A2"]
    s.value = "① 医療機関ベース（ハフ競合按分）  と  ② 集客ベース（来店客数）  の2つの独立した推計で比較"
    s.font = Font(size=10, color="5B6662")
    s.alignment = left
    ws.row_dimensions[2].height = 18

    # 2段ヘッダー（グループ見出し＋小見出し）
    hrow1, hrow2 = 4, 5
    groups = [("", 1), ("", 1), ("", 1), ("① 医療機関ベース", 2), ("② 集客ベース", 2), ("", 1)]
    col = 1
    for title, span in groups:
        if title and span > 1:
            ws.merge_cells(start_row=hrow1, start_column=col, end_row=hrow1, end_column=col + span - 1)
            cc = ws.cell(row=hrow1, column=col, value=title)
            cc.font = Font(bold=True, color="FFFFFF", size=10)
            cc.fill = PatternFill("solid", fgColor=teal)
            cc.alignment = center
            for k in range(span):
                ws.cell(row=hrow1, column=col + k).fill = PatternFill("solid", fgColor=teal)
        col += span
    subheads = ["ラベル", "店舗名 / メモ", "住所", "年間(枚)", "月間(枚)", "年間(枚)", "月間(枚)", "予測レンジ(年)"]
    for j, htxt in enumerate(subheads, start=1):
        cc = ws.cell(row=hrow2, column=j, value=htxt)
        cc.font = Font(bold=True, color="FFFFFF", size=10)
        cc.fill = PatternFill("solid", fgColor="16897E")
        cc.alignment = center
        cc.border = Border(left=thin, right=thin, top=thin, bottom=med_thin)

    # 勝者（集客 or 医療機関が最大）をハイライト
    def keyval(r):
        return r["foot_total"] if r["foot_total"] is not None else (r["med_total"] or 0)
    best_label = max(results, key=keyval)["label"] if results else None

    r0 = hrow2 + 1
    for i, r in enumerate(results):
        row = r0 + i
        ff = f"集客_{_sheet_name(r['label'])}"
        md = f"医療機関_{_sheet_name(r['label'])}"
        vals = {
            1: r["label"], 2: r["name"], 3: r["addr"],
            4: f"='{md}'!B7", 5: f"='{md}'!D7",
        }
        if r["foot_total"] is not None:
            vals[6] = f"='{ff}'!B18"
            vals[7] = f"='{ff}'!B19"
        vals[8] = (f'=IF(AND(ISNUMBER(D{row}),ISNUMBER(F{row})),'
                   f'TEXT(MIN(D{row},F{row}),"#,##0")&"〜"&TEXT(MAX(D{row},F{row}),"#,##0"),"—")')
        is_best = (r["label"] == best_label)
        base_fill = PatternFill("solid", fgColor="E8F3F1") if is_best else (
            PatternFill("solid", fgColor="F6F8F7") if i % 2 else None)
        for j in range(1, ncols + 1):
            cell = ws.cell(row=row, column=j, value=vals.get(j))
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if base_fill:
                cell.fill = base_fill
            if j == 1:
                cell.font = Font(bold=True, size=11, color=teal)
                cell.alignment = center
            elif j in (4, 5, 6, 7):
                cell.number_format = "#,##0"
                cell.alignment = right
                cell.font = Font(size=11, color=ink)
            elif j == 8:
                cell.alignment = center
            else:
                cell.alignment = left
        ws.row_dimensions[row].height = 22

    note = r0 + len(results) + 1
    ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=ncols)
    n = ws.cell(row=note, column=1,
                value="※ ①と②は同じ枚数を別データから見積もった2つの推計です（足し算しません）。"
                      "詳細な内訳は各「集客_」「医療機関_」シートを参照。")
    n.font = Font(italic=True, size=9, color="7A8481")
    ws.merge_cells(start_row=note + 1, start_column=1, end_row=note + 1, end_column=ncols)
    ws.cell(row=note + 1, column=1,
            value="※ 緑の行は①②の予測が最大の候補地。").font = Font(italic=True, size=9, color="7A8481")

    for col, w in zip("ABCDEFGH", [8, 24, 34, 13, 12, 13, 12, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False


def build_excel(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "比較サマリー"
    _build_summary_sheet(ws, results)
    for r in results:
        _build_footfall_sheet(wb, r)
        _build_medical_sheet(wb, r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════ メイン ════════════════════════════════
st.title("🏪 複数店舗（A / B / C）比較 — 処方箋獲得予測")
st.caption(
    "同じスーパーの複数の出店候補地をまとめて分析し、2トラック（医療機関ベース × 集客ベース）で"
    "比較します。ロジックの内訳はブラウザで確認でき、数式入りExcelにも書き出せます。"
)

st.markdown("#### 1. 候補地を入力")
st.caption("候補地ごとに ラベル・店舗名/メモ・住所・月間ユニーク客数（集客ベース用）を入力してください。行は追加できます。")

if "cand_df" not in st.session_state:
    st.session_state["cand_df"] = pd.DataFrame([
        {"ラベル": "A", "店舗名/メモ": "", "住所": "", "月間ユニーク客数": 0, "周知率": 1.0},
        {"ラベル": "B", "店舗名/メモ": "", "住所": "", "月間ユニーク客数": 0, "周知率": 1.0},
        {"ラベル": "C", "店舗名/メモ": "", "住所": "", "月間ユニーク客数": 0, "周知率": 1.0},
    ])

cand_edited = st.data_editor(
    st.session_state["cand_df"], num_rows="dynamic", use_container_width=True, key="cand_editor",
    column_config={
        "ラベル": st.column_config.TextColumn("ラベル", width="small"),
        "店舗名/メモ": st.column_config.TextColumn("店舗名/メモ"),
        "住所": st.column_config.TextColumn("住所", width="large"),
        "月間ユニーク客数": st.column_config.NumberColumn("月間ユニーク客数", min_value=0, step=500),
        "周知率": st.column_config.NumberColumn(
            "周知率", min_value=0.0, max_value=1.0, step=0.05, format="%.2f",
            help="館の来店客のうち、その薬局に接触・到達する割合。"
                 "食品スーパー=1.0／大型モール1階・主動線沿い=0.3／上層階・動線外=0.1。後から候補地ごとに変更可。",
        ),
    },
)
st.caption("💡 **周知率**：食品スーパー=**1.0**（既定）／大型モール1階・主動線=**0.3**／上層階・動線外=**0.1**。"
           "館の来店客数が大きくても、薬局に接触する割合を掛けて過大予測を防ぎます。")

col_run, col_clear = st.columns([3, 1])
run = col_run.button("▶ 未分析の候補地を分析（続きから）", type="primary", use_container_width=True)
if col_clear.button("🗑 結果をクリア", use_container_width=True):
    for k in ("mk_multi", "exp_multi", "med_edit", "ph_edit"):
        st.session_state[k] = {}
    st.session_state["multi_raw"] = []
    if os.path.exists(SAVE_FILE):
        try:
            os.remove(SAVE_FILE)
        except OSError:
            pass
    st.rerun()
st.caption("※ 分析は**1店ずつ完了と同時に保存**されます。途中で止まっても完了分は残り、もう一度押せば"
           "**続きから**処理します。3店同時＋全件取得ONは長時間になり中断されやすいので、"
           "**うまくいかない時は1〜2店ずつ、または全件取得をOFF**にしてお試しください。")

# 保存／再読込（検索や手修正の結果を残す。アプリを再起動しても復元）
col_save, col_load = st.columns(2)
if col_save.button("💾 現在の結果を保存（再検索不要に）", use_container_width=True,
                   disabled=not st.session_state.get("multi_raw")):
    try:
        save_session()
        st.success("保存しました。次回起動時に自動で復元されます（手動は『📂 保存を読込』）。")
    except Exception as e:
        st.error(f"保存に失敗: {e}")
if col_load.button("📂 保存を読込", use_container_width=True, disabled=not os.path.exists(SAVE_FILE)):
    try:
        load_session()
        st.rerun()
    except Exception as e:
        st.error(f"読込に失敗: {e}")

with st.expander("🔁 ファイルで保存／復元（URL共有・複数人・持ち運び用）", expanded=False):
    st.caption("上の『💾保存』はこのPC内に保存する方式で、**あなた一人のローカル利用向け**です。"
               "**URLを他の人に配って使う場合**は、各自が下のボタンで**自分のPCにダウンロード**し、"
               "使うときに**アップロードして復元**してください（他人と混ざらず、サーバー再起動でも消えません）。")
    dc1, dc2 = st.columns(2)
    dc1.download_button(
        "⬇️ 保存ファイルをダウンロード", data=(state_bytes() if st.session_state.get("multi_raw") else b"{}"),
        file_name="処方箋予測_保存.json", mime="application/json",
        use_container_width=True, disabled=not st.session_state.get("multi_raw"))
    up = dc2.file_uploader("⬆️ 保存ファイルから復元（アップロード）", type=["json"], key="restore_up")
    if up is not None:
        sig = (up.name, up.size)
        if st.session_state.get("_last_up") != sig:
            st.session_state["_last_up"] = sig
            try:
                _apply_state(json.loads(up.getvalue().decode("utf-8")))
                st.success("復元しました。")
                st.rerun()
            except Exception as e:
                st.error(f"復元に失敗: {e}")

# 起動時に一度だけ、前回保存を自動復元（消えても再検索しなくて済む）
if "auto_loaded" not in st.session_state:
    st.session_state["auto_loaded"] = True
    if not st.session_state.get("multi_raw") and os.path.exists(SAVE_FILE):
        try:
            load_session()
            st.info("前回の保存を復元しました（クリアしたい場合は『🗑 結果をクリア』）。")
        except Exception:
            pass

if run:
    scraper = get_scraper()
    assumptions = PredictionAssumptions()
    targets = [row for _, row in cand_edited.iterrows() if str(row.get("住所", "")).strip()]
    if not targets:
        st.warning("住所を1件以上入力してください。")
        st.stop()
    existing = st.session_state.setdefault("multi_raw", [])
    done_labels = {r["label"] for r in existing}
    todo = []
    for i, row in enumerate(targets):
        label = str(row.get("ラベル") or f"#{i+1}").strip()
        if label not in done_labels:
            todo.append((label, row))
    if not todo:
        st.info("入力中の候補地はすべて分析済みです。やり直すには『🗑 結果をクリア』を押してください。")
    else:
        st.info(f"未分析 {len(todo)}件を順に分析します（1件あたり数分。全件取得ONだと更に時間がかかります）。")
        overall = st.progress(0.0, text="開始…")
        for i, (label, row) in enumerate(todo):
            addr = str(row["住所"]).strip()
            uni = float(row.get("月間ユニーク客数") or 0)
            overall.progress(i / len(todo), text=f"[{label}] {addr} を分析中… ({i+1}/{len(todo)})")
            log = []
            prog = st.progress(0.0, text=f"[{label}] 収集中…")
            try:
                med, ph, clat, clon = run_analysis(
                    addr, int(radius_m), gate_m, int(max_detail), log, prog,
                    assumptions=assumptions, polygons=[], exclude_outside_med=True,
                )
                prog.empty()
            except Exception as e:
                prog.empty()
                st.error(f"[{label}] 分析に失敗: {e}")
                continue
            if clat is None:
                st.error(f"[{label}] 住所の座標が取得できませんでした：{addr}")
                continue
            exposure = float(row.get("周知率") if row.get("周知率") is not None else 1.0)
            # ★完了と同時に保存（中断されても、ここまで完了した店は残る）
            st.session_state["multi_raw"].append({
                "label": label, "name": str(row.get("店舗名/メモ") or ""),
                "addr": addr, "uni": uni, "exposure": exposure,
                "clat": clat, "clon": clon, "med": med, "ph": ph})
        overall.progress(1.0, text="完了")
        overall.empty()
        try:
            save_session()  # 検索完了時に自動保存（消えても再検索不要）
        except Exception:
            pass
        st.rerun()


# ── 結果の比較表示（毎回、現在の設定＋手修正で再計算） ──────────────────────────
raws = st.session_state.get("multi_raw", [])
if raws:
    computed = [compute_candidate(r) for r in raws]

    st.markdown("#### 2. 比較結果")
    rows = []
    for c in computed:
        med, foot = c["med_total"], c["foot_total"]
        vals = [v for v in (med, foot) if v is not None]
        rng = f"{min(vals):,.0f}〜{max(vals):,.0f}" if len(vals) == 2 else "—"
        rows.append({
            "ラベル": c["label"], "店舗名/メモ": c["name"], "住所": c["addr"][:24],
            "① 医療機関(年)": round(med) if med is not None else None,
            "① 医療機関(月)": round(med / 12) if med is not None else None,
            "② 集客(年)": round(foot) if foot is not None else None,
            "② 集客(月)": round(foot / 12) if foot is not None else None,
            "予測レンジ(年)": rng, "面競合数": c["comp_n"], "寄与医療機関数": len(c["huff_rows"]),
        })
    st.dataframe(
        pd.DataFrame(rows), hide_index=True, use_container_width=True,
        column_config={cc: st.column_config.NumberColumn(cc, format="%d 枚")
                       for cc in ["① 医療機関(年)", "① 医療機関(月)", "② 集客(年)", "② 集客(月)"]},
    )
    ranked = sorted(computed, key=lambda c: (c["foot_total"] or c["med_total"] or 0), reverse=True)
    st.success(f"🏆 最大の候補地： **{ranked[0]['label']}**"
               f"（{ranked[0]['name'] or ranked[0]['addr'][:20]}）")

    # ── 3. 候補地ごとの詳細（ロジックの内訳＋面/門前の目視修正） ──────────────
    st.markdown("#### 3. 候補地ごとの詳細（ロジックの内訳）")
    sel = st.selectbox("詳細を見る候補地", [c["label"] for c in computed],
                       format_func=lambda x: f"{x}　" + next((c["name"] or c["addr"][:16]
                                                              for c in computed if c["label"] == x), ""))
    c = next(x for x in computed if x["label"] == sel)

    m1, m2, m3 = st.columns(3)
    m1.metric("① 医療機関ベース（ハフ）", f"{c['med_total']:,.0f} 枚/年",
              f"月 {c['med_total']/12:,.0f} 枚")
    if c["foot_total"] is not None:
        m2.metric("② 集客ベース", f"{c['foot_total']:,.0f} 枚/年", f"月 {c['foot_total']/12:,.0f} 枚")
    else:
        m2.metric("② 集客ベース", "未入力")
    m3.metric("面競合数 / 除外", f"{c['comp_n']} / {c['comp_excluded']}")

    st.markdown("##### ① 医療機関ベース：ハフの取り分内訳")
    st.caption("各クリニックが出す処方箋（原資）を、自店の重み ÷（自店の重み＋競合の重み合計）の"
               "取り分率で獲得します。合計＝①予測。")
    hb_df = pd.DataFrame([{
        "医療機関": row["clinic"], "距離(m)": round(row["dist"]),
        "年間院外処方(原資)": round(row["pool"]),
        "自店の重み": round(row["self_w"], 3), "競合の重み合計": round(row["comp_w"], 3),
        "取り分率": round(row["share"], 3), "獲得(枚/年)": round(row["captured"]),
    } for row in c["huff_rows"]])
    st.dataframe(hb_df, hide_index=True, use_container_width=True, column_config={
        "取り分率": st.column_config.NumberColumn("取り分率", format="%.3f"),
        "獲得(枚/年)": st.column_config.NumberColumn("獲得(枚/年)", format="%d 枚"),
    })

    med_flags = [clinic_flag(f, int(med_high_thr)) for f in c["med"]]
    n_alert = sum(1 for x in med_flags if x)
    if n_alert:
        st.warning(f"⚠️ 医療機関に **{n_alert}件** の要確認あり（外れ値/外来不明）。下の表『検証』列を確認し、"
                   "外来数・院外区分・座標を必要に応じて修正してください。")
    with st.expander("🔧 医療機関の確認・修正（座標／外来数／院外区分／漏れの追加・削除 → ①に反映）", expanded=bool(n_alert)):
        st.caption(
            "『距離(m)』が実態と違う施設は座標がズレています（緯度経度を直すのが最も正確／距離を直接入力も可）。"
            "『院外区分』は院外のみ=100%・院内外どちらも=院外率・院内のみ=0で原資に反映（手で修正可）。"
            "『外来(人/日)』が不明のクリニックは既定値を使用（下の警告参照）。行の追加/削除で漏れ・誤検出を補正。"
        )
        med_edit = st.session_state.setdefault("med_edit", {})
        recs = med_edit.get(sel, [])
        fmap = {facility_key(f): fl for f, fl in zip(c["med"], med_flags)}
        disp = pd.DataFrame([{
            "医療機関": r.get("name"),
            "距離(m)": (round(haversine(c["clat"], c["clon"], _num(r.get("lat")), _num(r.get("lon"))))
                       if (_num(r.get("lat")) is not None and _num(r.get("lon")) is not None) else None),
            "緯度": _num(r.get("lat")), "経度": _num(r.get("lon")),
            "外来(人/日)": r.get("op"),
            "院外区分": (r.get("cat") if r.get("cat") in _EXT_CATS else "不明"),
            "検証": fmap.get(r.get("_key"), (
                "外来不明→既定使用" if _num(r.get("op")) is None else
                (f"要確認：{int(_num(r.get('op')))}人/日" if _num(r.get("op")) >= med_high_thr else ""))),
            "_key": r.get("_key"),
        } for r in recs])
        ed = st.data_editor(
            disp, hide_index=True, use_container_width=True, num_rows="dynamic",
            key=f"med_edit_{sel}", disabled=["検証", "_key"],
            column_config={
                "緯度": st.column_config.NumberColumn("緯度", format="%.6f"),
                "経度": st.column_config.NumberColumn("経度", format="%.6f"),
                "距離(m)": st.column_config.NumberColumn("距離(m)", help="正しい距離を直接入力しても補正できます（向き保持）。最も正確なのは緯度経度。"),
                "外来(人/日)": st.column_config.NumberColumn("外来(人/日)", min_value=0, step=1),
                "院外区分": st.column_config.SelectboxColumn("院外区分", options=_EXT_CATS, width="medium"),
            },
        )
        new_recs = resolve_edit(ed, "医療機関", "外来(人/日)", "op", recs, c["clat"], c["clon"],
                                cat_col="院外区分")
        changed = (_recs_sig(new_recs, "op") != _recs_sig(recs, "op")
                   or [r.get("cat") for r in new_recs] != [r.get("cat") for r in recs])
        if changed:
            med_edit[sel] = new_recs
            st.rerun()

    if c["foot"]:
        st.markdown("##### ② 集客ベース：内訳")
        fo, fp = c["foot"], c["fp"]
        exp_all = st.session_state.setdefault("exp_multi", {})
        new_exp = st.number_input(
            f"周知率（{sel}）— 館の来店客のうち薬局に接触する割合", 0.0, 1.0,
            float(c["exposure"]), 0.05, format="%.2f", key=f"exp_{sel}",
            help="食品スーパー=1.0／大型モール1階・主動線=0.3／上層階・動線外=0.1。変更すると②が再計算されます。",
        )
        if abs(new_exp - c["exposure"]) > 1e-9:
            exp_all[sel] = new_exp
            st.rerun()
        st.markdown(
            f"- 館の来店客数 {c['uni']:,.0f}人 × **周知率 {c['exposure']:.2f}** = 有効客数 **{c['eff_uni']:,.0f}人**"
            f"（65+ {fo['u65']:,.0f} / 65− {fo['u_under']:,.0f}）\n"
            f"- 年間受診延べ {fo['annual_visits']:,.0f}回 → 院外処方プール {fo['rx_pool']:,.0f}枚\n"
            f"- 利用率 {fp.use_rate:.1%} ÷ (面競合の実効パワー {c['comp_power']:.1f}"
            f"〔面{c['comp_n']}店・距離減衰λ={fp.competitor_decay_m:.0f}m〕 + 1) = シェア {fo['share']:.2%}\n"
            f"- **獲得 = {fo['total']:,.0f} 枚/年**"
        )

    with st.expander("🔧 薬局の確認・修正（座標／実績／面・門前／漏れの追加・削除 → ①②に反映）", expanded=False):
        st.caption(
            "座標(緯度・経度)がズレている薬局は正しい値に直し（距離(m)直接入力も可）、漏れている薬局は行追加、"
            "誤検出は行削除。『面/門前』もここで修正でき、②の面競合に即反映されます"
            "（『最寄りクリニック(m)』が近い店が門前の目安）。"
        )
        ph_edit = st.session_state.setdefault("ph_edit", {})
        mk_all = st.session_state.setdefault("mk_multi", {})
        mk = mk_all.setdefault(sel, {})
        precs = ph_edit.get(sel, [])
        clmap = {r["key"]: r for r in c["classified"]}
        pdisp = pd.DataFrame([{
            "薬局": r.get("name"),
            "距離(m)": (round(haversine(c["clat"], c["clon"], _num(r.get("lat")), _num(r.get("lon"))))
                       if (_num(r.get("lat")) is not None and _num(r.get("lon")) is not None) else None),
            "緯度": _num(r.get("lat")), "経度": _num(r.get("lon")),
            "実績(枚/年)": r.get("rx"),
            "最寄りクリニック(m)": (round(clmap[r.get("_key")]["nearest_clinic"])
                                if (r.get("_key") in clmap and clmap[r.get("_key")]["nearest_clinic"] < 1e8) else None),
            "面/門前": ("面" if mk.get(r.get("_key"),
                       (clmap[r.get("_key")]["auto_menkata"] if r.get("_key") in clmap else True)) else "門前"),
            "_key": r.get("_key"),
        } for r in precs])
        ped = st.data_editor(
            pdisp, hide_index=True, use_container_width=True, num_rows="dynamic",
            key=f"ph_edit_{sel}", disabled=["距離(m)", "最寄りクリニック(m)", "_key"],
            column_config={
                "緯度": st.column_config.NumberColumn("緯度", format="%.6f"),
                "経度": st.column_config.NumberColumn("経度", format="%.6f"),
                "距離(m)": st.column_config.NumberColumn("距離(m)", help="正しい距離を直接入力しても補正できます（向き保持）。最も正確なのは緯度経度。"),
                "実績(枚/年)": st.column_config.NumberColumn("実績(枚/年)", min_value=0, step=100),
                "面/門前": st.column_config.SelectboxColumn("面/門前", options=["面", "門前"], width="small",
                                                       help="面＝集客の競合に数える／門前＝競合から外す"),
            },
        )
        pnew = resolve_edit(ped, "薬局", "実績(枚/年)", "rx", precs, c["clat"], c["clon"])
        auto_men = {r["key"]: r["auto_menkata"] for r in c["classified"]}
        new_mk = {}
        for _, row in ped.iterrows():
            k = row.get("_key")
            if not (isinstance(k, str) and k):
                continue
            is_men = (row.get("面/門前") == "面")
            if k in auto_men and is_men != auto_men[k]:
                new_mk[k] = is_men
        if _recs_sig(pnew, "rx") != _recs_sig(precs, "rx") or new_mk != mk:
            ph_edit[sel] = pnew
            mk_all[sel] = new_mk
            st.rerun()

    # ── 4. Excel ─────────────────────────────────────────────────────────────
    st.markdown("#### 4. 数式入りExcelで書き出し")
    st.caption("集客シートは客数・競合・係数、医療機関シートは原資・重みを編集すると、"
               "獲得枚数がExcelの数式で自動再計算されます（ブラウザを閉じても手元で編集可能）。")
    st.download_button(
        "📊 比較Excel（数式入り）をダウンロード",
        data=build_excel(computed), file_name="処方箋予測_複数店舗比較.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
