# -*- coding: utf-8 -*-
"""
薬局 出店候補地 分析ツール（単一ファイル版）
============================================================
とあるスーパーの複数の出店候補地（A点・B点・C点…）をまとめて分析し、
「医療機関ベース（ハフ按分）」と「集客ベース（来店客数）」の2トラックで比較する。

このファイルは単体で動きます（既存の 260702_Prescription Analysis_v2.py は
コピー元として無改変で温存。実行には .venv の依存ライブラリが必要）。

主な機能:
- 候補地ごとに ①ハフの取り分内訳・②集客の内訳 を表示、数式入りExcelに書き出し
- 商圏マップ（候補地＋周辺の医療機関・薬局。手動追加/削除/座標修正も反映）
- 診療科別の処方箋発行率（整形外科など投薬の少ない科を反映。値は編集可）
"""
import streamlit as st

st.set_page_config(page_title="薬局 出店候補地 分析ツール", page_icon="🏪", layout="wide")

# ════════════ 予測モデル／スクレイパー（260702_Prescription Analysis_v2.py より無改変で取り込み） ════════════
# ※ 元ファイルのモジュール説明はコメント化（Streamlitのマジック表示で本文に出るのを防ぐため削除）。
import csv
import io
import math
import re
import time
import urllib.parse
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import folium
from folium.plugins import Draw
import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup
from streamlit_folium import st_folium

# ─── ページ設定（必ずファイル先頭のstコマンドより前に置く） ─────────────────────


# ─── 定数 ─────────────────────────────────────────────────────────────────────
MHLW_DOMAIN = "https://www.iryou.teikyouseido.mhlw.go.jp"
MHLW_BASE   = MHLW_DOMAIN + "/znk-web"
WORKING_DAYS = 305

# ─── 処方箋獲得予測モデル（流入率アサンプション） ──────────────────────────────
# 各医療機関 → 候補薬局地点までの直線距離帯ごとの「流入率」。
# 医療機関の外来患者のうち、この地点の薬局に処方箋を持ち込む割合を表す
# 経験値。(上限距離m, 流入率) の昇順リスト。最初に dist<=上限 に合致した帯を採用。
DEFAULT_INFLOW_BANDS: List[Tuple[float, float]] = [
    (50.0,   0.570),   # 〜50m（門前）
    (500.0,  0.070),   # 50m〜500m
    (1000.0, 0.050),   # 500m〜1km
    (2000.0, 0.012),   # 1km〜2km
    (3000.0, 0.004),   # 2km〜3km
    # 3km超は 0%（帯に該当しなければ流入0）
]


@dataclass
class PredictionAssumptions:
    """処方箋獲得予測で使う調整可能なアサンプション一式。"""
    bands: List[Tuple[float, float]] = field(
        default_factory=lambda: list(DEFAULT_INFLOW_BANDS)
    )
    annual_days_mode: str = "weekly"       # "weekly"=週診療日数×52 / "fixed"=固定日数
    fixed_annual_days: int = WORKING_DAYS  # weekly不明時のフォールバック日数
    external_factor_gairai: float = 1.0    # 院外処方あり の寄与係数
    external_factor_inhouse: float = 0.0   # 院内処方のみ の寄与係数（原則0）
    external_factor_unknown: float = 1.0   # 院内外不明 の寄与係数
    issue_rate: float = 1.0                # 処方箋発行率（流入率に織込済なら1.0）
    discount_contested_monzen: bool = False  # 門前競合クリニックを面レートに引下げるか
    cosmetic_factor: float = 0.0           # 美容・自由診療クリニックの寄与係数（保険処方箋ほぼ0）
    dental_factor: float = 0.05            # 歯科診療所の寄与係数（外来1回あたり発行率が医科より大幅に低い）


def inflow_rate_for_distance(
    dist_m: Optional[float], bands: List[Tuple[float, float]]
) -> float:
    """距離(m)に対応する流入率を返す。どの帯にも該当しなければ0。"""
    if dist_m is None:
        return 0.0
    for upper, rate in sorted(bands, key=lambda b: b[0]):
        if dist_m <= upper:
            return rate
    return 0.0


def inflow_band_label(dist_m: Optional[float], bands: List[Tuple[float, float]]) -> str:
    """距離が属する帯の人間可読ラベル（例「〜50m（門前）」）を返す。"""
    if dist_m is None:
        return "座標なし"
    sorted_bands = sorted(bands, key=lambda b: b[0])
    prev = 0.0
    for upper, _rate in sorted_bands:
        if dist_m <= upper:
            if upper <= 50:
                return f"〜{int(upper)}m（門前）"
            lo = f"{int(prev)}m" if prev < 1000 else f"{prev/1000:.0f}km"
            hi = f"{int(upper)}m" if upper < 1000 else f"{upper/1000:.0f}km"
            return f"{lo}〜{hi}"
        prev = upper
    top = sorted_bands[-1][0]
    return f"{top/1000:.0f}km超（流入0）"


def _external_factor(rx_summary: str, a: PredictionAssumptions) -> float:
    """院内外処方の別から寄与係数を決める。"""
    if rx_summary.startswith("院外処方あり"):
        return a.external_factor_gairai
    if rx_summary == "院内処方のみ":
        return a.external_factor_inhouse
    return a.external_factor_unknown


def facility_key(fac: "MedFacility") -> str:
    """医療機関を一意に識別するキー（手動上書き辞書のキーに使用）。"""
    return fac.kikan_cd or f"name:{fac.name}"


def compute_pharmacy_proximity(
    med_facs: List["MedFacility"], pharmacies: List["PharmacyFacility"]
) -> None:
    """
    各医療機関について、最も近い既存薬局までの距離を計算して書き戻す。
    門前占有チェック（そのクリニックの門前に既に別薬局が張り付いているか）に使う。
    候補地の新店は pharmacies に含まれないため、全て競合薬局として扱われる。
    """
    ph_coords = [p for p in pharmacies if p.lat is not None and p.lon is not None]
    for fac in med_facs:
        if fac.lat is None or fac.lon is None:
            fac.nearest_pharmacy_dist_m = None
            fac.nearest_pharmacy_name = ""
            continue
        best_d = float("inf")
        best_name = ""
        for p in ph_coords:
            d = haversine(fac.lat, fac.lon, p.lat, p.lon)
            if d < best_d:
                best_d = d
                best_name = p.name
        if best_name:
            fac.nearest_pharmacy_dist_m = best_d
            fac.nearest_pharmacy_name = best_name
        else:
            fac.nearest_pharmacy_dist_m = None
            fac.nearest_pharmacy_name = ""


def compute_capture_prediction(
    med_facs: List["MedFacility"],
    a: PredictionAssumptions,
    op_override: Optional[Dict[str, float]] = None,
) -> Dict[str, float]:
    """
    各医療機関について、候補地点（商圏中心）が獲得できる年間処方箋枚数を計算し、
    MedFacility に結果を書き戻す。合計値等のサマリーdictを返す。

        年間外来延べ数 = 1日平均外来患者数 × 年間診療日数
        獲得処方箋     = 年間外来延べ数 × 院外処方係数 × 発行率 × 流入率(距離)

    op_override: {facility_key: 外来患者数} でナビィ値を手動上書き（大病院のHP値等）。
                 ナビィ原値 fac.daily_outpatients は保持したまま計算にのみ反映する。

    門前占有: 候補地が門前バンド(≤先頭帯上限)に入るクリニックで、既に別の薬局が
             同じ門前圏に張り付いている場合 monzen_contested=True を立てる。
             流入率(門前0.570)は「自店が門前になれる」前提の実績値のため、
             椅子が埋まっているクリニックは過大評価になりうる旨をアラートする。
             discount_contested_monzen=True のときは面レート(第2帯)へ自動引下げ。
    """
    op_override = op_override or {}
    sorted_bands = sorted(a.bands, key=lambda b: b[0])
    monzen_radius = sorted_bands[0][0] if sorted_bands else 50.0
    men_rate = sorted_bands[1][1] if len(sorted_bands) >= 2 else 0.0
    total = 0.0
    n_contrib = 0
    n_no_op = 0
    n_contested = 0
    for fac in med_facs:
        # 門前占有の物理判定（外来数の有無に関わらず算出）
        fac.monzen_occupied = (
            fac.nearest_pharmacy_dist_m is not None
            and fac.nearest_pharmacy_dist_m <= monzen_radius
        )
        # 候補地から見て門前バンドに入るクリニックか
        cand_is_monzen = fac.distance_m is not None and fac.distance_m <= monzen_radius
        fac.monzen_contested = bool(cand_is_monzen and fac.monzen_occupied)

        # 年間診療日数
        if a.annual_days_mode == "weekly" and fac.weekly_op_days:
            days = fac.weekly_op_days * 52.0
        else:
            days = float(a.fixed_annual_days)
        fac.annual_op_days_used = days

        # 外来患者数（手動上書き優先）
        ov = op_override.get(facility_key(fac))
        if ov is not None and ov > 0:
            eff_op: Optional[float] = float(ov)
            fac.outpatient_manual = True
        else:
            eff_op = fac.daily_outpatients
            fac.outpatient_manual = False

        if eff_op is None:
            fac.annual_op_visits = None
            fac.external_rx_factor = _external_factor(fac.rx_summary, a)
            fac.inflow_rate = 0.0 if not fac.in_area else inflow_rate_for_distance(fac.distance_m, a.bands)
            fac.inflow_band = ("商圏外（ポリゴン）" if not fac.in_area
                               else inflow_band_label(fac.distance_m, a.bands))
            fac.captured_rx = None
            if fac.in_area:
                n_no_op += 1
            continue

        annual_visits = eff_op * days
        factor = _external_factor(fac.rx_summary, a)
        # 美容・自由診療は保険処方箋がほぼ発生しない／歯科は発行率が医科より大幅に低い
        if fac.is_cosmetic:
            factor *= a.cosmetic_factor
        elif fac.facility_category == "歯科診療所":
            factor *= a.dental_factor
        rate = inflow_rate_for_distance(fac.distance_m, a.bands)

        # 門前競合クリニックは、任意で面レートへ引下げ（デフォルトは引下げず表示のみ）
        if fac.monzen_contested and a.discount_contested_monzen:
            rate = men_rate

        # 商圏ポリゴン外（川・線路等で分断）のクリニックは寄与から除外
        band_label = inflow_band_label(fac.distance_m, a.bands)
        if not fac.in_area:
            rate = 0.0
            band_label = "商圏外（ポリゴン）"

        captured = annual_visits * factor * a.issue_rate * rate

        fac.annual_op_visits = annual_visits
        fac.external_rx_factor = factor
        fac.inflow_rate = rate
        fac.inflow_band = band_label
        fac.captured_rx = captured
        if captured > 0:
            n_contrib += 1
            total += captured
        # 門前競合かつ実際に門前レートで寄与しているクリニックのみアラート対象
        if fac.monzen_contested and captured > 0:
            n_contested += 1

    return {
        "total_annual_rx": total,
        "total_daily_rx": total / a.fixed_annual_days if a.fixed_annual_days else 0.0,
        "n_contributing": n_contrib,
        "n_no_outpatient": n_no_op,
        "n_contested_monzen": n_contested,
        "n_outside_area": sum(1 for f in med_facs if not f.in_area),
    }


def predict_at_point(
    lat: float,
    lon: float,
    med_facs: List["MedFacility"],
    a: PredictionAssumptions,
    op_override: Optional[Dict[str, float]] = None,
) -> float:
    """
    任意地点の年間獲得処方箋数を計算する（MedFacilityの状態は変更しない純関数）。
    実績照合タブで「既存薬局の位置にモデルを当てたらいくつになるか」を出すのに使う。
    """
    op_override = op_override or {}
    total = 0.0
    for fac in med_facs:
        if fac.lat is None or fac.lon is None:
            continue
        d = haversine(lat, lon, fac.lat, fac.lon)
        rate = inflow_rate_for_distance(d, a.bands)
        if rate <= 0:
            continue
        ov = op_override.get(facility_key(fac))
        op = float(ov) if (ov is not None and ov > 0) else fac.daily_outpatients
        if not op:
            continue
        if a.annual_days_mode == "weekly" and fac.weekly_op_days:
            days = fac.weekly_op_days * 52.0
        else:
            days = float(a.fixed_annual_days)
        factor = _external_factor(fac.rx_summary, a)
        if fac.is_cosmetic:
            factor *= a.cosmetic_factor
        elif fac.facility_category == "歯科診療所":
            factor *= a.dental_factor
        total += op * days * factor * a.issue_rate * rate
    return total


# ─── 2トラック予測：①医療機関ベース（ハフ競合按分）／②集客ベース（来店客数） ────
# 検証（面型275店 vs ナビィ実績）で、加算型は面型を中央値2.75倍過大・実績とほぼ無相関だった。
# 主因は「各クリニック外来の固定割合を、周辺に何店薬局があろうとこの1店に独占計上」していたこと。
# ①はこれを競合薬局との按分（ハフ＝引力×距離）に置き換えて過大を是正する。
# ②はスーパー等の来店客数から直接、面で取れる枚数を見積もる独立トラック。両者を併記する。

@dataclass
class HuffParams:
    """医療機関ベース（ハフ）予測のパラメータ。既定は面型275店の検証で
    過大がほぼ解消した設定（λ=250m・門前×8・純距離＝競合の引力は一律1）。"""
    enabled: bool = True
    lambda_m: float = 250.0                 # 距離減衰の距離定数（小さいほど近距離に集中）
    monzen_boost: float = 8.0               # 門前(≤monzen_radius)の引力ブースト
    monzen_radius: float = 50.0
    candidate_attractiveness: float = 1.0   # 候補店の引力（集客力/規模）。1.0=全国平均並み
    weight_by_power: bool = False           # 競合薬局を実績枚数で引力加重するか（既定OFF＝一律1）
    national_avg_rx: float = 12000.0        # 引力換算の基準（全国平均 年間枚数）
    reach_m: float = 3000.0                 # 商圏（3km円）


@dataclass
class FootfallParams:
    """集客ベース（来店客数）予測のパラメータ。いただいた店舗ファイル式を年齢2区分に拡張。"""
    enabled: bool = False
    store_format: str = "食品スーパー"
    unique_customers_monthly: float = 0.0   # 月間ユニーク客数（会員数 or POS客数÷来店回数で算出）
    ratio_65plus: float = 0.30              # ユニーク客のうち65歳以上の比率
    visits_month_65plus: float = 3.0        # 65歳以上の月平均受診回数
    visits_month_under65: float = 1.3       # 65歳未満の月平均受診回数
    issue_rate: float = 0.8054              # 処方箋発行率
    external_rate: float = 0.8313           # 院外処方率
    use_rate: float = 0.137                 # 当該薬局利用率
    national_avg_rx: float = 12000.0        # 競合パワー換算の基準
    menkata_monzen_dist: float = 50.0       # これ以内にクリニックがある薬局は門前と自動判定(後で目視修正可)
    menkata_main_rx: float = 15000.0        # 年間実績がこれ以上の薬局はメイン薬局とみなし面競合から除外(0=無効)
    competitor_decay_m: float = 1000.0      # 面競合を候補地からの距離で減衰(exp(-d/λ))。0で減衰なし(全店フラット)


# 店舗形態プリセット（来店回数などの既定値）
FORMAT_PRESETS = {
    "大型GMS/モール": {"visit_freq": 4.0, "r65": 0.28},
    "食品スーパー":   {"visit_freq": 4.0, "r65": 0.32},
    "ドラッグ路面":   {"visit_freq": 3.0, "r65": 0.35},
    "独立/その他":     {"visit_freq": 4.0, "r65": 0.30},
}


def _pharmacy_attractiveness(ph: "PharmacyFacility", national_avg: float) -> float:
    """既存薬局の引力＝年間実績枚数÷全国平均。不明は1.0（平均並み）とみなす。"""
    rx = getattr(ph, "annual_rx_count", None)
    if rx and rx > 0:
        return max(rx / national_avg, 0.05)
    return 1.0


def _clinic_annual_rx_pool(
    fac: "MedFacility", a: PredictionAssumptions, op_override: Dict[str, float]
) -> float:
    """クリニックが年間に発生させる院外処方の総量（枚）。predict と同じ係数で算出。"""
    ov = op_override.get(facility_key(fac))
    eff_op = float(ov) if (ov is not None and ov > 0) else fac.daily_outpatients
    if not eff_op:
        return 0.0
    if a.annual_days_mode == "weekly" and fac.weekly_op_days:
        days = fac.weekly_op_days * 52.0
    else:
        days = float(a.fixed_annual_days)
    factor = _external_factor(fac.rx_summary, a)
    if fac.is_cosmetic:
        factor *= a.cosmetic_factor
    elif fac.facility_category == "歯科診療所":
        factor *= a.dental_factor
    return eff_op * days * factor * a.issue_rate


def compute_huff_prediction(
    med_facs: List["MedFacility"],
    pharmacies: List["PharmacyFacility"],
    cand_lat: float,
    cand_lon: float,
    a: PredictionAssumptions,
    hp: HuffParams,
    op_override: Optional[Dict[str, float]] = None,
) -> Dict[str, object]:
    """
    ①医療機関ベース（ハフ）：各クリニックの年間院外処方プールを、周辺の全薬局へ
    「引力×距離減衰」で按分し、候補店の取り分だけを合算する。
        取り分率_自店 = A_自店·w(d_自店) / Σ_全薬局 A_k·w(d_k),  w(d)=exp(−d/λ)·門前boost
    全薬局の取り分は合計1（＝クリニックの総処方箋数）に保存され、独占・二重計上が起きない。
    """
    op_override = op_override or {}
    comps: List[Tuple[float, float, float]] = []
    for p in pharmacies:
        if p.lat is None or p.lon is None:
            continue
        A = _pharmacy_attractiveness(p, hp.national_avg_rx) if hp.weight_by_power else 1.0
        comps.append((p.lat, p.lon, A))

    def bw(d: float, A: float) -> float:
        val = math.exp(-d / hp.lambda_m)
        if d <= hp.monzen_radius:
            val *= hp.monzen_boost
        return A * val

    total = 0.0
    rows: List[Dict[str, object]] = []
    for f in med_facs:
        if f.lat is None or f.lon is None:
            continue
        if not getattr(f, "in_area", True):
            continue
        d_self = haversine(cand_lat, cand_lon, f.lat, f.lon)
        if d_self > hp.reach_m:
            continue
        pool = _clinic_annual_rx_pool(f, a, op_override)
        if pool <= 0:
            continue
        num = bw(d_self, hp.candidate_attractiveness)
        den = num
        for (plat, plon, A) in comps:
            dk = haversine(plat, plon, f.lat, f.lon)
            if dk <= hp.reach_m:
                den += bw(dk, A)
        if den <= 0:
            continue
        share = num / den
        cap = pool * share
        total += cap
        rows.append({
            "clinic": f.name,
            "dist_m": d_self,
            "pool": pool,
            "share": share,
            "captured": cap,
        })
    rows.sort(key=lambda r: r["captured"], reverse=True)
    return {"total": total, "rows": rows, "n_competitors": len(comps)}


def pharmacy_key(p: "PharmacyFacility") -> str:
    """薬局を一意に識別するキー（面/門前の手動修正の保存キーに使用）。"""
    return p.kikan_cd or f"name:{p.name}"


def classify_menkata(
    pharmacies: List["PharmacyFacility"],
    med_facs: List["MedFacility"],
    cand_lat: float,
    cand_lon: float,
    monzen_dist: float = 50.0,
    main_rx_threshold: float = 15000.0,
    reach_m: float = 3000.0,
) -> List[Dict[str, object]]:
    """
    候補地の商圏（reach_m 円）内の各薬局を「面／門前」に自動判定する（目視修正の土台）。
    自動で門前とみなす条件：最寄りクリニックが monzen_dist 以内、または年間実績が
    main_rx_threshold 以上（特定クリニックのメイン薬局。0で無効）。
    戻り値は薬局ごとの情報dict（key/name/候補地距離/最寄りクリニック距離/実績/自動=面か）。
    """
    facs = [f for f in med_facs if f.lat is not None and f.lon is not None]
    out: List[Dict[str, object]] = []
    for p in pharmacies:
        if p.lat is None or p.lon is None:
            continue
        d_cand = haversine(cand_lat, cand_lon, p.lat, p.lon)
        if d_cand > reach_m:
            continue
        nd = min((haversine(p.lat, p.lon, f.lat, f.lon) for f in facs), default=9e9)
        rx = getattr(p, "annual_rx_count", None)
        is_monzen = (nd <= monzen_dist) or bool(main_rx_threshold and rx and rx >= main_rx_threshold)
        out.append({
            "key": pharmacy_key(p), "name": p.name, "d_cand": d_cand,
            "nearest_clinic": nd, "rx": rx, "auto_menkata": (not is_monzen),
        })
    out.sort(key=lambda r: r["d_cand"])
    return out


def footfall_competitor_power(
    classified: List[Dict[str, object]],
    menkata_override: Optional[Dict[str, bool]] = None,
    competitor_decay_m: float = 1000.0,
    national_avg: float = 12000.0,
) -> Tuple[float, int, int]:
    """
    集客ベース②のシェア分母。classify_menkata の結果に手動修正（menkata_override: key→面か）を
    重ね、面と判定された薬局だけを「年間枚数÷全国平均（引力）× 候補地からの距離減衰 exp(-d/λ)」で
    重み付けして合計する（遠い面薬局はスーパー客の選択肢に入りにくいので実効競合を減らす）。
    戻り値: (面競合パワー合計, 面競合店数, 門前扱いで除外した店数)
    """
    menkata_override = menkata_override or {}
    power = 0.0
    n = 0
    excluded = 0
    for r in classified:
        is_menkata = menkata_override.get(r["key"], r["auto_menkata"])
        if not is_menkata:
            excluded += 1
            continue
        rx = r["rx"]
        base = (rx / national_avg) if (rx and rx > 0) else 1.0
        w = (math.exp(-r["d_cand"] / competitor_decay_m)
             if competitor_decay_m and competitor_decay_m > 0 else 1.0)
        power += base * w
        n += 1
    return power, n, excluded


def compute_footfall_prediction(
    fp: FootfallParams, competitor_power: float
) -> Optional[Dict[str, float]]:
    """
    ②集客ベース：スーパー来店客のうち何割が処方箋を持ち込むかで枚数を見積もる。
    いただいた式（商圏人口は分母分子で消える）を年齢2区分に拡張：
        獲得 = [Σ_年齢(ユニーク客数_age × 月受診数_age × 12)] × 発行率 × 院外率
               × 当該薬局利用率 ÷ (競合面薬局パワー + 1)
    """
    if not fp.enabled or fp.unique_customers_monthly <= 0:
        return None
    u = fp.unique_customers_monthly
    u65 = u * fp.ratio_65plus
    u_under = u * (1.0 - fp.ratio_65plus)
    annual_visits = (u65 * fp.visits_month_65plus
                     + u_under * fp.visits_month_under65) * 12.0
    rx_pool = annual_visits * fp.issue_rate * fp.external_rate
    denom = competitor_power + 1.0
    total = rx_pool * fp.use_rate / denom if denom > 0 else 0.0
    return {
        "total": total, "annual_visits": annual_visits, "rx_pool": rx_pool,
        "denom": denom, "u65": u65, "u_under": u_under, "share": fp.use_rate / denom,
    }


OVERPASS_MIRRORS = [
    "https://overpass.kumi.systems/api/interpreter",
    "https://overpass.osm.ch/api/interpreter",
    "https://overpass.openstreetmap.ru/api/interpreter",
]

# kikanCd 先頭桁 → kikanKbn の推定マッピング
KIKAN_KBN_MAP = {"1": [1, 2], "2": [2, 1], "3": [3, 2], "4": [4, 2], "5": [5, 2]}

OSM_SPECIALTY_MAP: Dict[str, str] = {
    "general": "一般内科", "general_practitioner": "一般内科",
    "internal_medicine": "一般内科", "internal": "一般内科",
    "cardiology": "循環器内科", "gastroenterology": "消化器内科",
    "diabetes": "糖尿病内科", "endocrinology": "糖尿病内科",
    "neurology": "神経内科", "pulmonology": "呼吸器内科",
    "surgery": "外科", "orthopaedics": "整形外科", "orthopedics": "整形外科",
    "dermatology": "皮膚科", "ophthalmology": "眼科",
    "otolaryngology": "耳鼻咽喉科", "ent": "耳鼻咽喉科",
    "psychiatry": "精神科", "mental_health": "精神科",
    "paediatrics": "小児科", "pediatrics": "小児科",
    "gynaecology": "産婦人科", "obstetrics": "産婦人科",
    "urology": "泌尿器科", "rehabilitation": "リハビリ科",
    "dentist": "歯科", "dental": "歯科",
}


# ─── データクラス ──────────────────────────────────────────────────────────────
@dataclass
class MedFacility:
    name: str
    address: str = ""
    href: str = ""
    pref_cd: str = ""
    kikan_cd: str = ""
    kikan_kbn: int = 2
    lat: Optional[float] = None
    lon: Optional[float] = None
    distance_m: Optional[float] = None
    source: str = "osm"
    inhouse_rx: str = "—"
    outpatient_rx: str = "—"
    rx_summary: str = "不明"
    daily_outpatients: Optional[int] = None
    daily_outpatients_source: str = "—"
    weekly_op_days: Optional[float] = None
    specialties: str = ""
    facility_category: str = "診療所"
    detail_fetched: bool = False
    detail_url: str = ""
    distance_note: str = ""
    raw_fields: Dict[str, str] = field(default_factory=dict)
    # ── 処方箋獲得予測（compute_capture_prediction が書き戻す） ──
    outpatient_manual: bool = False        # 外来患者数を手動上書きしたか
    annual_op_days_used: Optional[float] = None
    annual_op_visits: Optional[float] = None
    external_rx_factor: float = 1.0
    inflow_rate: float = 0.0
    inflow_band: str = ""
    captured_rx: Optional[float] = None
    # ── 門前占有チェック（既存の門前薬局に椅子を取られていないか） ──
    nearest_pharmacy_dist_m: Optional[float] = None
    nearest_pharmacy_name: str = ""
    monzen_occupied: bool = False          # このクリニックの門前(≤50m)に既存薬局あり
    monzen_contested: bool = False         # かつ候補地も門前バンド＝獲得が競合する
    # ── データ品質検証（get_facility_detail が書き込む） ──
    beds: Optional[int] = None             # 届出/許可病床数（合計）
    is_cosmetic: bool = False              # 美容・自由診療らしき施設（名称/診療科から判定）
    op_flag: str = ""                      # 外来患者数の異常値フラグ（空=正常）
    op_suggested: Optional[int] = None     # 年間値入力疑い時の補正候補（÷305）
    coord_source: str = ""                 # 座標の出典（ナビィ埋込 or ジオコーディング）
    # ── 商圏ポリゴン判定（apply_area_flags が書き込む） ──
    in_area: bool = True                   # 商圏内か（円形モードは常にTrue）


@dataclass
class PharmacyFacility:
    name: str
    address: str
    href: str = ""
    pref_cd: str = ""
    kikan_cd: str = ""
    lat: Optional[float] = None
    lon: Optional[float] = None
    distance_m: Optional[float] = None
    source: str = "mhlw"
    pharmacy_type: str = "不明"
    nearest_clinic_name: str = "—"
    nearest_clinic_dist_m: Optional[float] = None
    annual_rx_count: Optional[int] = None
    annual_rx_source: str = "—"
    detail_fetched: bool = False
    detail_url: str = ""
    raw_fields: Dict[str, str] = field(default_factory=dict)
    in_area: bool = True                   # 商圏ポリゴン内か（円形モードは常にTrue）


# ─── ユーティリティ ────────────────────────────────────────────────────────────
def haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    R = 6_371_000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def name_similarity(a: str, b: str) -> float:
    a_chars = set(re.sub(r"[　\s・（）()「」]", "", a))
    b_chars = set(re.sub(r"[　\s・（）()「」]", "", b))
    if not a_chars or not b_chars:
        return 0.0
    return len(a_chars & b_chars) / max(len(a_chars), len(b_chars))


def guess_kikan_kbn(kikan_cd: str) -> List[int]:
    prefix = kikan_cd[0] if kikan_cd else "2"
    return KIKAN_KBN_MAP.get(prefix, [2, 1, 3])


# ─── 商圏ポリゴン（手描きエリア）ユーティリティ ─────────────────────────────────
# ポリゴンは [(lat, lon), ...] の頂点リスト。複数ポリゴン＝リストのリストで持つ。

def point_in_polygon(lat: float, lon: float, poly: List[Tuple[float, float]]) -> bool:
    """レイキャスティング法による内外判定（商圏スケールでは平面近似で十分）。"""
    if len(poly) < 3:
        return False
    x, y = lon, lat
    inside = False
    j = len(poly) - 1
    for i in range(len(poly)):
        yi, xi = poly[i][0], poly[i][1]
        yj, xj = poly[j][0], poly[j][1]
        if (yi > y) != (yj > y):
            x_cross = (xj - xi) * (y - yi) / (yj - yi + 1e-12) + xi
            if x < x_cross:
                inside = not inside
        j = i
    return inside


def point_in_any_polygon(
    lat: Optional[float], lon: Optional[float],
    polygons: List[List[Tuple[float, float]]],
) -> bool:
    if lat is None or lon is None:
        return False
    return any(point_in_polygon(lat, lon, p) for p in polygons)


def polygons_from_map_output(map_output: Optional[dict]) -> List[List[Tuple[float, float]]]:
    """st_folium の戻り値（all_drawings の GeoJSON）から頂点リスト群を取り出す。"""
    polys: List[List[Tuple[float, float]]] = []
    if not map_output:
        return polys
    for feat in (map_output.get("all_drawings") or []):
        try:
            geom = feat.get("geometry", {})
            if geom.get("type") != "Polygon":
                continue
            ring = geom.get("coordinates", [[]])[0]  # GeoJSONは[lon, lat]順
            pts = [(float(c[1]), float(c[0])) for c in ring if len(c) >= 2]
            if len(pts) >= 3:
                polys.append(pts)
        except (TypeError, ValueError, IndexError):
            continue
    return polys


def polygon_max_radius_m(
    center_lat: float, center_lon: float,
    polygons: List[List[Tuple[float, float]]],
) -> float:
    """住所（候補地）から全ポリゴン頂点への最大距離＝収集円の半径。"""
    dmax = 0.0
    for poly in polygons:
        for lat, lon in poly:
            dmax = max(dmax, haversine(center_lat, center_lon, lat, lon))
    return dmax


def polygons_area_km2(polygons: List[List[Tuple[float, float]]]) -> float:
    """ポリゴン群の概算面積（km²）。重心緯度での正距円筒近似＋靴紐公式。"""
    total = 0.0
    for poly in polygons:
        if len(poly) < 3:
            continue
        lat0 = sum(p[0] for p in poly) / len(poly)
        k_lat = 111_320.0                                   # 1度あたりm（南北）
        k_lon = 111_320.0 * math.cos(math.radians(lat0))    # 1度あたりm（東西）
        pts = [((lon * k_lon), (lat * k_lat)) for lat, lon in poly]
        s = 0.0
        j = len(pts) - 1
        for i in range(len(pts)):
            s += pts[j][0] * pts[i][1] - pts[i][0] * pts[j][1]
            j = i
        total += abs(s) / 2.0
    return total / 1_000_000.0


def apply_area_flags(
    med_facs: List["MedFacility"],
    pharmacies: List["PharmacyFacility"],
    polygons: List[List[Tuple[float, float]]],
    exclude_med_outside: bool,
) -> Tuple[int, int]:
    """
    ポリゴンで商圏内外フラグを付け直す。(圏外医療機関数, 圏外薬局数) を返す。
    ポリゴン未指定（円形モード）なら全て圏内。
    座標なしの施設は判定不能のため圏内扱い（予測には距離が必要なので実害なし）。
    """
    if not polygons:
        for f in med_facs:
            f.in_area = True
        for p in pharmacies:
            p.in_area = True
        return 0, 0
    n_med_out = n_ph_out = 0
    for f in med_facs:
        if f.lat is None or f.lon is None:
            f.in_area = True
            continue
        inside = point_in_any_polygon(f.lat, f.lon, polygons)
        f.in_area = inside if exclude_med_outside else True
        if not inside:
            n_med_out += 1
    for p in pharmacies:
        if p.lat is None or p.lon is None:
            p.in_area = True
            continue
        p.in_area = point_in_any_polygon(p.lat, p.lon, polygons)
        if not p.in_area:
            n_ph_out += 1
    return n_med_out, n_ph_out


# ─── Overpass ─────────────────────────────────────────────────────────────────
def _overpass_post(query: str, timeout: int = 40, retries: int = 2) -> Optional[dict]:
    for attempt in range(retries + 1):
        for url in OVERPASS_MIRRORS:
            try:
                r = requests.post(url, data={"data": query}, timeout=timeout)
                if r.status_code == 200:
                    return r.json()
                if r.status_code in (429, 503):
                    time.sleep(5 + attempt * 5)
                    break
            except requests.exceptions.Timeout:
                continue
            except Exception:
                continue
        if attempt < retries:
            time.sleep(3 + attempt * 3)
    return None


# ─── OSM 検索 ─────────────────────────────────────────────────────────────────
def _parse_osm_pharmacy_elements(
    elements: list, center_lat: float, center_lon: float
) -> List[PharmacyFacility]:
    pharmacies: List[PharmacyFacility] = []
    seen_ids: set = set()
    for el in elements:
        el_id = el.get("id")
        if el_id in seen_ids:
            continue
        seen_ids.add(el_id)
        tags = el.get("tags", {})
        name = tags.get("name:ja") or tags.get("name", "")
        branch = tags.get("branch", "")
        if branch and branch not in name:
            name = f"{name}{branch}"
        if not name:
            continue
        if el["type"] == "node":
            f_lat, f_lon = el.get("lat"), el.get("lon")
        else:
            center = el.get("center", {})
            f_lat, f_lon = center.get("lat"), center.get("lon")
        if f_lat is None or f_lon is None:
            continue
        addr_parts = [
            tags.get("addr:province", ""),
            tags.get("addr:city", "") or tags.get("addr:district", ""),
            tags.get("addr:suburb", ""),
            tags.get("addr:housenumber", ""),
        ]
        address = re.sub(r"\s+", "", "".join(p for p in addr_parts if p))
        dist = haversine(center_lat, center_lon, f_lat, f_lon)
        pharmacies.append(PharmacyFacility(
            name=name, address=address, source="osm",
            lat=f_lat, lon=f_lon, distance_m=dist,
        ))
    return pharmacies


def search_osm_pharmacies(lat: float, lon: float, radius_m: int) -> List[PharmacyFacility]:
    query = f"""
[out:json][timeout:50];
(
  node["amenity"="pharmacy"](around:{radius_m},{lat},{lon});
  way["amenity"="pharmacy"](around:{radius_m},{lat},{lon});
  node["shop"="chemist"](around:{radius_m},{lat},{lon});
  way["shop"="chemist"](around:{radius_m},{lat},{lon});
);
out center;
"""
    data = _overpass_post(query)
    if not data:
        return []
    result = _parse_osm_pharmacy_elements(data.get("elements", []), lat, lon)
    result.sort(key=lambda x: x.distance_m or 9_999_999)
    return result


def search_osm_medical(lat: float, lon: float, radius_m: int) -> List[MedFacility]:
    query = f"""
[out:json][timeout:40];
(
  node["amenity"~"^(clinic|hospital|doctors|medical_centre)$"](around:{radius_m},{lat},{lon});
  way["amenity"~"^(clinic|hospital|doctors|medical_centre)$"](around:{radius_m},{lat},{lon});
  node["healthcare"]["healthcare"!~"^(pharmacy|chemist|dispensary|yes)$"](around:{radius_m},{lat},{lon});
  way["healthcare"]["healthcare"!~"^(pharmacy|chemist|dispensary|yes)$"](around:{radius_m},{lat},{lon});
);
out center;
"""
    # 薬局キーワードフィルター（名前ベース）
    _PHARMA_RE = re.compile(
        r'薬局|ドラッグ|ファーマ|調剤|くすり|クスリ|drug\s*store|pharmacy', re.IGNORECASE
    )

    data = _overpass_post(query)
    if not data:
        return []
    facilities: List[MedFacility] = []
    for el in data.get("elements", []):
        tags = el.get("tags", {})
        name = tags.get("name:ja") or tags.get("name", "")
        if not name:
            continue
        # 薬局・ドラッグストアを除外
        if _PHARMA_RE.search(name):
            continue
        if el["type"] == "node":
            f_lat, f_lon = el.get("lat"), el.get("lon")
        else:
            c = el.get("center", {})
            f_lat, f_lon = c.get("lat"), c.get("lon")
        if f_lat is None or f_lon is None:
            continue
        sp_en = (tags.get("healthcare:speciality", "")
                 or tags.get("speciality", "")
                 or tags.get("medical_system:western", "")).lower()
        sp_ja = OSM_SPECIALTY_MAP.get(sp_en, "")
        amenity = tags.get("amenity", "")
        healthcare = tags.get("healthcare", "")
        if amenity == "hospital" or healthcare == "hospital":
            cat = "病院"
        elif healthcare == "dentist" or "dentist" in sp_en:
            cat = "診療所"
            if not sp_ja:
                sp_ja = "歯科"
        else:
            cat = "診療所"

        addr_parts = [
            tags.get("addr:province", ""),
            tags.get("addr:city", "") or tags.get("addr:district", ""),
            tags.get("addr:suburb", ""),
            tags.get("addr:housenumber", ""),
        ]
        address = re.sub(r"\s+", "", "".join(p for p in addr_parts if p))
        dist = haversine(lat, lon, f_lat, f_lon)
        facilities.append(MedFacility(
            name=name, address=address, source="osm",
            lat=f_lat, lon=f_lon, distance_m=dist,
            specialties=sp_ja, facility_category=cat,
        ))
    facilities.sort(key=lambda x: x.distance_m or 9_999_999)
    return facilities


# ─── ジオコーダー ──────────────────────────────────────────────────────────────
class GeocoderService:
    GSI_URL       = "https://msearch.gsi.go.jp/address-search/AddressSearch"
    NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
    HEADERS       = {"User-Agent": "AreaAnalysisTool/1.0"}
    LAT_MIN, LAT_MAX = 24.0, 46.0
    LON_MIN, LON_MAX = 122.0, 154.0

    def _is_japan(self, lat, lon) -> bool:
        return self.LAT_MIN <= lat <= self.LAT_MAX and self.LON_MIN <= lon <= self.LON_MAX

    def _clean(self, address: str) -> str:
        a = re.sub(r"〒\s*\d{3}[-−]\d{4}\s*", "", address)
        a = re.sub(r"Googleマップ.*", "", a).strip()
        trans = str.maketrans(
            "０１２３４５６７８９ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ"
            "ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ－−‐",
            "0123456789abcdefghijklmnopqrstuvwxyz"
            "ABCDEFGHIJKLMNOPQRSTUVWXYZ---",
        )
        a = a.translate(trans).replace("　", " ")
        a = re.sub(r"(\d+)\s*丁目\s*(\d+)\s*番地?\s*(\d+)\s*号?", r"\1-\2-\3", a)
        a = re.sub(r"(\d+)\s*丁目\s*(\d+)\s*番地?", r"\1-\2", a)
        a = re.sub(r"(\d+)\s*番地?\s*(\d+)\s*号", r"\1-\2", a)
        a = re.sub(r"(\d+)\s*番地", r"\1", a)
        return re.sub(r"\s+", " ", a).strip()

    def _shorten(self, address: str) -> str:
        a = re.sub(r"(\d+(?:[-]\d+)+)\s+[　-鿿＀-￯A-Za-z].+$", r"\1", address)
        if a != address:
            return a.strip()
        a = re.sub(r"\s*\d+\s*(?:階|[Ff]|号室|番地)\b.*$", "", address)
        a = re.sub(r"\s+[゠-ヿ]{3,}.*$", "", a)
        return a.strip()

    def _gsi(self, q: str) -> Optional[Tuple[float, float]]:
        try:
            r = requests.get(self.GSI_URL, params={"q": q}, headers=self.HEADERS, timeout=6)
            if r.status_code == 200:
                data = r.json()
                if data:
                    coords = data[0].get("geometry", {}).get("coordinates", [])
                    if len(coords) == 2:
                        lon, lat = float(coords[0]), float(coords[1])
                        if self._is_japan(lat, lon):
                            return lat, lon
        except Exception:
            pass
        return None

    def _nominatim(self, q: str) -> Optional[Tuple[float, float]]:
        try:
            r = requests.get(
                self.NOMINATIM_URL,
                params={"q": q + " 日本", "format": "json", "limit": 1},
                headers=self.HEADERS, timeout=8,
            )
            if r.status_code == 200:
                data = r.json()
                if data:
                    lat, lon = float(data[0]["lat"]), float(data[0]["lon"])
                    if self._is_japan(lat, lon):
                        return lat, lon
        except Exception:
            pass
        return None

    def geocode(self, address: str) -> Optional[Tuple[float, float]]:
        clean = self._clean(address)
        short = self._shorten(clean)
        has_short = short and short != clean
        result = self._gsi(clean)
        if result:
            return result
        if has_short:
            result = self._gsi(short)
            if result:
                return result
        time.sleep(1.0)
        result = self._nominatim(clean)
        if result:
            return result
        if has_short:
            time.sleep(0.5)
            result = self._nominatim(short)
            if result:
                return result
        return None

    def geocode_with_verification(
        self, address: str
    ) -> Tuple[Optional[Tuple[float, float]], str]:
        """GSI と Nominatim を両方試して結果を比較し、距離ノートを返す。"""
        clean = self._clean(address)
        short = self._shorten(clean)
        has_short = short and short != clean
        gsi_result = self._gsi(clean) or (self._gsi(short) if has_short else None)
        time.sleep(0.8)
        nom_result = self._nominatim(clean) or (self._nominatim(short) if has_short else None)
        if gsi_result and nom_result:
            diff = haversine(gsi_result[0], gsi_result[1], nom_result[0], nom_result[1])
            if diff <= 150:
                return gsi_result, f"確認済({diff:.0f}m差)"
            elif diff <= 400:
                return gsi_result, f"中程度({diff:.0f}m差)"
            else:
                return gsi_result, f"要確認({diff:.0f}m差・目視推奨)"
        elif gsi_result:
            return gsi_result, "GSIのみ取得"
        elif nom_result:
            return nom_result, "Nominatimのみ取得"
        return None, "取得失敗"

    def geocode_by_name(
        self, name: str, near_lat: float, near_lon: float, radius_km: float = 25
    ) -> Optional[Tuple[float, float]]:
        """施設名でNominatimをバウンディングボックス付き検索（住所不明時のフォールバック用）。"""
        delta = radius_km / 111.0
        viewbox = f"{near_lon - delta},{near_lat + delta},{near_lon + delta},{near_lat - delta}"
        try:
            r = requests.get(
                self.NOMINATIM_URL,
                params={
                    "q": name,
                    "format": "json",
                    "limit": 3,
                    "countrycodes": "jp",
                    "viewbox": viewbox,
                    "bounded": 1,
                },
                headers=self.HEADERS,
                timeout=8,
            )
            if r.status_code == 200:
                for item in r.json():
                    lat, lon = float(item["lat"]), float(item["lon"])
                    if self._is_japan(lat, lon):
                        return lat, lon
        except Exception:
            pass
        return None


# ─── フィールドパーサ群 ────────────────────────────────────────────────────────
def _get_field(fields: Dict[str, str], keys: List[str]) -> Optional[str]:
    for k in keys:
        if k in fields:
            return fields[k]
    for k in keys:
        for fk, fv in fields.items():
            if k in fk:
                return fv
    return None


def _infer_rx_type(full_text: str) -> str:
    lines = [line.strip() for line in full_text.split("\n") if line.strip()]
    for line in lines:
        if "院外処方" in line:
            if any(w in line for w in ["有り", "有", "あり", "可能", "実施"]):
                return "院外処方あり"
            if any(w in line for w in ["無し", "無", "なし", "不可"]):
                return "院内処方のみ"
        if "処方せん" in line or "処方箋" in line:
            if any(w in line for w in ["交付", "発行", "有"]):
                return "院外処方あり"
        if "院内処方" in line:
            if any(w in line for w in ["有り", "有", "あり"]):
                return "院内処方のみ"
    if "院外処方" in full_text or "処方せんの交付" in full_text:
        return "院外処方あり（推定）"
    return "不明"


def _extract_outpatient_from_table(table) -> Optional[int]:
    rows = table.find_all("tr")
    if not rows:
        return None
    header_cells = rows[0].find_all(["th", "td"])
    header_texts = [re.sub(r"\s+", "", c.get_text(strip=True)) for c in header_cells]
    gairaikan_col: Optional[int] = None
    for i, h in enumerate(header_texts):
        if h.startswith("外来患者") and "月平均" not in h:
            gairaikan_col = i
            break
    if gairaikan_col is not None and len(rows) >= 2:
        for row in rows[1:]:
            cells = row.find_all(["th", "td"])
            if not cells:
                continue
            row_label = cells[0].get_text(strip=True)
            if "前年度" not in row_label:
                continue
            if gairaikan_col >= len(cells):
                continue
            val = cells[gairaikan_col].get_text(strip=True)
            if not val or re.fullmatch(r"[－\-−—―\s]*", val):
                continue
            m = re.search(r"(\d+\.?\d*)", val)
            if m:
                n = float(m.group(1))
                if 0 < n <= 10_000:
                    return int(round(n))
    outpatient_row_kw = ["外来患者", "外来数", "１日平均", "1日平均", "日平均外来"]
    for row in rows:
        cells = row.find_all(["th", "td"])
        if len(cells) < 2:
            continue
        label = re.sub(r"\s+", "", cells[0].get_text(strip=True))
        if not any(kw in label for kw in outpatient_row_kw):
            continue
        if any(ex in label for ex in ["紹介", "入院", "在宅"]):
            continue
        # 多列行（label + 入院6列 + 外来 + 歯科 = 9セル以上）は外来列(cells[7])を指定。
        # 先頭から数値を拾うと入院列を誤取得するため（2026-07 実データ検証で確認）。
        scan_cells = [cells[7]] if len(cells) >= 9 else cells[1:]
        for cell in scan_cells:
            val = cell.get_text(strip=True)
            m = re.search(r"(\d+\.?\d*)", val)
            if m:
                n = float(m.group(1))
                if 0 < n <= 10_000:
                    return int(round(n))
    # 注: 旧実装にあった「任意の前年度行から最初の数値を拾う」汎用フォールバックは
    # 入院列を誤取得するリスクがあるため削除（同じ行は raw_fields 経由の
    # 「前年度フィールド・外来列」パスが正しい列位置で処理する）。
    return None


def _parse_outpatients_from_stats_table(soup: BeautifulSoup) -> Optional[int]:
    for item_div in soup.find_all("div", class_="item"):
        h3 = item_div.find("h3")
        if not h3 or "患者数" not in h3.get_text(strip=True):
            continue
        for table in item_div.find_all("table"):
            result = _extract_outpatient_from_table(table)
            if result is not None:
                return result
    for data_th in soup.find_all("th", class_="ptn4ItemName"):
        if "前年度" not in data_th.get_text(strip=True):
            continue
        table = data_th.find_parent("table")
        if table is None:
            continue
        result = _extract_outpatient_from_table(table)
        if result is not None:
            return result
    for table in soup.find_all("table"):
        result = _extract_outpatient_from_table(table)
        if result is not None:
            return result
    return None


def _parse_daily_outpatients(
    fields: Dict[str, str],
    full_text: str,
    soup: Optional[BeautifulSoup] = None,
) -> Tuple[Optional[int], str]:
    if soup is not None:
        result = _parse_outpatients_from_stats_table(soup)
        if result is not None:
            return result, "ナビィ（実績統計表）"
    v_zennen = fields.get("前年度１日平均患者数", "")
    if v_zennen:
        # 「前年度１日平均患者数」は多列（例: 396.3人/-/-/-/-/-/824人/-）。
        # 列は [入院各病床…, 外来, 歯科] の順で、外来は index=6（7列目）。
        # 旧実装の patient_vals[-2] はダッシュ列が findall で欠落すると
        # 入院列を誤取得したため、列位置を保持して外来列を選ぶよう修正。
        cells = [c.strip() for c in v_zennen.split("/")]
        def _num(cell: str) -> Optional[float]:
            m = re.search(r"(\d+\.?\d*)", cell)
            if not m:
                return None
            try:
                x = float(m.group(1))
            except ValueError:
                return None
            return x if 0 < x <= 10_000 else None
        # ① 外来列（index 6）を最優先
        if len(cells) >= 7:
            n = _num(cells[6])
            if n is not None:
                return int(round(n)), "ナビィ（前年度フィールド・外来列）"
        # ② 列数が想定外の場合は、数値を持つ末尾寄りの列を外来とみなす
        numeric = [(i, _num(c)) for i, c in enumerate(cells)]
        numeric = [(i, x) for i, x in numeric if x is not None]
        if numeric:
            # 歯科（最終列）を避けつつ、外来に相当する後方の列を採用
            idx, val = numeric[-1] if len(numeric) == 1 else numeric[-2] if len(cells) == len(numeric) else numeric[-1]
            if val is not None:
                return int(round(val)), "ナビィ（前年度フィールド）"
    candidate_keys = [
        "1日あたりの外来患者の平均数", "外来患者の平均数", "1日平均外来患者数",
        "一日平均外来患者数", "外来患者数（1日平均）", "前年度の１日平均外来患者数",
        "前年度１日平均外来患者数", "１日平均外来患者数", "1日あたり外来患者数",
        "外来（1日平均）", "外来患者の延数", "外来患者数",
    ]
    for k_target in candidate_keys:
        # 部分一致で「紹介を受けた外来患者数（月平均）」等の別指標を誤取得しないよう
        # 完全一致 → 除外語なしの部分一致 の順で自前ルックアップする
        v = fields.get(k_target)
        if v is None:
            for fk, fv in fields.items():
                if k_target in fk and not any(ng in fk for ng in ("紹介", "月平均", "在宅", "入院")):
                    v = fv
                    break
        if v:
            m_person = re.search(r"(\d+\.?\d*)\s*人", v)
            if m_person:
                n = float(m_person.group(1))
                if 0 < n <= 10_000:
                    return int(round(n)), f"ナビィ（{k_target}）"
            nums = re.findall(r"[0-9,]+\.?[0-9]*", v)
            for n_str in nums:
                try:
                    n = float(n_str.replace(",", ""))
                    if 1 <= n <= 3_000:
                        return int(round(n)), f"ナビィ（{k_target}）"
                    if n > 3_000:
                        return max(1, int(n / WORKING_DAYS)), f"ナビィ（{k_target}・年間÷305）"
                except ValueError:
                    pass
    # 注: \d は全角数字（例: ツールチップ内「１日平均」の「１」）にもマッチし
    # 誤って外来=1人を返していたため、テキスト解析は半角 [0-9] に限定する。
    for pat, label in [
        (r"1日あたりの外来患者の平均数[^0-9]{0,20}([0-9]{1,4})", "ナビィ（テキスト解析）"),
        (r"前年度の?１?日平均外来患者数[^0-9]{0,15}([0-9]{1,4})", "ナビィ（テキスト解析）"),
        (r"外来患者の平均数[^0-9]{0,15}([0-9]{1,4})", "ナビィ（テキスト解析）"),
        (r"1日平均外来患者数[^0-9]{0,15}([0-9]{1,4})", "ナビィ（テキスト解析）"),
        (r"外来患者[^0-9]{0,10}1日平均[^0-9]{0,10}([0-9]{1,4})", "ナビィ（テキスト解析）"),
        (r"外来[^0-9]{0,8}([0-9]{1,3})\s*人[/／]日", "ナビィ（テキスト解析）"),
        (r"前年度[^0-9]{0,20}外来[^0-9]{0,10}([0-9]{1,4}\.?[0-9]*)\s*人", "ナビィ（テキスト解析）"),
    ]:
        m = re.search(pat, full_text)
        if m:
            try:
                n = float(m.group(1).replace(",", ""))
                if 1 <= n <= 3_000:
                    return int(round(n)), label
            except ValueError:
                pass
    return None, "—"


def _parse_weekly_days(
    fields: Dict[str, str],
    full_text: str,
    soup: Optional[BeautifulSoup],
) -> Optional[float]:
    candidate_keys = [
        "週の診療日数", "週診療日数", "週あたり診療日数",
        "診療日数（週）", "平均診療日（週）", "診療日（週平均）",
    ]
    v = _get_field(fields, candidate_keys)
    if v:
        m = re.search(r"(\d+\.?\d*)\s*日", v)
        if m:
            n = float(m.group(1))
            if 0.5 <= n <= 7:
                return n
    schedule_keys = [
        "診療時間（診療科目別の）", "診療科目別の診療時間",
        "外来受付時間（診療科目別の）", "診療時間帯",
    ]
    sv = _get_field(fields, schedule_keys)
    if sv and "/" in sv:
        parts = [p.strip() for p in sv.split("/")]
        day_indices = set()
        for i, p in enumerate(parts[:8]):
            if p and p not in ["-", "−", "—", "休", "×"] and re.search(r"\d{1,2}:\d{2}", p):
                day_indices.add(i)
        if day_indices:
            return float(len(day_indices))
    if soup:
        days = _count_open_days_from_hours_table(soup)
        if days:
            return float(days)
    for pat in [
        r"週\s*(\d+\.?\d*)\s*日",
        r"週に平均\s*(\d+\.?\d*)\s*日",
        r"(\d+\.?\d*)\s*日[／/]週",
    ]:
        m = re.search(pat, full_text)
        if m:
            n = float(m.group(1))
            if 0.5 <= n <= 7:
                return n
    return None


def _count_open_days_from_hours_table(soup: BeautifulSoup) -> Optional[int]:
    WEEKDAY_CHARS = ["月", "火", "水", "木", "金", "土", "日"]
    open_days: set = set()
    for table in soup.find_all("table"):
        headers = []
        first_row = table.find("tr")
        if first_row:
            cells = first_row.find_all(["th", "td"])
            headers = [c.get_text(strip=True) for c in cells]
        header_days = []
        for i, h in enumerate(headers):
            for wd in WEEKDAY_CHARS:
                if wd in h:
                    header_days.append((i, wd))
        if header_days:
            for row in table.find_all("tr")[1:]:
                cells = row.find_all(["th", "td"])
                for col_i, wd in header_days:
                    if col_i < len(cells):
                        v = cells[col_i].get_text(strip=True)
                        if v and v not in ["×", "✗", "−", "-", "休", "—", ""]:
                            if not re.fullmatch(r"[×✗−\-休―‐ー]", v):
                                open_days.add(wd)
        for row in table.find_all("tr"):
            cells = row.find_all(["th", "td"])
            if not cells:
                continue
            header = cells[0].get_text(strip=True)
            for wd in WEEKDAY_CHARS:
                if wd in header:
                    for cell in cells[1:]:
                        v = cell.get_text(strip=True)
                        if re.search(r"\d{1,2}:\d{2}", v):
                            open_days.add(wd)
    return len(open_days) if open_days else None


def _parse_specialties(fields: Dict[str, str], full_text: str) -> str:
    candidate_keys = ["診療科目", "診療科", "標榜診療科", "診療科名"]
    v = _get_field(fields, candidate_keys)
    if v:
        parts = re.split(r"[、。\n/／・]", v)
        sp_list = [p.strip() for p in parts if p.strip() and len(p.strip()) <= 15]
        return "、".join(sp_list[:6])
    return ""


def _parse_annual_rx_count(
    fields: Dict[str, str], full_text: str
) -> Tuple[Optional[int], str]:
    candidate_keys = [
        "処方箋受付回数（年間）", "処方箋受付枚数（年間）",
        "処方箋受付回数", "処方箋受付枚数",
        "調剤処方箋の受付枚数", "取扱処方箋数",
        "総取扱処方箋数", "年間処方箋受付回数",
        "年間処方箋取扱枚数", "処方箋枚数",
    ]
    for k in candidate_keys:
        v = None
        if k in fields:
            v = fields[k]
        else:
            for fk, fv in fields.items():
                if k in fk:
                    v = fv
                    break
        if v:
            m = re.search(r"([0-9,]+)", v)
            if m:
                try:
                    n = int(m.group(1).replace(",", ""))
                    if 100 <= n <= 10_000_000:
                        return n, f"ナビィ（{k}）"
                    if n == 0:
                        # 明示的な「0件」報告（漢方専門・OTC併設・調剤実績なし等）。
                        # 取得失敗(None)と区別して返す＝競合分析・実績照合で除外できる。
                        return 0, "ナビィ（報告0件）"
                except ValueError:
                    pass
    for pat, label in [
        (r"処方箋受付(?:回数|枚数)[^0-9]{0,10}([0-9,]+)", "ナビィ（テキスト解析）"),
        (r"取扱処方箋(?:数|枚)[^0-9]{0,10}([0-9,]+)", "ナビィ（テキスト解析）"),
        (r"処方箋[^0-9]{0,8}([0-9,]+)\s*(?:回|枚|件)", "ナビィ（テキスト解析）"),
    ]:
        m = re.search(pat, full_text)
        if m:
            try:
                n = int(m.group(1).replace(",", ""))
                if 100 <= n <= 10_000_000:
                    return n, label
            except ValueError:
                pass
    return None, "—"


# ─── データ品質検証ヘルパー ─────────────────────────────────────────────────────
def _extract_coords_from_html(html: str) -> Optional[Tuple[float, float]]:
    """
    ナビィ詳細ページに埋め込まれた正確な緯度経度を抽出する（地図リンク q=lat,lon）。
    住所ジオコーディング（誤差±30〜80m）より高精度で、門前50m判定の信頼性が大きく上がる。
    """
    m = re.search(r"q=([2-4][0-9]\.[0-9]{3,}),\s*(1[23][0-9]\.[0-9]{3,})", html)
    if not m:
        m = re.search(r"([2-4][0-9]\.[0-9]{4,})\s*,\s*(1[23][0-9]\.[0-9]{4,})", html)
    if m:
        lat, lon = float(m.group(1)), float(m.group(2))
        if 24.0 <= lat <= 46.0 and 122.0 <= lon <= 154.0:  # 日本国内チェック
            return lat, lon
    return None


def _parse_total_beds(fields: Dict[str, str]) -> Optional[int]:
    """「届出又は許可病床数」フィールドから合計病床数（最終列）を取る。"""
    for k, v in fields.items():
        if k.startswith("届出又は許可病床数"):
            nums = re.findall(r"([0-9]+)床", v)
            if nums:
                return int(nums[-1])
    return None


# 美容・自由診療らしき施設（保険処方箋がほぼ発生しない）の名称/診療科パターン。
# 誤検出を避けるため明白なキーワードに限定（例:「形成外科」単体は保険診療なので含めない）。
_COSMETIC_RE = re.compile(
    r"美容外科|美容皮膚|美容クリニック|美容医療|ＡＧＡ|AGA|スキンクリニック"
    r"|脱毛|アートメイク|植毛|メンズライフ|包茎", re.IGNORECASE
)


def _detect_cosmetic(name: str, specialties: str) -> bool:
    return bool(_COSMETIC_RE.search(name or "") or _COSMETIC_RE.search(specialties or ""))


def _validate_outpatients(fac: "MedFacility") -> Tuple[str, Optional[int]]:
    """
    取得した外来患者数の妥当性を検証し (フラグ文字列, 補正候補値) を返す。空文字=正常。

    実データ検証（2026-07・3エリア46施設）で確認された誤報告パターン:
      - 年間値が1日欄に入力（例: 外来列=13,736人 → 実際は約45人/日）
      - 月間値らしき高値（診療所で700人/日超）
    補正候補は年間値÷305日で算出（週診療日数は誤登録がありうるため固定日数を使う）。
    """
    # raw多列フィールドの外来列（キャップで弾かれた大きな生値も見る）
    raw6: Optional[float] = None
    raw = fac.raw_fields.get("前年度１日平均患者数", "") if fac.raw_fields else ""
    if raw:
        cells = [c.strip() for c in raw.split("/")]
        if len(cells) >= 7:
            m = re.search(r"([0-9]+\.?[0-9]*)", cells[6])
            if m:
                raw6 = float(m.group(1))

    op = fac.daily_outpatients

    # 年間値入力疑い: 生値が2,000超で、÷305が現実的な1日患者数に収まる
    if raw6 is not None and raw6 > 2_000:
        est = raw6 / WORKING_DAYS
        if 5 <= est <= 600:
            return (f"年間値入力疑い（外来列の生値={raw6:,.0f}）", int(round(est)))

    if op is None:
        return ("", None)

    if fac.facility_category == "病院":
        beds = fac.beds or 0
        if op > max(2_500, beds * 6):
            return ("過大値疑い（病床規模と不整合）", None)
        return ("", None)

    # 診療所
    if op >= 1_000:
        est = op / WORKING_DAYS
        if 5 <= est <= 600:
            return ("年間値入力疑い", int(round(est)))
        return ("過大値疑い", None)
    if op >= 700:
        return ("高値・要確認（月間値の可能性）", None)
    return ("", None)


# ─── MHLWスクレイパー ──────────────────────────────────────────────────────────
class MHLWScraper:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "ja-JP,ja;q=0.9",
        })
        self._ready = False

    def _init(self) -> bool:
        if self._ready:
            return True
        try:
            r = self.session.get(f"{MHLW_BASE}/juminkanja/S2320/initialize", timeout=15)
            self._ready = r.status_code == 200
        except Exception:
            self._ready = False
        return self._ready

    def search_pharmacies_by_latlon(
        self,
        lat: float, lon: float,
        radius_m: int,
        center_name: str = "",
        max_pages: int = 8,
    ) -> Tuple[List[PharmacyFacility], str]:
        """ナビィ薬局タブ（S2300/yakkyokuSearch）で薬局を緯度経度検索する。"""
        if not self._init():
            return [], "MHLW接続エラー"
        dist_code = "00" if radius_m <= 1_000 else ("01" if radius_m <= 5_000 else "")
        cn = urllib.parse.quote(center_name or "検索地点")
        try:
            self.session.get(f"{MHLW_BASE}/juminkanja/S2300/initializeYakk", timeout=12)
            r1 = self.session.get(
                f"{MHLW_BASE}/juminkanja/S2300/yakkyokuSearch",
                params={
                    "iyakuKbn": "2", "lang": "ja",
                    "latitude": str(lat), "longitude": str(lon),
                    "distanceFromCenterPoint": dist_code,
                    "centerPointName": cn,
                    "selectCenterPoint": "3",
                    "specifyDateAndTime": "01",
                    "XCHARSET": "utf-8",
                },
                timeout=15,
            )
            j = r1.json()
            if j.get("code") != "0":
                return [], f"薬局ナビィエラー: {j.get('messages')}"
            search_id = j["result"]["id"]
            self.session.get(
                f"{MHLW_BASE}/juminkanja/S2300/yakkyokuSearch",
                params={
                    "id": search_id,
                    "latitude": str(lat), "longitude": str(lon),
                    "distanceFromCenterPoint": dist_code,
                    "selectCenterPoint": "3",
                    "specifyDateAndTime": "01",
                    "XCHARSET": "utf-8",
                },
                timeout=15,
            )
        except Exception as e:
            return [], f"薬局ナビィ例外: {e}"

        all_ph: List[PharmacyFacility] = []
        total = 0
        for page in range(max_pages):
            try:
                r3 = self.session.get(
                    f"{MHLW_BASE}/juminkanja/S2400/initialize",
                    params={"id": search_id, "page": page, "size": 20, "sortNo": 2},
                    timeout=15,
                )
                if r3.status_code != 200:
                    break
                phs, t = self._parse_pharmacy_list(r3.text)
                if page == 0:
                    total = t
                if not phs:
                    break
                all_ph.extend(phs)
                if len(all_ph) >= total:
                    break
                time.sleep(0.3)
            except Exception:
                break

        dist_str = f"{radius_m // 1000}km" if radius_m >= 1000 else f"{radius_m}m"
        return all_ph, f"ナビィ薬局: {dist_str}圏内 全{total}件 / 取得{len(all_ph)}件"

    def _parse_pharmacy_list(self, html: str) -> Tuple[List[PharmacyFacility], int]:
        soup = BeautifulSoup(html, "html.parser")
        results: List[PharmacyFacility] = []
        total = 0
        m = re.search(r"(\d{1,6})\s*件", soup.get_text())
        if m:
            total = int(m.group(1))
        for item in soup.select("div.resultItems div.item"):
            h3 = item.find("h3", class_="name")
            if not h3:
                continue
            link = h3.find("a", href=True)
            if not link:
                continue
            name = link.get_text(strip=True)
            href = link.get("href", "")
            if href.startswith("/"):
                href = MHLW_DOMAIN + href
            pref_cd = re.search(r"prefCd=(\d+)", href)
            kikan_cd = re.search(r"kikanCd=(\w+)", href)
            pref_cd  = pref_cd.group(1)  if pref_cd  else ""
            kikan_cd = kikan_cd.group(1) if kikan_cd else ""
            raw_text = item.get_text(separator=" ", strip=True)
            addr_m = re.search(r"〒\s*[\d-]+\s+(.+?)(?:Googleマップ|$)", raw_text)
            address = addr_m.group(1).strip() if addr_m else ""
            results.append(PharmacyFacility(
                name=name, address=address, href=href,
                pref_cd=pref_cd, kikan_cd=kikan_cd, source="mhlw",
            ))
        return results, total

    def search_medical_by_latlon(
        self,
        lat: float, lon: float,
        radius_m: int,
        center_name: str = "",
        max_pages: int = 6,
    ) -> Tuple[List[MedFacility], str]:
        """ナビィ S2320 → S2400 で医療機関（病院・診療所）を検索する。"""
        if not self._init():
            return [], "MHLW接続エラー"
        dist_code = "00" if radius_m <= 1_000 else ("01" if radius_m <= 5_000 else "")
        try:
            self.session.get(f"{MHLW_BASE}/juminkanja/S2320/initsearch", timeout=12)
            r2 = self.session.get(
                f"{MHLW_BASE}/juminkanja/S2320/search",
                params={
                    "specifyDateAndTime": "01",
                    "centerPointName": urllib.parse.quote(center_name or "検索地点"),
                    "latitude": str(lat), "longitude": str(lon),
                    "selectCenterPoint": "",
                    "distanceFromCenterPoint": dist_code,
                    "medicalCare": ["1", "2"],
                    "searchTypes": "01-2",
                },
                timeout=15,
            )
            j = r2.json()
            if j.get("code") != "0":
                return [], f"MHLW search エラー: {j.get('messages')}"
            redirect_url = j["result"]["redirectUrl"]
        except Exception as e:
            return [], f"MHLW search 例外: {e}"

        all_facs: List[MedFacility] = []
        total = 0
        for page in range(max_pages):
            try:
                sep = "&" if "?" in redirect_url else "?"
                r3 = self.session.get(
                    f"{redirect_url}{sep}page={page}&size=20&sortNo=2", timeout=15
                )
                if r3.status_code != 200:
                    break
                facs, t = self._parse_med_list(r3.text)
                if page == 0:
                    total = t
                if not facs:
                    break
                all_facs.extend(facs)
                if len(all_facs) >= total:
                    break
                time.sleep(0.3)
            except Exception:
                break
        dist_str = f"{radius_m // 1000}km" if radius_m >= 1000 else f"{radius_m}m"
        return all_facs, f"MHLW医療機関: {dist_str}圏内 全{total}件/取得{len(all_facs)}件"

    def _parse_med_list(self, html: str) -> Tuple[List[MedFacility], int]:
        """S2400 医療機関一覧HTMLからMedFacilityリストを生成する（hrefからpref_cd/kikan_cd/kikan_kbn抽出）。"""
        soup = BeautifulSoup(html, "html.parser")
        results: List[MedFacility] = []
        total = 0
        m = re.search(r"(\d{1,6})\s*件", soup.get_text())
        if m:
            total = int(m.group(1))
        for item in soup.find_all("div", class_="item"):
            h3 = item.find("h3", class_="name")
            if not h3:
                continue
            link = h3.find("a", href=True)
            if not link:
                continue
            name = link.get_text(strip=True)
            if not name:
                continue
            href = link.get("href", "")
            if href.startswith("/"):
                href = MHLW_DOMAIN + href
            qp = dict(urllib.parse.parse_qsl(urllib.parse.urlparse(href).query))
            pref_cd   = qp.get("prefCd", "")
            kikan_cd  = qp.get("kikanCd", "")
            kikan_kbn = int(qp.get("kikanKbn", "2"))
            results.append(MedFacility(
                name=name, source="mhlw",
                pref_cd=pref_cd, kikan_cd=kikan_cd, kikan_kbn=kikan_kbn,
            ))
        return results, max(total, len(results))

    def get_facility_detail(self, fac: MedFacility) -> bool:
        """
        MHLW 詳細ページを取得・パース。
        住所取得 + 院内外処方・外来患者数・診療日数を同時取得。
        """
        self._init()
        if not (fac.pref_cd and fac.kikan_cd):
            return False
        known_kbn = fac.kikan_kbn
        kbn_list = [known_kbn] + [k for k in guess_kikan_kbn(fac.kikan_cd) if k != known_kbn]
        soup = None
        used_kbn = None
        for kbn in kbn_list:
            url = (f"{MHLW_BASE}/juminkanja/S2430/initialize"
                   f"?prefCd={fac.pref_cd}&kikanCd={fac.kikan_cd}&kikanKbn={kbn}")
            try:
                r = self.session.get(url, timeout=12)
                if r.status_code != 200:
                    continue
                candidate_soup = BeautifulSoup(r.text, "html.parser")
                text = candidate_soup.get_text()
                if "E-0109" in text or "データは存在しません" in text:
                    continue
                if fac.name[:4] in text or len(text) > 50_000:
                    soup = candidate_soup
                    used_kbn = kbn
                    fac.detail_url = url
                    fac.kikan_kbn = kbn
                    raw_html = r.text
                    break
            except Exception:
                continue
        if soup is None:
            return False

        # ── 座標（ナビィ埋込の正確な緯度経度を最優先） ─────────────────────
        coords = _extract_coords_from_html(raw_html)
        if coords:
            fac.lat, fac.lon = coords
            fac.coord_source = "ナビィ埋込座標"

        # ── 全 tr/dl フィールドを収集 ──────────────────────────────────────
        all_fields: Dict[str, str] = {}
        for row in soup.find_all("tr"):
            cells = row.find_all(["th", "td"])
            if len(cells) >= 2:
                k = cells[0].get_text(strip=True)
                v = " / ".join(c.get_text(strip=True) for c in cells[1:] if c.get_text(strip=True))
                if k and v:
                    all_fields[k] = v
        for dl in soup.find_all("dl"):
            for dt, dd in zip(dl.find_all("dt"), dl.find_all("dd")):
                k = dt.get_text(strip=True)
                v = dd.get_text(strip=True)
                if k and v:
                    all_fields[k] = v
        fac.raw_fields = all_fields
        full_text = soup.get_text(separator="\n", strip=True)

        # ── 住所取得（まだ空の場合のみ） ───────────────────────────────────
        if not fac.address:
            for row in soup.find_all("tr"):
                cells = row.find_all(["th", "td"])
                if len(cells) >= 2:
                    key = cells[0].get_text(strip=True)
                    if re.search(r"所在地|住所", key):
                        val = cells[1].get_text(" ", strip=True)
                        val = re.sub(r"〒\s*\d{3}[-－]\d{4}\s*", "", val).strip()
                        val = re.sub(r"\s+", " ", val).strip()
                        if val:
                            fac.address = val[:120]
                            break
            if not fac.address:
                m = re.search(r"〒\s*[\d-]+\s+(.+?)(?:Tel|TEL|電話|Googleマップ|\n|$)", full_text)
                if m:
                    addr = re.sub(r"\s+", " ", m.group(1)).strip()
                    if addr:
                        fac.address = addr[:120]

        # ── 施設カテゴリ ──────────────────────────────────────────────────
        if used_kbn == 1:
            fac.facility_category = "病院"
        elif used_kbn == 3:
            fac.facility_category = "歯科診療所"
        elif "病院" in fac.name:
            fac.facility_category = "病院"

        # ── 院内処方 / 院外処方 ───────────────────────────────────────────
        inhouse = _get_field(all_fields, [
            "院内処方の有無", "院内処方", "調剤（院内処方）", "院内調剤",
        ])
        outpatient = _get_field(all_fields, [
            "院外処方の有無", "院外処方", "調剤（院外処方）", "院外調剤",
            "処方せんの交付", "処方箋の交付",
        ])
        fac.inhouse_rx    = inhouse    or "—"
        fac.outpatient_rx = outpatient or "—"
        if outpatient and "有" in outpatient:
            fac.rx_summary = "院外処方あり"
        elif inhouse and "有" in inhouse and (not outpatient or "無" in outpatient or "不可" in outpatient):
            fac.rx_summary = "院内処方のみ"
        elif inhouse or outpatient:
            fac.rx_summary = f"院内:{inhouse or '—'} / 院外:{outpatient or '—'}"
        else:
            fac.rx_summary = _infer_rx_type(full_text)

        # ── 1日平均外来患者数 ────────────────────────────────────────────
        fac.daily_outpatients, fac.daily_outpatients_source = \
            _parse_daily_outpatients(all_fields, full_text, soup)

        # 歯科診療所は外来列(index6)が空で歯科列(index7)に患者数が入る
        if fac.daily_outpatients is None and used_kbn == 3:
            raw = all_fields.get("前年度１日平均患者数", "")
            cells = [c.strip() for c in raw.split("/")] if raw else []
            if len(cells) >= 8:
                m = re.search(r"([0-9]+\.?[0-9]*)", cells[7])
                if m:
                    n = float(m.group(1))
                    if 0 < n <= 1_000:
                        fac.daily_outpatients = int(round(n))
                        fac.daily_outpatients_source = "ナビィ（歯科患者列）"

        # ── 週診療日数 ───────────────────────────────────────────────────
        fac.weekly_op_days = _parse_weekly_days(all_fields, full_text, soup)
        # 妥当性クランプ: 週8日等の解析ミス→7日 / 外来数十人規模なのに週1-2日は
        # 診療時間表の解析ミスの可能性が高い→欠測扱い（固定日数フォールバック）
        if fac.weekly_op_days and fac.weekly_op_days > 7:
            fac.weekly_op_days = 7.0
        if (fac.weekly_op_days and fac.weekly_op_days <= 2
                and (fac.daily_outpatients or 0) >= 30):
            fac.weekly_op_days = None

        # ── 診療科目 ─────────────────────────────────────────────────────
        if not fac.specialties:
            fac.specialties = _parse_specialties(all_fields, full_text)

        # ── データ品質検証 ────────────────────────────────────────────────
        fac.beds = _parse_total_beds(all_fields)
        fac.is_cosmetic = _detect_cosmetic(fac.name, fac.specialties)
        fac.op_flag, fac.op_suggested = _validate_outpatients(fac)

        fac.detail_fetched = True
        return True

    def get_pharmacy_detail(self, ph: PharmacyFacility) -> bool:
        """ナビィ薬局詳細ページから総取扱処方箋数を取得する。"""
        self._init()
        if not (ph.pref_cd and ph.kikan_cd):
            return False
        url = (f"{MHLW_BASE}/juminkanja/S2430/initialize"
               f"?prefCd={ph.pref_cd}&kikanCd={ph.kikan_cd}&kikanKbn=5")
        try:
            r = self.session.get(url, timeout=12)
            if r.status_code != 200:
                return False
            soup = BeautifulSoup(r.text, "html.parser")
            text = soup.get_text(separator="\n", strip=True)
            if "E-0109" in text or "データは存在しません" in text:
                return False
        except Exception:
            return False

        ph.detail_url = url

        # 座標（ナビィ埋込の正確な緯度経度があればジオコーディング値より優先）
        coords = _extract_coords_from_html(r.text)
        if coords:
            ph.lat, ph.lon = coords

        all_fields: Dict[str, str] = {}
        for row in soup.find_all("tr"):
            cells = row.find_all(["th", "td"])
            if len(cells) >= 2:
                k = cells[0].get_text(strip=True)
                v = " / ".join(c.get_text(strip=True) for c in cells[1:] if c.get_text(strip=True))
                if k and v:
                    all_fields[k] = v
        for dl in soup.find_all("dl"):
            for dt, dd in zip(dl.find_all("dt"), dl.find_all("dd")):
                k, v = dt.get_text(strip=True), dd.get_text(strip=True)
                if k and v:
                    all_fields[k] = v
        ph.raw_fields = all_fields
        ph.annual_rx_count, ph.annual_rx_source = _parse_annual_rx_count(all_fields, text)
        ph.detail_fetched = True
        return True


# ─── 門前/面 判定 ──────────────────────────────────────────────────────────────
def assign_monzen_to_pharmacies(
    pharmacies: List[PharmacyFacility],
    med_facilities: List[MedFacility],
    threshold_m: float = 50.0,
) -> List[str]:
    """各薬局に最近接の医療機関を割り当て、閾値以内なら門前薬局と判定する。"""
    debug: List[str] = []
    facs_with_coords = [f for f in med_facilities if f.lat is not None and f.lon is not None]
    debug.append(
        f"門前判定: 薬局={len(pharmacies)}件 "
        f"医療機関(座標あり)={len(facs_with_coords)}件 閾値={threshold_m:.0f}m"
    )
    for ph in pharmacies:
        if ph.lat is None or ph.lon is None:
            ph.pharmacy_type = "不明"
            continue
        best_dist = float("inf")
        best_fac: Optional[MedFacility] = None
        for fac in facs_with_coords:
            d = haversine(ph.lat, ph.lon, fac.lat, fac.lon)
            if d < best_dist:
                best_dist = d
                best_fac = fac
        ph.nearest_clinic_dist_m = best_dist if best_fac else None
        ph.nearest_clinic_name   = best_fac.name if best_fac else "—"
        if best_fac and best_dist <= threshold_m:
            ph.pharmacy_type = "門前薬局"
            debug.append(f"  [門前] {ph.name[:20]} → {best_fac.name[:20]} ({best_dist:.0f}m)")
        elif best_fac:
            ph.pharmacy_type = "面薬局"
            debug.append(
                f"  [面] {ph.name[:20]} → 最近接: {best_fac.name[:20]} {best_dist:.0f}m > {threshold_m:.0f}m"
            )
        else:
            ph.pharmacy_type = "不明"
            debug.append(f"  [不明] {ph.name[:20]} → 医療機関データなし")
    return debug


# ─── キャッシュ ────────────────────────────────────────────────────────────────
@st.cache_resource
def get_scraper() -> MHLWScraper:
    return MHLWScraper()


@st.cache_resource
def get_geocoder() -> GeocoderService:
    return GeocoderService()


# ─── メイン分析処理 ────────────────────────────────────────────────────────────
def run_analysis(
    address: str,
    radius_m: int,
    gate_m: int,
    max_detail: int,
    log: List[str],
    prog,
    assumptions: Optional[PredictionAssumptions] = None,
    polygons: Optional[List[List[Tuple[float, float]]]] = None,
    exclude_outside_med: bool = True,
) -> Tuple[List[MedFacility], List[PharmacyFacility], float, float]:
    scraper = get_scraper()
    geocoder = get_geocoder()

    # ─────────────────────────────────────────────────────────────────────
    # Phase 1: 初回データ収集
    # ─────────────────────────────────────────────────────────────────────

    # Step 1: 住所ジオコーディング → 中心座標
    prog.progress(3, text="Step1: 住所をジオコーディング中…")
    t0 = time.time()
    coords = geocoder.geocode(address)
    if not coords:
        st.error(f"住所「{address}」の座標取得に失敗しました。より詳細な住所を入力してください。")
        st.stop()
    center_lat, center_lon = coords
    log.append(
        f"[Step1] ジオコーディング完了: lat={center_lat:.5f}, lon={center_lon:.5f} "
        f"({time.time()-t0:.1f}s)"
    )

    # Step 2: OSM 薬局検索
    prog.progress(8, text="Step2: OSMから薬局を取得中…")
    t0 = time.time()
    time.sleep(2)
    ph_osm = search_osm_pharmacies(center_lat, center_lon, radius_m)
    ph_merged: List[PharmacyFacility] = list(ph_osm)
    log.append(f"[Step2] OSM薬局: {len(ph_osm)}件取得 ({time.time()-t0:.1f}s)")

    # Step 3: OSM 医療機関検索
    prog.progress(14, text="Step3: OSMから医療機関を取得中…")
    t0 = time.time()
    med_radius = radius_m + gate_m
    time.sleep(2)
    med_osm = search_osm_medical(center_lat, center_lon, med_radius)
    log.append(f"[Step3] OSM医療機関: {med_radius}m圏内 {len(med_osm)}件 ({time.time()-t0:.1f}s)")
    if len(med_osm) == 0:
        log.append("[Step3] ⚠️ OSM医療機関0件 → ナビィデータのみで処理します")

    # Step 4: ナビィ薬局リスト取得 → 住所geocoding → OSMとマージ
    prog.progress(20, text="Step4: ナビィから薬局リストを取得中…")
    t0 = time.time()
    navvi_phs, navvi_ph_msg = scraper.search_pharmacies_by_latlon(
        center_lat, center_lon, radius_m=radius_m,
        center_name=address[:20], max_pages=8,
    )
    log.append(f"[Step4] {navvi_ph_msg}")
    existing_names = [p.name for p in ph_merged]
    added_navvi_ph = 0
    for i, nph in enumerate(navvi_phs):
        if i % 5 == 0:
            prog.progress(20, text=f"Step4: ナビィ薬局 座標取得中 {i+1}/{len(navvi_phs)}件…")
        is_dup = any(name_similarity(nph.name, en) >= 0.65 for en in existing_names)
        if is_dup:
            for osm_ph in ph_merged:
                if name_similarity(nph.name, osm_ph.name) >= 0.65 and not osm_ph.pref_cd:
                    osm_ph.pref_cd  = nph.pref_cd
                    osm_ph.kikan_cd = nph.kikan_cd
                    osm_ph.href     = nph.href
            continue
        if nph.address:
            gc = geocoder.geocode(nph.address)
            if gc:
                nph.lat, nph.lon = gc
                nph.distance_m = haversine(center_lat, center_lon, nph.lat, nph.lon)
                if nph.distance_m > radius_m * 1.1:
                    time.sleep(0.15)
                    continue
            time.sleep(0.15)
        ph_merged.append(nph)
        existing_names.append(nph.name)
        added_navvi_ph += 1
    ph_merged.sort(key=lambda x: x.distance_m or 9_999_999)
    no_coord_ph = sum(1 for p in ph_merged if p.lat is None)
    log.append(
        f"[Step4] ナビィ固有追加: {added_navvi_ph}件 合計: {len(ph_merged)}件 "
        f"（座標なし: {no_coord_ph}件） ({time.time()-t0:.1f}s)"
    )

    # Step 5: ナビィ医療機関リスト取得 → get_facility_detail で住所+詳細取得 → geocoding → OSMとマージ
    prog.progress(30, text="Step5: ナビィから医療機関リストを取得中…")
    t0 = time.time()
    navvi_meds, med_msg = scraper.search_medical_by_latlon(
        center_lat, center_lon, radius_m=med_radius,
        center_name=address[:20], max_pages=6,
    )
    log.append(f"[Step5] {med_msg}")

    # 薬局名フィルター（ナビィ医療機関検索に薬局が混入する場合を除外）
    _PHARMA_NAME_RE = re.compile(
        r'薬局|ドラッグ|ファーマ|調剤|くすり|クスリ|drug\s*store|pharmacy', re.IGNORECASE
    )
    med_existing_names = [f.name for f in med_osm]
    med_existing_kikan_cds = {f.kikan_cd for f in med_osm if f.kikan_cd}
    med_targets = [
        f for f in navvi_meds
        if f.pref_cd and f.kikan_cd
        and f.kikan_kbn != 5                          # kikanKbn=5は薬局
        and not _PHARMA_NAME_RE.search(f.name)        # 名前ベースでも除外
        and f.kikan_cd not in med_existing_kikan_cds
        and not any(name_similarity(f.name, en) >= 0.65 for en in med_existing_names)
    ][:50]

    geocode_ok, geocode_fail, detail_fail = 0, 0, 0
    for i, nmf in enumerate(med_targets):
        prog.progress(
            30 + int(20 * i / max(len(med_targets), 1)),
            text=f"Step5: 医療機関 詳細+住所取得中 {i+1}/{len(med_targets)}件: {nmf.name[:15]}…",
        )
        # get_facility_detail で住所取得 + 詳細データを同時取得
        ok = scraper.get_facility_detail(nmf)
        if not ok:
            detail_fail += 1
        if nmf.lat is not None:
            # 詳細ページのナビィ埋込座標（高精度）を取得済 → ジオコーディング不要
            nmf.distance_m = haversine(center_lat, center_lon, nmf.lat, nmf.lon)
            geocode_ok += 1
        elif nmf.address:
            gc = geocoder.geocode(nmf.address)
            if gc:
                nmf.lat, nmf.lon = gc
                nmf.distance_m = haversine(center_lat, center_lon, nmf.lat, nmf.lon)
                geocode_ok += 1
            else:
                geocode_fail += 1
        else:
            geocode_fail += 1
        med_osm.append(nmf)
        med_existing_names.append(nmf.name)
        if nmf.kikan_cd:
            med_existing_kikan_cds.add(nmf.kikan_cd)
        time.sleep(0.2)

    med_osm.sort(key=lambda x: x.distance_m or 9_999_999)
    log.append(
        f"[Step5] 医療機関詳細+住所取得: 成功={geocode_ok}件 "
        f"詳細失敗={detail_fail}件 geocoding失敗={geocode_fail}件 "
        f"合計={len(med_osm)}件 ({time.time()-t0:.1f}s)"
    )

    # ─────────────────────────────────────────────────────────────────────
    # Phase 2: 推考フェーズ（必須・スキップ不可）
    # ─────────────────────────────────────────────────────────────────────

    # Step 6: 【医療機関 漏れ確認】ナビィ再検索
    prog.progress(52, text="Step6（推考①）: 医療機関 漏れ確認中…")
    t0 = time.time()
    existing_med_kikan_cds = {f.kikan_cd for f in med_osm if f.kikan_cd}
    existing_med_names     = [f.name for f in med_osm]
    verify_meds, _ = scraper.search_medical_by_latlon(
        center_lat, center_lon, radius_m=med_radius,
        center_name=address[:20], max_pages=6,
    )
    added_med = 0
    for vf in verify_meds:
        if vf.kikan_cd and vf.kikan_cd in existing_med_kikan_cds:
            continue
        if any(name_similarity(vf.name, en) >= 0.65 for en in existing_med_names):
            continue
        if not (vf.pref_cd and vf.kikan_cd):
            continue
        ok = scraper.get_facility_detail(vf)
        if vf.lat is not None:
            vf.distance_m = haversine(center_lat, center_lon, vf.lat, vf.lon)
        elif vf.address:
            gc = geocoder.geocode(vf.address)
            if gc:
                vf.lat, vf.lon = gc
                vf.distance_m = haversine(center_lat, center_lon, vf.lat, vf.lon)
        vf.source = "mhlw(推考①追加)"
        med_osm.append(vf)
        existing_med_names.append(vf.name)
        if vf.kikan_cd:
            existing_med_kikan_cds.add(vf.kikan_cd)
        added_med += 1
        time.sleep(0.2)
    log.append(
        f"[Step6] 推考①: 医療機関 {added_med}件追加 "
        f"（再検索{len(verify_meds)}件確認） ({time.time()-t0:.1f}s)"
    )

    # Step 7: 【薬局 漏れ確認】ナビィ再検索
    prog.progress(62, text="Step7（推考②）: 薬局 漏れ確認中…")
    t0 = time.time()
    existing_ph_kikan_cds = {p.kikan_cd for p in ph_merged if p.kikan_cd}
    existing_ph_names     = [p.name for p in ph_merged]
    verify_phs, _ = scraper.search_pharmacies_by_latlon(
        center_lat, center_lon, radius_m=radius_m,
        center_name=address[:20], max_pages=8,
    )
    added_ph = 0
    for vph in verify_phs:
        if vph.kikan_cd and vph.kikan_cd in existing_ph_kikan_cds:
            continue
        if any(name_similarity(vph.name, en) >= 0.65 for en in existing_ph_names):
            continue
        if vph.address:
            gc = geocoder.geocode(vph.address)
            if gc:
                vph.lat, vph.lon = gc
                vph.distance_m = haversine(center_lat, center_lon, vph.lat, vph.lon)
                if vph.distance_m > radius_m * 1.1:
                    time.sleep(0.15)
                    continue
            time.sleep(0.15)
        vph.source = "mhlw(推考②追加)"
        ph_merged.append(vph)
        existing_ph_names.append(vph.name)
        if vph.kikan_cd:
            existing_ph_kikan_cds.add(vph.kikan_cd)
        added_ph += 1
    ph_merged.sort(key=lambda x: x.distance_m or 9_999_999)
    log.append(
        f"[Step7] 推考②: 薬局 {added_ph}件追加 "
        f"（再検索{len(verify_phs)}件確認） ({time.time()-t0:.1f}s)"
    )

    # Step 8: 【距離整合性チェック】
    prog.progress(70, text="Step8（推考③）: 距離整合性チェック中…")
    t0 = time.time()
    warnings_med = 0
    for fac in med_osm:
        if fac.lat is not None and fac.lon is not None:
            actual_dist = haversine(center_lat, center_lon, fac.lat, fac.lon)
            fac.distance_m = actual_dist
            if actual_dist > med_radius + 500:
                warnings_med += 1
                log.append(
                    f"[Step8] ⚠️ 圏外疑い(医療): {fac.name[:20]} "
                    f"距離={actual_dist:.0f}m > {med_radius + 500}m"
                )

    removed_ph = 0
    ph_filtered: List[PharmacyFacility] = []
    for ph in ph_merged:
        if ph.distance_m is not None and ph.distance_m > radius_m * 1.1:
            log.append(f"[Step8] 除外(薬局距離超過): {ph.name[:20]} {ph.distance_m:.0f}m")
            removed_ph += 1
        else:
            ph_filtered.append(ph)
    ph_merged = ph_filtered
    log.append(
        f"[Step8] 推考③: 距離整合性チェック完了 "
        f"医療機関警告={warnings_med}件 薬局除外={removed_ph}件 ({time.time()-t0:.1f}s)"
    )

    # ─────────────────────────────────────────────────────────────────────
    # Phase 3: 詳細取得 & 判定
    # ─────────────────────────────────────────────────────────────────────

    # Step 9: 薬局詳細ページから年間処方箋数取得
    prog.progress(74, text="Step9: 薬局詳細（年間処方箋数）を取得中…")
    t0 = time.time()
    ph_targets = [p for p in ph_merged if p.pref_cd and p.kikan_cd and not p.detail_fetched][:max_detail]
    log.append(f"[Step9] 薬局詳細取得対象: {len(ph_targets)}件")
    for i, ph in enumerate(ph_targets):
        prog.progress(
            74 + int(18 * i / max(len(ph_targets), 1)),
            text=f"Step9: 薬局詳細取得中 ({i+1}/{len(ph_targets)}): {ph.name[:20]}…",
        )
        scraper.get_pharmacy_detail(ph)
        time.sleep(0.5)
    fetched_ph = sum(1 for p in ph_merged if p.detail_fetched)
    # 詳細ページで座標が更新された薬局の距離を再計算（門前判定の精度向上）
    for ph in ph_merged:
        if ph.lat is not None and ph.lon is not None:
            ph.distance_m = haversine(center_lat, center_lon, ph.lat, ph.lon)
    ph_merged.sort(key=lambda x: x.distance_m or 9_999_999)
    log.append(f"[Step9] 薬局詳細取得完了: {fetched_ph}件 ({time.time()-t0:.1f}s)")

    # Step 10: 門前/面 判定
    prog.progress(93, text="Step10: 門前/面 判定中…")
    t0 = time.time()
    debug_lines = assign_monzen_to_pharmacies(ph_merged, med_osm, threshold_m=float(gate_m))
    log.extend(debug_lines)
    n_monzen = sum(1 for p in ph_merged if p.pharmacy_type == "門前薬局")
    n_men    = sum(1 for p in ph_merged if p.pharmacy_type == "面薬局")
    log.append(
        f"[Step10] 門前/面判定完了: 門前={n_monzen}件 面={n_men}件 "
        f"不明={len(ph_merged)-n_monzen-n_men}件 ({time.time()-t0:.1f}s)"
    )

    # Step 11: 門前占有チェック（各クリニックの最近接 既存薬局距離を計算）
    prog.progress(95, text="Step11: 門前占有チェック中…")
    compute_pharmacy_proximity(med_osm, ph_merged)

    # Step 11.5: 商圏ポリゴンの内外判定（ポリゴンモードのみ）
    if polygons:
        n_med_out, n_ph_out = apply_area_flags(
            med_osm, ph_merged, polygons, exclude_outside_med
        )
        log.append(
            f"[Step11.5] 商圏ポリゴン判定: ポリゴン{len(polygons)}個 "
            f"圏外医療機関={n_med_out}件 圏外薬局={n_ph_out}件 "
            f"（医療機関の寄与除外: {'ON' if exclude_outside_med else 'OFF'}）"
        )

    # Step 12: 処方箋獲得予測（候補地点＝商圏中心 が獲得する年間処方箋枚数）
    prog.progress(96, text="Step12: 処方箋獲得予測を計算中…")
    a = assumptions or PredictionAssumptions()
    summary = compute_capture_prediction(med_osm, a)
    log.append(
        f"[Step12] 処方箋獲得予測: 年間 {summary['total_annual_rx']:,.0f} 枚 "
        f"（寄与医療機関={summary['n_contributing']}件 / "
        f"外来数なし={summary['n_no_outpatient']}件 / "
        f"門前競合={summary.get('n_contested_monzen', 0)}件）"
    )

    # Step 12: 結果をsession_stateに保存（呼び出し元で保存）
    med_osm.sort(key=lambda x: x.distance_m or 9_999_999)
    ph_merged.sort(key=lambda x: x.distance_m or 9_999_999)
    log.append(
        f"[完了] 医療機関={len(med_osm)}件 薬局={len(ph_merged)}件 "
        f"（座標あり医療機関: {sum(1 for f in med_osm if f.lat is not None)}件）"
    )
    return med_osm, ph_merged, center_lat, center_lon


# ─── セッション初期化 ──────────────────────────────────────────────────────────
_defaults = {
    "med_results":    [],
    "ph_results":     [],
    "center_lat":     None,
    "center_lon":     None,
    "search_log":     [],
    "last_address":   "",
    "area_polygons":  [],
    "collected_radius": None,
    "searched_polygon_mode": False,
}
for k, v in _defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ─── UI ───────────────────────────────────────────────────────────────────────

# ════════════════════════════════ アプリ本体 ════════════════════════════════
import io
import math
import os
import re
from dataclasses import replace

import folium
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from streamlit_folium import st_folium


def _num(v):
    """pandas由来のNaN/None/空 を None に、数値は float にする。"""
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return None if f != f else f  # NaN 除外


def _reposition(clat, clon, flat, flon, target_dist):
    """候補地→現在地の向きを保ったまま、距離だけ target_dist(m) に補正した座標を返す。"""
    cur = haversine(clat, clon, flat, flon)
    if cur <= 0 or target_dist <= 0:
        return flat, flon
    ratio = target_dist / cur
    return clat + (flat - clat) * ratio, clon + (flon - clon) * ratio


def _recs_sig(recs, numfield):
    """編集レコードの正規化シグネチャ（float揺れによる無限再実行を防ぐ比較用）。"""
    out = []
    for r in recs:
        la, lo, nv = _num(r.get("lat")), _num(r.get("lon")), _num(r.get(numfield))
        out.append(((r.get("name") or "").strip(),
                    round(la, 6) if la is not None else None,
                    round(lo, 6) if lo is not None else None,
                    round(nv, 1) if nv is not None else None))
    return out


_EXT_CATS = ["院外のみ", "院内外どちらも", "院内のみ", "不明"]

# ── 診療科別の処方箋発行率（＝そもそも処方箋を出す割合。院外率とは別物） ──────────────
# 【根拠】
#  ・院外“処方率”は厚労省・社会医療診療行為別統計 令和6年で全体81.4%（＝院外率0.8313に反映済）。
#  ・受診のうち投薬に至る割合（発行率）の診療科差は、日医総研RR113「診療所の診療科特性」表5.3.1
#    （厚労省 社会医療診療行為別統計より作成）の“診療科別・入院外・投薬の点数構成比(2020)”で裏づけ：
#      内科20.4% 皮膚22.0% 産婦21.1% 外科16.3% 精神15.1% 耳鼻14.9% 泌尿14.8% 整形12.8% 眼科11.8% 小児11.1%
#    整形外科は投薬が低く代わりに「その他(リハビリ等)」28.8%・画像13.4%、眼科は検査43.6%＋手術20.0%が主体。
#  ・ただし点数構成比は“収益の内訳”で発行「頻度」とは別物（小児=医学管理料主体/精神=精神療法主体で
#    点数は低いが投薬頻度は高い）。そこで点数構成比を方向性の根拠とし、臨床実態で補正した初期値とした。
#  ・下表の件数加重平均は約0.803で、再較正時のフラット値0.8054とほぼ一致（＝全体水準は保ちつつ科別に再配分）。
#  ・値はサイドバー（科ごと）・医療機関表（施設ごと）で編集可能。
DEFAULT_ISSUE = 0.8054  # その他・不明・大病院はこの一律値（従来値）
DEPT_DEFAULTS = [
    ("内科系", 0.90), ("精神科", 0.92), ("小児科", 0.88), ("耳鼻咽喉科", 0.85),
    ("皮膚科", 0.85), ("泌尿器科", 0.75), ("産婦人科", 0.60), ("眼科", 0.65),
    ("整形外科", 0.55), ("外科", 0.60), ("リハビリ科", 0.20), ("美容", 0.00),
    ("病院", DEFAULT_ISSUE), ("その他", DEFAULT_ISSUE),
]
DEPT_OPTIONS = [k for k, _ in DEPT_DEFAULTS]
# 判定キーワード（上から順に最初に一致したバケットを採用。整形は外科より前・内科系は外科より前）
DEPT_KWS = [
    ("美容", ["美容"]),
    ("整形外科", ["整形"]),
    ("リハビリ科", ["リハビリ", "リハビリテーション"]),
    ("眼科", ["眼科"]),
    ("耳鼻咽喉科", ["耳鼻", "咽喉"]),
    ("皮膚科", ["皮膚", "スキン"]),
    ("泌尿器科", ["泌尿"]),
    ("産婦人科", ["産婦", "婦人科", "産科"]),
    ("精神科", ["精神", "心療", "メンタル"]),
    ("小児科", ["小児"]),
    ("内科系", ["内科", "糖尿", "代謝", "循環器", "呼吸器", "消化器", "胃腸", "腎臓",
              "内分泌", "血液", "神経内科", "アレルギー", "リウマチ", "感染症", "ペイン", "在宅"]),
    ("外科", ["外科"]),
]


def get_dept_rates():
    """診療科→発行率の現在値（サイドバーで編集可能。session_stateに保持）。"""
    dr = st.session_state.get("dept_rates")
    if not dr:
        dr = {k: v for k, v in DEPT_DEFAULTS}
        st.session_state["dept_rates"] = dr
    return dr


def bucket_of_med(fac):
    """医療機関の診療科バケットを判定。大病院(20床以上/病院)は『病院』、美容は『美容』。"""
    if getattr(fac, "is_cosmetic", False):
        return "美容"
    cat = getattr(fac, "facility_category", "") or ""
    beds = getattr(fac, "beds", 0) or 0
    if ("病院" in cat and "診療所" not in cat) or (beds and beds >= 20):
        return "病院"  # 大病院は外来が全科合算のため一律（=DEFAULT_ISSUE）
    specs = getattr(fac, "specialties", "") or ""
    if not isinstance(specs, str):     # 念のためリスト等でも文字列化
        specs = " ".join(str(x) for x in specs)
    name = getattr(fac, "name", "") or ""
    for probe in (specs, name):  # まず標榜診療科、無ければ名称から推定
        if not probe:
            continue
        for bname, kws in DEPT_KWS:
            if any(kw in probe for kw in kws):
                return bname
    return "その他"


def eff_issue_rate(dept, override, dept_rates):
    """施設の実効発行率＝手入力override（あれば）／無ければ診療科バケットの現在値。"""
    if override is not None:
        return override
    return dept_rates.get(dept, DEFAULT_ISSUE)


# ── 診療（開局）曜日・時間の抽出（ナビィ詳細ページの raw_fields から） ─────────────────
# ナビィは月〜日＋祝の8列を "/" 連結した時間割を持つ。医療機関＝「診療時間（診療科目別の）…」、
# 薬局＝「時間帯１」。曜日・時間は必ずこの8列時間割から index で導出する
# （"営業日""診療日" 等の直接キーは、他項目やツールチップ文言への部分一致で誤取得するため使わない）。
_WEEK_LABELS = ["月", "火", "水", "木", "金", "土", "日", "祝"]
_SCHEDULE_KEYS = [
    "時間帯１", "時間帯1",                 # 薬局（開局時間）
    "診療時間（診療科目別の）",             # 医療機関（診療科目別の診療時間）
    "外来受付時間（診療科目別の）",         # 医療機関（外来受付時間）
]


def parse_open_schedule(fields):
    """raw_fields から (開いている曜日, 代表的な時間帯) を作る。取得できなければ ("", "")。
    例: 内山皮膚科→("月火水金土","09:00-12:00") / マルヤマ薬局→("月火水木金土","09:00-19:00")。"""
    if not fields:
        return "", ""
    sv = None
    for key in _SCHEDULE_KEYS:                 # 8列の時間割フィールドを探す
        for fk, fv in fields.items():
            if key in fk and "/" in fv and re.search(r"\d{1,2}:\d{2}", fv):
                sv = fv
                break
        if sv:
            break
    if not sv:
        return "", ""
    parts = [p.strip() for p in sv.split("/")]
    open_days, times = [], []
    for i, p in enumerate(parts[:8]):          # 0..7 = 月火水木金土日祝
        if p and re.search(r"\d{1,2}:\d{2}", p):
            if i < len(_WEEK_LABELS):
                open_days.append(_WEEK_LABELS[i])
            times.append(re.sub(r"\s+", "", p))
    days_str = "".join(open_days)
    hours_str = max(set(times), key=times.count) if times else ""   # 代表＝最頻の時間帯
    return days_str, hours_str


def rx_category(fac):
    """ナビィの院内/院外処方フィールドから、院外のみ/院内外どちらも/院内のみ/不明を判定。"""
    ih = (getattr(fac, "inhouse_rx", "") or "")
    op = (getattr(fac, "outpatient_rx", "") or "")
    has_in, has_out = ("有" in ih), ("有" in op)
    if has_out and has_in:
        return "院内外どちらも"
    if has_out and not has_in:
        return "院外のみ"
    if has_in and not has_out:
        return "院内のみ"
    s = getattr(fac, "rx_summary", "") or ""
    if s.startswith("院外処方あり"):
        return "院内外どちらも"
    if s == "院内処方のみ":
        return "院内のみ"
    return "不明"


def ext_coef(cat, ext_rate):
    """院外区分→院外係数。院外のみ=1.0／院内外どちらも・不明=院外率／院内のみ=0。"""
    if cat == "院外のみ":
        return 1.0
    if cat == "院内のみ":
        return 0.0
    return ext_rate


def clinic_pool(fac, cat, a, issue_rate, ext_rate, unknown_op):
    """クリニックの年間院外処方（原資）＝外来×診療日数×発行率×院外係数（美容/歯科は減係数）。"""
    op = fac.daily_outpatients
    if not op:
        op = unknown_op  # 外来不明→既定値（下目）
    days = (fac.weekly_op_days * 52.0
            if (a.annual_days_mode == "weekly" and fac.weekly_op_days) else float(a.fixed_annual_days))
    coef = ext_coef(cat, ext_rate)
    if getattr(fac, "is_cosmetic", False):
        coef *= a.cosmetic_factor
    elif getattr(fac, "facility_category", "") == "歯科診療所":
        coef *= a.dental_factor
    ir = getattr(fac, "issue_eff", None)  # 診療科別の実効発行率（未設定なら一律値）
    if ir is None:
        ir = issue_rate
    return op * days * ir * coef


def clinic_flag(fac, high_thr):
    """外れ値・不明のアラート文字列（空=正常）。"""
    flags = []
    op = fac.daily_outpatients
    if not op:
        flags.append("外来不明→既定値使用")
    elif getattr(fac, "facility_category", "") != "病院" and op >= high_thr:
        flags.append(f"要確認：{int(op)}人/日は多め")
    cf = getattr(fac, "op_flag", "") or ""
    if cf:
        flags.append(cf)
    return " / ".join(flags)


def resolve_edit(ed, name_col, num_disp, num_store, stored, clat, clon, cat_col=None,
                 dept_col=None, issue_col=None):
    """
    編集後の data_editor 内容を、保存用レコードに解決する。行ごとに：
      ・緯度/経度を編集した → その座標を採用
      ・距離(m)だけ編集した → 元座標の向きを保って距離を補正
      ・新規行 → 緯度経度があれば採用、無く距離だけなら候補地の真北に仮置き
    """
    smap = {s.get("_key"): s for s in stored if isinstance(s.get("_key"), str) and s.get("_key")}
    out = []
    for _, row in ed.iterrows():
        name = (str(row.get(name_col)) if row.get(name_col) is not None else "").strip()
        lat, lon = _num(row.get("緯度")), _num(row.get("経度"))
        dist, nv = _num(row.get("距離(m)")), _num(row.get(num_disp))
        key = row.get("_key")
        key = key if (isinstance(key, str) and key) else None
        s = smap.get(key)
        if s is not None:
            slat, slon = _num(s.get("lat")), _num(s.get("lon"))
            sdist = (haversine(clat, clon, slat, slon)
                     if (slat is not None and slon is not None) else None)
            coords_edited = (slat is not None and lat is not None and lon is not None
                             and (abs(lat - slat) > 1e-9 or abs(lon - slon) > 1e-9))
            if coords_edited:
                pass  # 編集した緯度経度を採用
            elif dist is not None and sdist is not None and abs(dist - sdist) > 1:
                lat, lon = _reposition(clat, clon, slat, slon, dist)  # 距離だけ補正
            else:
                lat, lon = slat, slon  # 変更なし
        else:
            if (lat is None or lon is None) and dist is not None and dist > 0:
                lat, lon = clat + dist / 111000.0, clon  # 新規・距離のみ→真北に仮置き
        if name or (lat is not None and lon is not None):
            rec = {"name": name, "lat": lat, "lon": lon, num_store: nv, "_key": key}
            if cat_col:
                cv = row.get(cat_col)
                rec["cat"] = cv if cv in _EXT_CATS else "不明"
            if dept_col:
                dv = row.get(dept_col)
                rec["dept"] = dv if (isinstance(dv, str) and dv in DEPT_OPTIONS) else None
            if issue_col:
                rec["issue_disp"] = _num(row.get(issue_col))
            out.append(rec)
    return out


def effective_facilities(raw, clat, clon, label):
    """
    生データ＋候補地ごとの編集（座標補正・削除・手動追加）を反映した実効の医療機関/薬局リストを返す。
    編集内容は session_state['med_edit'][label] / ['ph_edit'][label] に行レコードで保持。
    """
    med_edit = st.session_state.setdefault("med_edit", {})
    ph_edit = st.session_state.setdefault("ph_edit", {})
    if label not in med_edit:
        med_edit[label] = [{"name": f.name, "lat": f.lat, "lon": f.lon,
                            "op": f.daily_outpatients, "cat": rx_category(f),
                            "dept": bucket_of_med(f), "issue": None,
                            "_key": facility_key(f)} for f in raw["med"]]
    if label not in ph_edit:
        ph_edit[label] = [{"name": p.name, "lat": p.lat, "lon": p.lon,
                           "rx": p.annual_rx_count, "_key": pharmacy_key(p)} for p in raw["ph"]]
    raw_med_map = {facility_key(f): f for f in raw["med"]}
    raw_ph_map = {pharmacy_key(p): p for p in raw["ph"]}

    dept_rates = get_dept_rates()
    med_eff = []
    for r in med_edit[label]:
        name = (r.get("name") or "").strip()
        lat, lon = _num(r.get("lat")), _num(r.get("lon"))
        if not name or lat is None or lon is None:
            continue
        op = _num(r.get("op"))
        base = raw_med_map.get(r.get("_key"))
        if base is not None:
            f = replace(base, lat=lat, lon=lon,
                        daily_outpatients=(int(op) if op else base.daily_outpatients))
        else:
            f = MedFacility(name=name, lat=lat, lon=lon,
                            daily_outpatients=(int(op) if op else None),
                            rx_summary="院外処方あり", facility_category="診療所",
                            source="手動追加")
        cat = r.get("cat")
        f.rx_cat = cat if cat in _EXT_CATS else rx_category(f)
        # 診療科バケット（手修正があれば優先）と実効発行率
        dept = r.get("dept")
        if dept not in dept_rates:
            dept = bucket_of_med(f)
        f.dept_name = dept
        f.issue_eff = eff_issue_rate(dept, _num(r.get("issue")), dept_rates)
        f.distance_m = haversine(clat, clon, lat, lon)
        med_eff.append(f)

    ph_eff = []
    for r in ph_edit[label]:
        name = (r.get("name") or "").strip()
        lat, lon = _num(r.get("lat")), _num(r.get("lon"))
        if not name or lat is None or lon is None:
            continue
        rx = _num(r.get("rx"))
        base = raw_ph_map.get(r.get("_key"))
        if base is not None:
            p = replace(base, lat=lat, lon=lon,
                        annual_rx_count=(int(rx) if rx else base.annual_rx_count))
        else:
            p = PharmacyFacility(name=name, address="", lat=lat, lon=lon,
                                 annual_rx_count=(int(rx) if rx else None), source="手動追加")
        p.distance_m = haversine(clat, clon, lat, lon)
        ph_eff.append(p)
    return med_eff, ph_eff


@st.cache_resource
def get_scraper():
    return MHLWScraper()


# ════════════════════════════════ サイドバー ════════════════════════════════
with st.sidebar:
    st.header("共通設定")
    radius_m = st.slider("商圏半径 (m)", 500, 5000, 3000, 100,
                         help="スーパー商圏に準拠。全候補地に共通で使います。")
    fetch_all_ph = st.checkbox(
        "商圏内の薬局を全件取得（重み付けを最も正確に・ただし遅い）", value=True,
        help="ON：商圏内の全薬局の詳細（正確な座標・実績）を取得し、門前判定・実績を最も正確にします。"
             "OFF：近い順に指定件数だけ詳細取得（遠い競合も座標は取得済みで按分には入ります）。",
    )
    if fetch_all_ph:
        max_detail = 9999
        st.caption("→ 商圏内の薬局を全件取得します（速度優先にしたい場合はチェックを外すと件数指定が出ます）。")
    else:
        max_detail = int(st.slider("詳細取得件数（薬局）", 5, 120, 30, 5,
                                   help="ナビィから実績・座標を取る薬局の上限（多いほど正確・遅い）。"))
    gate_m = 50

    st.divider()
    st.subheader("🛒 集客ベースの前提（全候補地に共通）")
    ff_r65 = st.number_input("65歳以上の比率（商圏の高齢化率）", 0.0, 1.0, 0.30, 0.01,
                             help="来店客の年齢構成。会員の年齢データがあればその比率、なければ商圏の高齢化率。")
    c1, c2 = st.columns(2)
    ff_v65 = c1.number_input("65+ の月受診回数", 0.0, 6.0, 3.0, 0.1)
    ff_vu65 = c2.number_input("65- の月受診回数", 0.0, 6.0, 1.3, 0.1)
    c3, c4, c5 = st.columns(3)
    ff_issue = c3.number_input("発行率", 0.0, 1.0, 0.8054, 0.0001, format="%.4f")
    ff_ext = c4.number_input("院外率", 0.0, 1.0, 0.8313, 0.0001, format="%.4f")
    ff_use = c5.number_input("利用率", 0.0, 1.0, 0.137, 0.001, format="%.3f")
    c6, c7 = st.columns(2)
    ff_monzen = c6.number_input("門前しきい値(m)", 0, 300, 50, 10,
                                help="最寄りクリニックがこの距離以内の薬局は門前として自動判定→面競合から除外。")
    ff_decay = c7.number_input("面競合の距離減衰λ(m)", 0, 3000, 1000, 100,
                               help="遠い面競合を弱く数える。小さいほど自店シェア↑。")
    ff_main = st.number_input(
        "メイン薬局しきい値(枚/年・0=無効/既定)", 0, 100000, 0, 1000,
        help="既定0＝無効。実績が大きい面薬局は“面の強豪”なので、除外せずパワー加重で強い競合として"
             "数えます（門前かどうかは距離＝門前しきい値で判定）。門前で大量の店だけ外したい場合のみ値を入れます。",
    )

    with st.expander("⚙️ 詳細設定（ハフ按分・通常は変更不要）", expanded=False):
        huff_lambda = st.slider(
            "距離減衰 λ (m)", 150, 1200, 300, 50,
            help="発行率×院外係数を入れた新原資に合わせ、地方69シード・面型106店で再較正した値"
                 "（予測/実績 中央値0.99）。",
        )
        huff_boost = st.slider("門前ブースト", 1.0, 15.0, 6.0, 0.5)
        huff_monzen_r = st.slider(
            "門前ブースト半径 (m)", 30, 150, 50, 10,
            help="①医療機関ベースで、クリニックからこの距離以内の薬局に門前ブーストを掛けます。"
                 "実質“門前”が80〜100mにある場合はここを広げてください（既定50mは275店検証時の値のため、"
                 "広げた場合は再較正が望ましい）。",
        )
        huff_candA = st.number_input("候補店の引力（大型店は上げる）", 0.2, 10.0, 1.0, 0.1)
        st.caption("── 医療機関ベースの原資（外来→処方箋の換算）──")
        med_unknown_op = st.number_input(
            "外来不明クリニックの既定外来数(人/日)", 0, 500, 30, 5,
            help="ナビィで外来患者数が取得できないクリニックに入れる値（未入力なので下目に設定）。",
        )
        med_high_thr = st.number_input(
            "外れ値アラートしきい値（診療所・人/日）", 50, 2000, 200, 10,
            help="診療所でこの人数以上なら『要確認』を表示（月間値・誤登録の疑い）。",
        )
        st.caption("── 診療科別の処方箋発行率（①医療機関ベースの原資に使用・編集可）──")
        _dept_rates = get_dept_rates()
        _dr_df = pd.DataFrame([{"診療科": k, "発行率": _dept_rates.get(k, v)}
                               for k, v in DEPT_DEFAULTS])
        _dr_ed = st.data_editor(
            _dr_df, hide_index=True, use_container_width=True, key="dept_rates_editor",
            disabled=["診療科"],
            column_config={"発行率": st.column_config.NumberColumn(
                "発行率", min_value=0.0, max_value=1.0, step=0.01, format="%.2f")},
        )
        _new_dr = {}
        for _, _rr in _dr_ed.iterrows():
            _k = _rr["診療科"]
            _v = _num(_rr["発行率"])
            _new_dr[_k] = _v if _v is not None else _dept_rates.get(_k, DEFAULT_ISSUE)
        if _new_dr != _dept_rates:
            st.session_state["dept_rates"] = _new_dr   # 保存のみ（この後の再計算で即反映・st.rerun不要）
        st.caption("整形外科0.55/リハビリ0.20/内科系0.90 等（受診のうち投薬に至る割合）。"
                   "診療科は施設ごとに医療機関表でも変更できます。大病院は一律で『病院』値。")
        st.caption("根拠：日医総研 診療所の診療科特性（厚労省 社会医療診療行為別統計より作成）の"
                   "診療科別・入院外・投薬の点数構成比（2020）＝内科20.4%/皮膚22.0%/整形12.8%/眼科11.8% 等。"
                   "件数加重平均≒0.80で全体水準は較正値と整合。")
        st.caption("※ ①の原資 = 外来×診療日数×**診療科別発行率**×院外係数"
                   "（院外のみ1.0／院内外どちらも=院外率/院内のみ0）。"
                   f"②集客ベースの発行率は上の {float(ff_issue):.4f}（館全体の平均）を使用。")

    st.caption("※ サイドバーや面/門前を変えると、再検索なしで比較表・Excelが即更新されます。")


def make_fp(uni):
    return FootfallParams(
        enabled=(uni > 0),
        unique_customers_monthly=float(uni), ratio_65plus=float(ff_r65),
        visits_month_65plus=float(ff_v65), visits_month_under65=float(ff_vu65),
        issue_rate=float(ff_issue), external_rate=float(ff_ext), use_rate=float(ff_use),
        menkata_monzen_dist=float(ff_monzen), menkata_main_rx=float(ff_main),
        competitor_decay_m=float(ff_decay),
    )


def make_hp():
    return HuffParams(lambda_m=float(huff_lambda), monzen_boost=float(huff_boost),
                      candidate_attractiveness=float(huff_candA), monzen_radius=float(huff_monzen_r))


# ── ハフの取り分内訳（クリニック1行ずつ・自店の重み/競合の重み合計を明示） ─────────
def huff_breakdown(med, ph, clat, clon, hp, a, issue_rate, ext_rate, unknown_op):
    comps = []
    for p in ph:
        if p.lat is None or p.lon is None:
            continue
        ak = _pharmacy_attractiveness(p, hp.national_avg_rx) if hp.weight_by_power else 1.0
        comps.append((p.lat, p.lon, ak))

    def bw(d, aa):
        v = math.exp(-d / hp.lambda_m)
        if d <= hp.monzen_radius:
            v *= hp.monzen_boost
        return aa * v

    rows = []
    for f in med:
        if f.lat is None or f.lon is None or not getattr(f, "in_area", True):
            continue
        d_self = haversine(clat, clon, f.lat, f.lon)
        if d_self > hp.reach_m:
            continue
        cat = getattr(f, "rx_cat", None) or rx_category(f)
        pool = clinic_pool(f, cat, a, issue_rate, ext_rate, unknown_op)
        if pool <= 0:
            continue
        self_w = bw(d_self, hp.candidate_attractiveness)
        den = self_w
        for (plat, plon, ak) in comps:
            dk = haversine(plat, plon, f.lat, f.lon)
            if dk <= hp.reach_m:
                den += bw(dk, ak)
        share = self_w / den if den > 0 else 0.0
        rows.append({"clinic": f.name, "dist": d_self, "pool": pool,
                     "dept": getattr(f, "dept_name", "その他"),
                     "issue": getattr(f, "issue_eff", issue_rate),
                     "self_w": self_w, "comp_w": den - self_w,
                     "share": share, "captured": pool * share})
    rows.sort(key=lambda r: r["captured"], reverse=True)
    return rows


# ── 生データ＋現在の設定/手修正から、両トラックを算出（再検索なしで即再計算） ──────
def compute_candidate(raw):
    clat, clon, uni = raw["clat"], raw["clon"], raw["uni"]
    label = raw["label"]
    a = PredictionAssumptions()
    hp = make_hp()
    # 座標補正・削除・手動追加を反映した実効の医療機関/薬局リスト（①②の両方に効く）
    med, ph = effective_facilities(raw, clat, clon, label)
    # 周知率（接触率）：館の来店客数のうち、その薬局に接触・到達する割合。手修正が優先。
    exposure = float(st.session_state.get("exp_multi", {}).get(label, raw.get("exposure", 1.0)))
    eff_uni = uni * exposure
    fp = make_fp(eff_uni)
    hb = huff_breakdown(med, ph, clat, clon, hp, a,
                        float(ff_issue), float(ff_ext), float(med_unknown_op))
    med_total = sum(r["captured"] for r in hb)
    classified = classify_menkata(ph, med, clat, clon,
                                  monzen_dist=fp.menkata_monzen_dist,
                                  main_rx_threshold=fp.menkata_main_rx, reach_m=hp.reach_m)
    override = st.session_state.get("mk_multi", {}).get(label, {})
    cpow, cn, cexc = footfall_competitor_power(classified, override, fp.competitor_decay_m,
                                               hp.national_avg_rx)
    foot = compute_footfall_prediction(fp, cpow)
    return {
        "label": label, "name": raw["name"], "addr": raw["addr"],
        "uni": uni, "exposure": exposure, "eff_uni": eff_uni,
        "med": med, "ph": ph, "clat": clat, "clon": clon, "hp": hp, "fp": fp,
        "huff_rows": hb, "med_total": med_total, "classified": classified, "override": override,
        "comp_power": cpow, "comp_n": cn, "comp_excluded": cexc,
        "foot_total": (foot["total"] if foot else None), "foot": foot,
    }


# ════════════════════════════════ 商圏マップ ════════════════════════════════
def _med_color(f):
    cat = getattr(f, "facility_category", "") or ""
    if getattr(f, "is_cosmetic", False):
        return "purple"
    if "病院" in cat and "診療所" not in cat:
        return "darkblue"
    return "blue"


def build_map(c, radius_m):
    """候補地＋商圏円＋周辺の医療機関/薬局を1枚に描く（スクショ用）。
    描画は実効リスト c['med']/c['ph'] を使うので、手動追加・削除・座標修正がそのまま反映される。"""
    clat, clon = c["clat"], c["clon"]
    m = folium.Map(location=[clat, clon], zoom_start=15, control_scale=True, tiles="OpenStreetMap")
    folium.Circle([clat, clon], radius=float(radius_m), color="#0F766E", weight=2,
                  fill=True, fill_color="#0F766E", fill_opacity=0.05,
                  tooltip=f"商圏 {int(radius_m)}m").add_to(m)
    folium.Marker(
        [clat, clon], tooltip=f"候補地 {c['label']}",
        popup=folium.Popup(
            f"<b>候補地 {c['label']}</b><br>{c['name'] or ''}<br>{c['addr']}", max_width=260),
        icon=folium.Icon(color="red", icon="star", prefix="fa")).add_to(m)

    for f in c["med"]:
        if f.lat is None or f.lon is None:
            continue
        d = round(haversine(clat, clon, f.lat, f.lon))
        op = f.daily_outpatients
        days, hours = parse_open_schedule(getattr(f, "raw_fields", None))
        sched = (f"<br>診療日: {days}" if days else "") + (f"<br>診療時間: {hours}" if hours else "")
        html = (f"<b>{f.name}</b><br>距離 {d}m<br>"
                f"診療科: {getattr(f, 'dept_name', '—')}"
                f"（発行率 {getattr(f, 'issue_eff', 0.0):.2f}）<br>"
                f"外来: {int(op) if op else '不明'} 人/日<br>"
                f"院外区分: {getattr(f, 'rx_cat', '—')}{sched}")
        folium.Marker([f.lat, f.lon], tooltip=f.name,
                      popup=folium.Popup(html, max_width=280),
                      icon=folium.Icon(color=_med_color(f), icon="plus", prefix="fa")).add_to(m)

    menkata = {cl["key"]: c["override"].get(cl["key"], cl["auto_menkata"])
               for cl in c["classified"]}
    for p in c["ph"]:
        if p.lat is None or p.lon is None:
            continue
        d = round(haversine(clat, clon, p.lat, p.lon))
        is_men = menkata.get(pharmacy_key(p), True)
        rx = p.annual_rx_count
        days, hours = parse_open_schedule(getattr(p, "raw_fields", None))
        sched = (f"<br>開局日: {days}" if days else "") + (f"<br>開局時間: {hours}" if hours else "")
        html = (f"<b>{p.name}</b><br>距離 {d}m<br>"
                f"区分: {'面' if is_men else '門前'}<br>"
                f"実績: {int(rx) if rx else '—'} 枚/年{sched}")
        folium.Marker([p.lat, p.lon], tooltip=p.name,
                      popup=folium.Popup(html, max_width=250),
                      icon=folium.Icon(color=("green" if is_men else "orange"),
                                       icon="medkit", prefix="fa")).add_to(m)

    legend = (
        '<div style="position:fixed;bottom:22px;left:22px;z-index:9999;background:white;'
        'padding:9px 12px;border:1px solid #B9C4C0;border-radius:8px;font-size:12px;'
        'line-height:1.7;box-shadow:0 1px 4px rgba(0,0,0,.15)">'
        '<b>凡例</b><br>'
        '<span style="color:#d63e3e">★</span> 候補地&nbsp;&nbsp;'
        '<span style="color:#0F766E">◯</span> 商圏<br>'
        '<span style="color:#3a87d6">●</span> 診療所&nbsp;'
        '<span style="color:#1b3a7a">●</span> 病院&nbsp;'
        '<span style="color:#9b3bd6">●</span> 美容<br>'
        '<span style="color:#2ca02c">●</span> 薬局(面)&nbsp;'
        '<span style="color:#e8820e">●</span> 薬局(門前)</div>')
    m.get_root().html.add_child(folium.Element(legend))
    return m


# ════════════════════════════════ 数式入りExcel ════════════════════════════════
_HDR = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="0F766E")
_HDR_FILL2 = PatternFill("solid", fgColor="B45309")
_INP_FILL = PatternFill("solid", fgColor="FFF7E0")   # 編集できる入力＝薄い黄色
_CALC_FILL = PatternFill("solid", fgColor="EEF2F1")  # 自動計算＝薄いグレー
_BOLD = Font(bold=True)


def _sheet_name(label):
    return re.sub(r"[\\/*?:\[\]]", "_", str(label))[:20]


def _build_footfall_sheet(wb, r):
    fp = r["fp"]
    ws = wb.create_sheet(f"集客_{_sheet_name(r['label'])}")
    ws["A1"] = f"② 集客ベース（来店客数）  {r['label']}  {r['name']}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "【黄色のセルは編集できます。編集すると下の「獲得」が自動で再計算されます】"
    ws["A2"].font = Font(italic=True, size=9, color="B45309")
    # 入力（B3..B12）※B5=有効客数は自動計算
    ws.cell(row=3, column=1, value="館の来店客数（月間）")
    ws.cell(row=3, column=2, value=r["uni"]).fill = _INP_FILL
    ws.cell(row=4, column=1, value="周知率（接触率）")
    ws.cell(row=4, column=2, value=r["exposure"]).fill = _INP_FILL
    ws.cell(row=5, column=1, value="有効客数 ＝ 来店客数 × 周知率")
    ec = ws.cell(row=5, column=2, value="=B3*B4")
    ec.fill = _CALC_FILL
    ec.font = _BOLD
    inputs = [
        ("65歳以上の比率", fp.ratio_65plus), ("65+ 月受診回数", fp.visits_month_65plus),
        ("65- 月受診回数", fp.visits_month_under65), ("処方箋発行率", fp.issue_rate),
        ("院外処方率", fp.external_rate), ("当該薬局利用率", fp.use_rate),
        ("面競合の距離減衰λ(m)", fp.competitor_decay_m),
    ]  # B6..B12
    for k, (lab, val) in enumerate(inputs):
        rr = 6 + k
        ws.cell(row=rr, column=1, value=lab)
        ws.cell(row=rr, column=2, value=val).fill = _INP_FILL
    computed = [
        ("年間受診延べ(回)", "=(B5*B6*B7+B5*(1-B6)*B8)*12"),
        ("院外処方プール(枚)", "=B14*B9*B10"),
        ("面競合の実効パワー", "=SUM(E22:E500)"),
        ("シェア", "=B11/(1+B16)"),
        ("獲得（年間・枚）", "=B15*B17"),
        ("獲得（月間・枚）", "=B18/12"),
    ]  # B14..B19
    for k, (lab, f) in enumerate(computed):
        rr = 14 + k
        ws.cell(row=rr, column=1, value=lab)
        c = ws.cell(row=rr, column=2, value=f)
        c.fill = _CALC_FILL
        if rr in (18, 19):
            c.font = _BOLD
    for j, htxt in enumerate(["競合薬局名", "候補地から(m)", "実績(枚)", "面=1/門前=0", "重み(自動)"], start=1):
        c = ws.cell(row=21, column=j, value=htxt)
        c.font = _HDR
        c.fill = _HDR_FILL
    for k, cl in enumerate(r["classified"]):
        rr = 22 + k
        eff = r["override"].get(cl["key"], cl["auto_menkata"])
        ws.cell(row=rr, column=1, value=cl["name"])
        ws.cell(row=rr, column=2, value=round(cl["d_cand"])).fill = _INP_FILL
        ws.cell(row=rr, column=3, value=int(cl["rx"]) if cl["rx"] else 0).fill = _INP_FILL
        ws.cell(row=rr, column=4, value=(1 if eff else 0)).fill = _INP_FILL
        ws.cell(row=rr, column=5,
                value=f"=IF(D{rr}=1,IF(C{rr}>0,C{rr}/12000,1)*EXP(-B{rr}/$B$12),0)").fill = _CALC_FILL
    for col, w in zip("ABCDE", [30, 14, 12, 14, 14]):
        ws.column_dimensions[col].width = w


def _build_medical_sheet(wb, r):
    hp = r["hp"]
    ws = wb.create_sheet(f"医療機関_{_sheet_name(r['label'])}")
    ws["A1"] = f"① 医療機関ベース（ハフ競合按分）  {r['label']}  {r['name']}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "【黄色=編集可。取り分率・獲得は自動再計算。『競合の重み合計』はアプリ計算値（λ変更は自店側のみ反映）】"
    ws["A2"].font = Font(italic=True, size=9, color="B45309")
    for k, (lab, val) in enumerate([("距離減衰λ(m)", hp.lambda_m), ("門前ブースト", hp.monzen_boost),
                                    ("候補店の引力", hp.candidate_attractiveness)]):
        ws.cell(row=4 + k, column=1, value=lab)
        ws.cell(row=4 + k, column=2, value=val)
    heads = ["クリニック名", "距離(m)", "年間院外処方(原資)", "自店の重み", "競合の重み合計",
             "取り分率", "獲得(枚/年)", "診療科", "発行率"]
    for j, htxt in enumerate(heads, start=1):
        c = ws.cell(row=9, column=j, value=htxt)
        c.font = _HDR
        c.fill = _HDR_FILL2
    rr = 10
    for row in r["huff_rows"]:
        ws.cell(row=rr, column=1, value=row["clinic"])
        ws.cell(row=rr, column=2, value=round(row["dist"]))
        ws.cell(row=rr, column=3, value=round(row["pool"])).fill = _INP_FILL
        ws.cell(row=rr, column=4, value=round(row["self_w"], 4)).fill = _INP_FILL
        ws.cell(row=rr, column=5, value=round(row["comp_w"], 4)).fill = _INP_FILL
        ws.cell(row=rr, column=6, value=f"=IF((D{rr}+E{rr})>0,D{rr}/(D{rr}+E{rr}),0)").fill = _CALC_FILL
        ws.cell(row=rr, column=7, value=f"=C{rr}*F{rr}").fill = _CALC_FILL
        ws.cell(row=rr, column=8, value=row.get("dept", "その他"))
        ws.cell(row=rr, column=9, value=round(row.get("issue", 0.0), 3))
        rr += 1
    ws.cell(row=rr, column=6, value="合計（＝①予測）").font = _BOLD
    tot = ws.cell(row=rr, column=7, value=f"=SUM(G10:G{rr-1})")
    tot.font = _BOLD
    ws["A7"] = "ハフ按分による予測（年間・枚）"
    ws["B7"] = f"=G{rr}"
    ws["B7"].font = _BOLD
    ws["C7"] = "（月間）"
    ws["D7"] = f"=B7/12"
    for col, w in zip("ABCDEFGHI", [28, 10, 18, 12, 14, 10, 14, 12, 9]):
        ws.column_dimensions[col].width = w


def _build_summary_sheet(ws, results):
    """お客様提示用のサマリー（比較表）。"""
    ink = "16211E"
    teal = "0F766E"
    thin = Side(style="thin", color="D7DEDB")
    med_thin = Side(style="thin", color="B9C4C0")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ncols = 8
    last_col = get_column_letter(ncols)
    # タイトル帯
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = "処方箋獲得予測  ─  出店候補地の比較"
    t.font = Font(bold=True, size=18, color=teal)
    t.alignment = left
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A2:{last_col}2")
    s = ws["A2"]
    s.value = "① 医療機関ベース（ハフ競合按分）  と  ② 集客ベース（来店客数）  の2つの独立した推計で比較"
    s.font = Font(size=10, color="5B6662")
    s.alignment = left
    ws.row_dimensions[2].height = 18

    # 2段ヘッダー（グループ見出し＋小見出し）
    hrow1, hrow2 = 4, 5
    groups = [("", 1), ("", 1), ("", 1), ("① 医療機関ベース", 2), ("② 集客ベース", 2), ("", 1)]
    col = 1
    for title, span in groups:
        if title and span > 1:
            ws.merge_cells(start_row=hrow1, start_column=col, end_row=hrow1, end_column=col + span - 1)
            cc = ws.cell(row=hrow1, column=col, value=title)
            cc.font = Font(bold=True, color="FFFFFF", size=10)
            cc.fill = PatternFill("solid", fgColor=teal)
            cc.alignment = center
            for k in range(span):
                ws.cell(row=hrow1, column=col + k).fill = PatternFill("solid", fgColor=teal)
        col += span
    subheads = ["ラベル", "店舗名 / メモ", "住所", "年間(枚)", "月間(枚)", "年間(枚)", "月間(枚)", "予測レンジ(年)"]
    for j, htxt in enumerate(subheads, start=1):
        cc = ws.cell(row=hrow2, column=j, value=htxt)
        cc.font = Font(bold=True, color="FFFFFF", size=10)
        cc.fill = PatternFill("solid", fgColor="16897E")
        cc.alignment = center
        cc.border = Border(left=thin, right=thin, top=thin, bottom=med_thin)

    # 勝者（集客 or 医療機関が最大）をハイライト
    def keyval(r):
        return r["foot_total"] if r["foot_total"] is not None else (r["med_total"] or 0)
    best_label = max(results, key=keyval)["label"] if results else None

    r0 = hrow2 + 1
    for i, r in enumerate(results):
        row = r0 + i
        ff = f"集客_{_sheet_name(r['label'])}"
        md = f"医療機関_{_sheet_name(r['label'])}"
        vals = {
            1: r["label"], 2: r["name"], 3: r["addr"],
            4: f"='{md}'!B7", 5: f"='{md}'!D7",
        }
        if r["foot_total"] is not None:
            vals[6] = f"='{ff}'!B18"
            vals[7] = f"='{ff}'!B19"
        vals[8] = (f'=IF(AND(ISNUMBER(D{row}),ISNUMBER(F{row})),'
                   f'TEXT(MIN(D{row},F{row}),"#,##0")&"〜"&TEXT(MAX(D{row},F{row}),"#,##0"),"—")')
        is_best = (r["label"] == best_label)
        base_fill = PatternFill("solid", fgColor="E8F3F1") if is_best else (
            PatternFill("solid", fgColor="F6F8F7") if i % 2 else None)
        for j in range(1, ncols + 1):
            cell = ws.cell(row=row, column=j, value=vals.get(j))
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if base_fill:
                cell.fill = base_fill
            if j == 1:
                cell.font = Font(bold=True, size=11, color=teal)
                cell.alignment = center
            elif j in (4, 5, 6, 7):
                cell.number_format = "#,##0"
                cell.alignment = right
                cell.font = Font(size=11, color=ink)
            elif j == 8:
                cell.alignment = center
            else:
                cell.alignment = left
        ws.row_dimensions[row].height = 22

    note = r0 + len(results) + 1
    ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=ncols)
    n = ws.cell(row=note, column=1,
                value="※ ①と②は同じ枚数を別データから見積もった2つの推計です（足し算しません）。"
                      "詳細な内訳は各「集客_」「医療機関_」シートを参照。")
    n.font = Font(italic=True, size=9, color="7A8481")
    ws.merge_cells(start_row=note + 1, start_column=1, end_row=note + 1, end_column=ncols)
    ws.cell(row=note + 1, column=1,
            value="※ 緑の行は①②の予測が最大の候補地。").font = Font(italic=True, size=9, color="7A8481")

    for col, w in zip("ABCDEFGH", [8, 24, 34, 13, 12, 13, 12, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False


def build_excel(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "比較サマリー"
    _build_summary_sheet(ws, results)
    for r in results:
        _build_footfall_sheet(wb, r)
        _build_medical_sheet(wb, r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════ メイン ════════════════════════════════
st.title("🏪 薬局 出店候補地 分析ツール")

st.markdown("#### 1. 候補地を入力")
st.caption("候補地ごとに ラベル・店舗名/メモ・住所・月間ユニーク客数（集客ベース用）を入力してください。行は追加できます。")

if "cand_df" not in st.session_state:
    st.session_state["cand_df"] = pd.DataFrame([
        {"ラベル": "A", "店舗名/メモ": "", "住所": "", "月間ユニーク客数": 0, "周知率": 1.0},
        {"ラベル": "B", "店舗名/メモ": "", "住所": "", "月間ユニーク客数": 0, "周知率": 1.0},
        {"ラベル": "C", "店舗名/メモ": "", "住所": "", "月間ユニーク客数": 0, "周知率": 1.0},
    ])

cand_edited = st.data_editor(
    st.session_state["cand_df"], num_rows="dynamic", use_container_width=True, key="cand_editor",
    column_config={
        "ラベル": st.column_config.TextColumn("ラベル", width="small"),
        "店舗名/メモ": st.column_config.TextColumn("店舗名/メモ"),
        "住所": st.column_config.TextColumn("住所", width="large"),
        "月間ユニーク客数": st.column_config.NumberColumn("月間ユニーク客数", min_value=0, step=500),
        "周知率": st.column_config.NumberColumn(
            "周知率", min_value=0.0, max_value=1.0, step=0.05, format="%.2f",
            help="館の来店客のうち、その薬局に接触・到達する割合。"
                 "食品スーパー=1.0／大型モール1階・主動線沿い=0.3／上層階・動線外=0.1。後から候補地ごとに変更可。",
        ),
    },
)
st.caption("💡 **周知率**：食品スーパー=**1.0**（既定）／大型モール1階・主動線=**0.3**／上層階・動線外=**0.1**。"
           "館の来店客数が大きくても、薬局に接触する割合を掛けて過大予測を防ぎます。")

col_run, col_clear = st.columns([3, 1])
run = col_run.button("▶ 未分析の候補地を分析（続きから）", type="primary", use_container_width=True)
if col_clear.button("🗑 結果をクリア", use_container_width=True):
    for k in ("mk_multi", "exp_multi", "med_edit", "ph_edit"):
        st.session_state[k] = {}
    st.session_state["multi_raw"] = []
    st.rerun()
st.caption("※ 分析は**1店ずつ順に処理**します。もう一度押せば**未分析の店だけ続きから**処理します。"
           "**1店ずつ分析 →「4. 数式入りExcel」でダウンロードして管理**するのがおすすめです"
           "（3店同時＋全件取得ONは長時間になり中断されやすいため）。")

if run:
    scraper = get_scraper()
    assumptions = PredictionAssumptions()
    targets = [row for _, row in cand_edited.iterrows() if str(row.get("住所", "")).strip()]
    if not targets:
        st.warning("住所を1件以上入力してください。")
        st.stop()
    existing = st.session_state.setdefault("multi_raw", [])
    done_labels = {r["label"] for r in existing}
    todo = []
    for i, row in enumerate(targets):
        label = str(row.get("ラベル") or f"#{i+1}").strip()
        if label not in done_labels:
            todo.append((label, row))
    if not todo:
        st.info("入力中の候補地はすべて分析済みです。やり直すには『🗑 結果をクリア』を押してください。")
    else:
        st.info(f"未分析 {len(todo)}件を順に分析します（1件あたり数分。全件取得ONだと更に時間がかかります）。")
        overall = st.progress(0.0, text="開始…")
        for i, (label, row) in enumerate(todo):
            addr = str(row["住所"]).strip()
            uni = float(row.get("月間ユニーク客数") or 0)
            overall.progress(i / len(todo), text=f"[{label}] {addr} を分析中… ({i+1}/{len(todo)})")
            log = []
            prog = st.progress(0.0, text=f"[{label}] 収集中…")
            try:
                med, ph, clat, clon = run_analysis(
                    addr, int(radius_m), gate_m, int(max_detail), log, prog,
                    assumptions=assumptions, polygons=[], exclude_outside_med=True,
                )
                prog.empty()
            except Exception as e:
                prog.empty()
                st.error(f"[{label}] 分析に失敗: {e}")
                continue
            if clat is None:
                st.error(f"[{label}] 住所の座標が取得できませんでした：{addr}")
                continue
            exposure = float(row.get("周知率") if row.get("周知率") is not None else 1.0)
            # ★完了と同時に保存（中断されても、ここまで完了した店は残る）
            st.session_state["multi_raw"].append({
                "label": label, "name": str(row.get("店舗名/メモ") or ""),
                "addr": addr, "uni": uni, "exposure": exposure,
                "clat": clat, "clon": clon, "med": med, "ph": ph})
        overall.progress(1.0, text="完了")
        overall.empty()
        st.rerun()


# ── 結果の比較表示（毎回、現在の設定＋手修正で再計算） ──────────────────────────
# ── 編集フォームは st.fragment で隔離：行追加・修正の間はアプリ全体を再実行しない ─────
# （data_editorに確定行を毎回戻すと追加行が重複増殖して重くなるため、編集中は戻さず、
#  「反映」ボタン押下時のみ確定＝session_stateへ保存し、ウィジェット状態をクリアして全体再計算）
@st.fragment
def render_med_editor(sel, c, med_high_thr_v):
    dr = get_dept_rates()
    med_flags = [clinic_flag(f, med_high_thr_v) for f in c["med"]]
    n_alert = sum(1 for x in med_flags if x)
    if n_alert:
        st.warning(f"⚠️ 医療機関に **{n_alert}件** の要確認あり（外れ値/外来不明）。『検証』列を確認し、"
                   "外来数・院外区分・座標を必要に応じて修正してください。")
    with st.expander("🔧 医療機関の確認・修正（座標／外来数／院外区分／診療科／漏れの追加・削除 → ①に反映）", expanded=bool(n_alert)):
        st.caption(
            "『距離(m)』が実態と違う施設は座標がズレています（緯度経度を直すのが最も正確／距離を直接入力も可）。"
            "行を追加・修正・削除したら、最後に **「反映して再計算」** を押すと①予測・マップに反映されます"
            "（編集中は全体を再計算しないので、続けて何行でもサクサク追加できます）。"
        )
        med_edit = st.session_state.setdefault("med_edit", {})
        recs = med_edit.get(sel, [])
        fmap = {facility_key(f): fl for f, fl in zip(c["med"], med_flags)}
        sched_map = {facility_key(f): parse_open_schedule(getattr(f, "raw_fields", None))
                     for f in c["med"]}

        def _disp_dept(r):
            d = r.get("dept")
            return d if d in DEPT_OPTIONS else "その他"

        disp = pd.DataFrame([{
            "医療機関": r.get("name"),
            "距離(m)": (round(haversine(c["clat"], c["clon"], _num(r.get("lat")), _num(r.get("lon"))))
                       if (_num(r.get("lat")) is not None and _num(r.get("lon")) is not None) else None),
            "緯度": _num(r.get("lat")), "経度": _num(r.get("lon")),
            "外来(人/日)": r.get("op"),
            "院外区分": (r.get("cat") if r.get("cat") in _EXT_CATS else "不明"),
            "診療科": _disp_dept(r),
            "発行率": eff_issue_rate(_disp_dept(r), _num(r.get("issue")), dr),
            "診療日": sched_map.get(r.get("_key"), ("", ""))[0],
            "診療時間": sched_map.get(r.get("_key"), ("", ""))[1],
            "検証": fmap.get(r.get("_key"), (
                "外来不明→既定使用" if _num(r.get("op")) is None else
                (f"要確認：{int(_num(r.get('op')))}人/日" if _num(r.get("op")) >= med_high_thr_v else ""))),
            "_key": r.get("_key"),
        } for r in recs])
        ed = st.data_editor(
            disp, hide_index=True, use_container_width=True, num_rows="dynamic",
            key=f"med_edit_{sel}", disabled=["検証", "_key", "診療日", "診療時間"],
            column_config={
                "緯度": st.column_config.NumberColumn("緯度", format="%.6f"),
                "経度": st.column_config.NumberColumn("経度", format="%.6f"),
                "距離(m)": st.column_config.NumberColumn("距離(m)", help="正しい距離を直接入力しても補正できます（向き保持）。最も正確なのは緯度経度。"),
                "外来(人/日)": st.column_config.NumberColumn("外来(人/日)", min_value=0, step=1),
                "院外区分": st.column_config.SelectboxColumn("院外区分", options=_EXT_CATS, width="medium"),
                "診療日": st.column_config.TextColumn("診療日", help="ナビィの診療時間表から抽出（開いている曜日）。"),
                "診療時間": st.column_config.TextColumn("診療時間", help="ナビィの診療時間表から抽出（代表的な時間帯）。"),
                "診療科": st.column_config.SelectboxColumn(
                    "診療科", options=DEPT_OPTIONS, width="small",
                    help="標榜診療科から自動判定。変えると発行率がその科の初期値になります。"),
                "発行率": st.column_config.NumberColumn(
                    "発行率", min_value=0.0, max_value=1.0, step=0.01, format="%.2f",
                    help="この施設だけ発行率を手入力で上書きできます（受診のうち投薬に至る割合）。"),
            },
        )
        if st.button("🔄 医療機関の修正を反映（再計算）", key=f"apply_med_{sel}", type="primary"):
            new_recs = resolve_edit(ed, "医療機関", "外来(人/日)", "op", recs, c["clat"], c["clon"],
                                    cat_col="院外区分", dept_col="診療科", issue_col="発行率")
            old_map = {r.get("_key"): r for r in recs if r.get("_key")}
            for nr in new_recs:
                disp_iss = nr.pop("issue_disp", None)
                old = old_map.get(nr.get("_key"))
                old_dept = (old or {}).get("dept")
                old_issue = (old or {}).get("issue")
                old_eff = eff_issue_rate(
                    old_dept if old_dept in dr else (nr.get("dept") or "その他"), old_issue, dr)
                ndept = nr.get("dept")
                if ndept is not None and ndept != old_dept:
                    nr["issue"] = None                 # 診療科を変更→その科の初期値に従う
                    nr["dept"] = ndept
                elif disp_iss is not None and abs(disp_iss - old_eff) > 1e-6:
                    nr["issue"] = disp_iss             # 発行率を手入力→この施設だけ上書き
                    nr["dept"] = old_dept or ndept or "その他"
                else:
                    nr["issue"] = old_issue
                    nr["dept"] = old_dept or ndept or "その他"
            med_edit[sel] = new_recs
            st.session_state.pop(f"med_edit_{sel}", None)   # ウィジェット状態をクリア＝追加行の重複増殖を防止
            st.rerun(scope="app")


@st.fragment
def render_ph_editor(sel, c):
    with st.expander("🔧 薬局の確認・修正（座標／実績／面・門前／漏れの追加・削除 → ①②に反映）", expanded=False):
        st.caption(
            "座標(緯度・経度)がズレている薬局は正しい値に直し（距離(m)直接入力も可）、漏れている薬局は行追加、"
            "誤検出は行削除。『面/門前』もここで修正できます。行を追加・修正したら、最後に "
            "**「反映して再計算」** を押すと①②に反映されます（編集中は全体を再計算しません）。"
        )
        ph_edit = st.session_state.setdefault("ph_edit", {})
        mk_all = st.session_state.setdefault("mk_multi", {})
        mk = mk_all.setdefault(sel, {})
        precs = ph_edit.get(sel, [])
        clmap = {r["key"]: r for r in c["classified"]}
        psched_map = {pharmacy_key(p): parse_open_schedule(getattr(p, "raw_fields", None))
                      for p in c["ph"]}
        pdisp = pd.DataFrame([{
            "薬局": r.get("name"),
            "距離(m)": (round(haversine(c["clat"], c["clon"], _num(r.get("lat")), _num(r.get("lon"))))
                       if (_num(r.get("lat")) is not None and _num(r.get("lon")) is not None) else None),
            "緯度": _num(r.get("lat")), "経度": _num(r.get("lon")),
            "実績(枚/年)": r.get("rx"),
            "最寄りクリニック(m)": (round(clmap[r.get("_key")]["nearest_clinic"])
                                if (r.get("_key") in clmap and clmap[r.get("_key")]["nearest_clinic"] < 1e8) else None),
            "面/門前": ("面" if mk.get(r.get("_key"),
                       (clmap[r.get("_key")]["auto_menkata"] if r.get("_key") in clmap else True)) else "門前"),
            "開局日": psched_map.get(r.get("_key"), ("", ""))[0],
            "開局時間": psched_map.get(r.get("_key"), ("", ""))[1],
            "_key": r.get("_key"),
        } for r in precs])
        ped = st.data_editor(
            pdisp, hide_index=True, use_container_width=True, num_rows="dynamic",
            key=f"ph_edit_{sel}",
            disabled=["距離(m)", "最寄りクリニック(m)", "_key", "開局日", "開局時間"],
            column_config={
                "緯度": st.column_config.NumberColumn("緯度", format="%.6f"),
                "経度": st.column_config.NumberColumn("経度", format="%.6f"),
                "距離(m)": st.column_config.NumberColumn("距離(m)", help="正しい距離を直接入力しても補正できます（向き保持）。最も正確なのは緯度経度。"),
                "実績(枚/年)": st.column_config.NumberColumn("実績(枚/年)", min_value=0, step=100),
                "面/門前": st.column_config.SelectboxColumn("面/門前", options=["面", "門前"], width="small",
                                                       help="面＝集客の競合に数える／門前＝競合から外す"),
                "開局日": st.column_config.TextColumn("開局日", help="ナビィの開局時間表から抽出（開いている曜日）。"),
                "開局時間": st.column_config.TextColumn("開局時間", help="ナビィの開局時間表から抽出（代表的な時間帯）。"),
            },
        )
        if st.button("🔄 薬局の修正を反映（再計算）", key=f"apply_ph_{sel}", type="primary"):
            pnew = resolve_edit(ped, "薬局", "実績(枚/年)", "rx", precs, c["clat"], c["clon"])
            auto_men = {r["key"]: r["auto_menkata"] for r in c["classified"]}
            new_mk = {}
            for _, row in ped.iterrows():
                k = row.get("_key")
                if not (isinstance(k, str) and k):
                    continue
                is_men = (row.get("面/門前") == "面")
                if k in auto_men and is_men != auto_men[k]:
                    new_mk[k] = is_men
            ph_edit[sel] = pnew
            mk_all[sel] = new_mk
            st.session_state.pop(f"ph_edit_{sel}", None)    # ウィジェット状態をクリア＝追加行の重複増殖を防止
            st.rerun(scope="app")


raws = st.session_state.get("multi_raw", [])
if raws:
    # 確定済みの手修正（session_state）を反映して計算（編集フォームは fragment で隔離）。
    computed = [compute_candidate(r) for r in raws]

    st.markdown("#### 2. 比較結果")
    rows = []
    for cc in computed:
        med, foot = cc["med_total"], cc["foot_total"]
        vals = [v for v in (med, foot) if v is not None]
        rng = f"{min(vals):,.0f}〜{max(vals):,.0f}" if len(vals) == 2 else "—"
        rows.append({
            "ラベル": cc["label"], "店舗名/メモ": cc["name"], "住所": cc["addr"][:24],
            "① 医療機関(年)": round(med) if med is not None else None,
            "① 医療機関(月)": round(med / 12) if med is not None else None,
            "② 集客(年)": round(foot) if foot is not None else None,
            "② 集客(月)": round(foot / 12) if foot is not None else None,
            "予測レンジ(年)": rng, "面競合数": cc["comp_n"], "寄与医療機関数": len(cc["huff_rows"]),
        })
    st.dataframe(
        pd.DataFrame(rows), hide_index=True, use_container_width=True,
        column_config={cn: st.column_config.NumberColumn(cn, format="%d 枚")
                       for cn in ["① 医療機関(年)", "① 医療機関(月)", "② 集客(年)", "② 集客(月)"]},
    )
    ranked = sorted(computed, key=lambda x: (x["foot_total"] or x["med_total"] or 0), reverse=True)
    st.success(f"🏆 最大の候補地： **{ranked[0]['label']}**"
               f"（{ranked[0]['name'] or ranked[0]['addr'][:20]}）")

    st.markdown("#### 3. 候補地ごとの詳細（ロジックの内訳）")
    sel = st.selectbox("詳細を見る候補地", [c["label"] for c in computed],
                       format_func=lambda x: f"{x}　" + next((c["name"] or c["addr"][:16]
                                                              for c in computed if c["label"] == x), ""))
    c = next(x for x in computed if x["label"] == sel)

    m1, m2, m3 = st.columns(3)
    m1.metric("① 医療機関ベース（ハフ）", f"{c['med_total']:,.0f} 枚/年",
              f"月 {c['med_total']/12:,.0f} 枚")
    if c["foot_total"] is not None:
        m2.metric("② 集客ベース", f"{c['foot_total']:,.0f} 枚/年", f"月 {c['foot_total']/12:,.0f} 枚")
    else:
        m2.metric("② 集客ベース", "未入力")
    m3.metric("面競合数 / 除外", f"{c['comp_n']} / {c['comp_excluded']}")

    st.markdown("##### 🗺 商圏マップ（周辺の医療機関・薬局）")
    show_map = st.checkbox(
        "地図を表示する（⚠️ 重い処理。**施設の追加・修正が終わってから**オンにしてスクショ推奨）",
        value=False, key=f"showmap_{sel}",
        help="地図の再描画は重いので、編集中はオフのままが快適です。オフでも①②の数字・下の内訳表は反映済みです。",
    )
    if show_map:
        st.caption("このままスクショしてお客様提示にお使いください。手動で追加・削除・座標修正した施設もそのまま反映されます。"
                   "★＝候補地／◯＝商圏／青＝診療所・濃青＝病院・紫＝美容／緑＝薬局(面)・橙＝薬局(門前)。"
                   "マーカーをクリックすると距離・診療科・外来・実績などが表示されます。")
        st_folium(build_map(c, radius_m), height=540, returned_objects=[],
                  use_container_width=True, key=f"map_{sel}")
    else:
        st.caption("地図はオフ中（編集がサクサク動きます）。編集が終わったら上のチェックをオンにして地図を表示・スクショしてください。")

    st.markdown("##### ① 医療機関ベース：ハフの取り分内訳")
    st.caption("各クリニックが出す処方箋（原資）を、自店の重み ÷（自店の重み＋競合の重み合計）の"
               "取り分率で獲得します。合計＝①予測。")
    hb_df = pd.DataFrame([{
        "医療機関": row["clinic"], "診療科": row.get("dept", "その他"),
        "発行率": round(row.get("issue", 0.0), 2), "距離(m)": round(row["dist"]),
        "年間院外処方(原資)": round(row["pool"]),
        "自店の重み": round(row["self_w"], 3), "競合の重み合計": round(row["comp_w"], 3),
        "取り分率": round(row["share"], 3), "獲得(枚/年)": round(row["captured"]),
    } for row in c["huff_rows"]])
    st.dataframe(hb_df, hide_index=True, use_container_width=True, column_config={
        "発行率": st.column_config.NumberColumn("発行率", format="%.2f"),
        "取り分率": st.column_config.NumberColumn("取り分率", format="%.3f"),
        "獲得(枚/年)": st.column_config.NumberColumn("獲得(枚/年)", format="%d 枚"),
    })

    if c["foot"]:
        st.markdown("##### ② 集客ベース：内訳")
        exp_all = st.session_state.setdefault("exp_multi", {})
        new_exp = st.number_input(
            f"周知率（{sel}）— 館の来店客のうち薬局に接触する割合", 0.0, 1.0,
            float(c["exposure"]), 0.05, format="%.2f", key=f"exp_{sel}",
            help="食品スーパー=1.0／大型モール1階・主動線=0.3／上層階・動線外=0.1。変更すると②が再計算されます。",
        )
        if abs(new_exp - c["exposure"]) > 1e-9:
            exp_all[sel] = new_exp
            st.rerun()
        fo, fp = c["foot"], c["fp"]
        st.markdown(
            f"- 館の来店客数 {c['uni']:,.0f}人 × **周知率 {c['exposure']:.2f}** = 有効客数 **{c['eff_uni']:,.0f}人**"
            f"（65+ {fo['u65']:,.0f} / 65− {fo['u_under']:,.0f}）\n"
            f"- 年間受診延べ {fo['annual_visits']:,.0f}回 → 院外処方プール {fo['rx_pool']:,.0f}枚\n"
            f"- 利用率 {fp.use_rate:.1%} ÷ (面競合の実効パワー {c['comp_power']:.1f}"
            f"〔面{c['comp_n']}店・距離減衰λ={fp.competitor_decay_m:.0f}m〕 + 1) = シェア {fo['share']:.2%}\n"
            f"- **獲得 = {fo['total']:,.0f} 枚/年**"
        )

    # ── 確認・修正フォーム（fragment：行追加してもアプリ全体は再実行しない） ──────────
    st.markdown("##### 🔧 確認・修正（漏れの追加・座標・面/門前・診療科・外来など）")
    st.caption("💡 施設の追加・修正は**何行でも続けて**行えます（サクサク動きます）。"
               "終わったら各フォームの **「反映して再計算」** を押すと、上の①②予測とマップに反映されます。")
    render_med_editor(sel, c, int(med_high_thr))
    render_ph_editor(sel, c)

    # ── 4. Excel ─────────────────────────────────────────────────────────────
    st.markdown("#### 4. 数式入りExcelで書き出し")
    st.caption("集客シートは客数・競合・係数、医療機関シートは原資・重みを編集すると、"
               "獲得枚数がExcelの数式で自動再計算されます（ブラウザを閉じても手元で編集可能）。")
    st.download_button(
        "📊 比較Excel（数式入り）をダウンロード",
        data=build_excel(computed), file_name="処方箋予測_複数店舗比較.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
